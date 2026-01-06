"""
Prima Nota Salari - Nuovo sistema di gestione stipendi.

Struttura dati:
- Dipendente
- Mese di competenza
- Anno
- Importo Busta (da file paghe.xlsx)
- Importo Bonifico (da file bonifici.xlsx)
- Saldo (differenza)
- Progressivo (riporto da mesi precedenti)
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from typing import Dict, Any, List, Optional
from datetime import datetime
import uuid
import logging
import io

from app.database import Database

logger = logging.getLogger(__name__)
router = APIRouter()

# Mapping mesi italiano -> numero
MESI_MAP = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12
}

MESI_NOMI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
             "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]


def normalize_name(name: str) -> str:
    """Normalizza nome dipendente per matching."""
    if not name:
        return ""
    return " ".join(name.strip().upper().split())


def get_mese_numero(mese_str: str) -> int:
    """Converte nome mese in numero."""
    if not mese_str:
        return 0
    return MESI_MAP.get(mese_str.lower().strip(), 0)


# ============== ENDPOINTS ==============

@router.get("/salari")
async def get_prima_nota_salari(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    dipendente: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """
    Recupera la prima nota salari con tutti i campi:
    - dipendente, mese, anno
    - importo_busta, importo_bonifico, saldo, progressivo
    - riconciliato
    """
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    if dipendente:
        query["dipendente"] = {"$regex": dipendente, "$options": "i"}
    
    salari = await db["prima_nota_salari"].find(
        query, {"_id": 0}
    ).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)
    
    return salari


@router.get("/salari/riepilogo")
async def get_riepilogo_salari(
    anno: int = Query(...),
    mese: Optional[int] = Query(None)
) -> Dict[str, Any]:
    """Riepilogo statistiche salari."""
    db = Database.get_db()
    
    query = {"anno": anno}
    if mese:
        query["mese"] = mese
    
    salari = await db["prima_nota_salari"].find(query, {"_id": 0}).to_list(5000)
    
    totale_buste = sum(s.get("importo_busta", 0) or 0 for s in salari)
    totale_bonifici = sum(s.get("importo_bonifico", 0) or 0 for s in salari)
    totale_saldo = sum(s.get("saldo", 0) or 0 for s in salari)
    
    # Conta dipendenti unici
    dipendenti_unici = len(set(s.get("dipendente", "") for s in salari))
    
    # Conta riconciliati
    riconciliati = sum(1 for s in salari if s.get("riconciliato"))
    
    return {
        "anno": anno,
        "mese": mese,
        "totale_records": len(salari),
        "dipendenti_unici": dipendenti_unici,
        "totale_buste": round(totale_buste, 2),
        "totale_bonifici": round(totale_bonifici, 2),
        "totale_saldo": round(totale_saldo, 2),
        "riconciliati": riconciliati,
        "da_riconciliare": len(salari) - riconciliati
    }


@router.post("/import-paghe")
async def import_paghe(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa file PAGHE (stipendi netti).
    Formato atteso: Dipendente | Mese | Anno | Stipendio Netto
    
    Aggiorna/crea record nella prima_nota_salari con importo_busta.
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {str(e)}")
    
    # Normalizza nomi colonne
    df.columns = [c.strip().lower() for c in df.columns]
    
    # Mapping colonne
    col_dipendente = None
    col_mese = None
    col_anno = None
    col_importo = None
    
    for c in df.columns:
        if 'dipendente' in c or 'nome' in c or 'cognome' in c:
            col_dipendente = c
        elif 'mese' in c:
            col_mese = c
        elif 'anno' in c:
            col_anno = c
        elif any(x in c for x in ['stipendio', 'netto', 'busta', 'paga', 'retribuzione']):
            col_importo = c
    
    # Fallback: se non trova colonna specifica, cerca 'importo' (ma non 'bonifico' o 'erogato')
    if not col_importo:
        for c in df.columns:
            if 'importo' in c and 'bonifico' not in c and 'erogato' not in c:
                col_importo = c
                break
    
    if not all([col_dipendente, col_mese, col_anno, col_importo]):
        raise HTTPException(
            status_code=400, 
            detail=f"Colonne richieste non trovate. Trovate: {list(df.columns)}. Mappate: dipendente={col_dipendente}, mese={col_mese}, anno={col_anno}, importo={col_importo}"
        )
    
    logger.info(f"IMPORT PAGHE - Colonne mappate: dipendente={col_dipendente}, mese={col_mese}, anno={col_anno}, importo={col_importo}")
    
    created = 0
    updated = 0
    errors = []
    
    # Raggruppa per dipendente/mese/anno e somma gli importi
    grouped_data = {}
    
    for idx, row in df.iterrows():
        try:
            dipendente = normalize_name(str(row[col_dipendente]))
            if not dipendente or dipendente == "NAN":
                continue
            
            mese_str = str(row[col_mese]).strip()
            mese = get_mese_numero(mese_str)
            if mese == 0:
                continue
            
            anno_val = row[col_anno]
            if isinstance(anno_val, datetime):
                anno = anno_val.year
            else:
                anno = int(anno_val)
            
            importo = float(row[col_importo]) if pd.notna(row[col_importo]) else 0
            
            # Chiave univoca
            key = (dipendente, anno, mese)
            if key not in grouped_data:
                grouped_data[key] = 0
            grouped_data[key] += importo
            
        except Exception as e:
            errors.append(f"Riga {idx + 2}: {str(e)}")
    
    # Inserisci/aggiorna nel database
    for (dipendente, anno, mese), importo_busta in grouped_data.items():
        # Cerca record esistente
        existing = await db["prima_nota_salari"].find_one({
            "dipendente": dipendente,
            "anno": anno,
            "mese": mese
        })
        
        if existing:
            # Aggiorna importo_busta
            await db["prima_nota_salari"].update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "importo_busta": round(importo_busta, 2),
                    "saldo": round((existing.get("importo_bonifico") or 0) - importo_busta, 2),
                    "updated_at": datetime.utcnow().isoformat()
                }}
            )
            updated += 1
        else:
            # Crea nuovo record
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1],
                "importo_busta": round(importo_busta, 2),
                "importo_bonifico": 0,
                "saldo": round(-importo_busta, 2),  # Bonifico(0) - Busta
                "progressivo": 0,
                "riconciliato": False,
                "created_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat()
            }
            await db["prima_nota_salari"].insert_one(new_record)
            created += 1
    
    # Ricalcola progressivi per tutti i dipendenti
    await ricalcola_progressivi_tutti(db)
    
    return {
        "success": True,
        "message": f"Import PAGHE completato",
        "created": created,
        "updated": updated,
        "errors": errors[:10] if errors else []
    }


@router.post("/import-bonifici")
async def import_bonifici(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa file BONIFICI (importi erogati).
    Formato atteso: Dipendente | Mese | Anno | Importo Erogato
    
    Aggiorna/crea record nella prima_nota_salari con importo_bonifico.
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {str(e)}")
    
    # Normalizza nomi colonne
    df.columns = [c.strip().lower() for c in df.columns]
    
    # Mapping colonne
    col_dipendente = None
    col_mese = None
    col_anno = None
    col_importo = None
    
    for c in df.columns:
        if 'dipendente' in c or 'nome' in c or 'cognome' in c:
            col_dipendente = c
        elif 'mese' in c:
            col_mese = c
        elif 'anno' in c:
            col_anno = c
        elif any(x in c for x in ['erogato', 'bonifico', 'pagato', 'versato', 'accredito']):
            col_importo = c
    
    # Fallback: se non trova colonna specifica, cerca 'importo' (ma non 'stipendio' o 'netto')
    if not col_importo:
        for c in df.columns:
            if 'importo' in c and 'stipendio' not in c and 'netto' not in c and 'busta' not in c:
                col_importo = c
                break
    
    if not all([col_dipendente, col_mese, col_anno, col_importo]):
        raise HTTPException(
            status_code=400, 
            detail=f"Colonne richieste non trovate. Trovate: {list(df.columns)}"
        )
    
    created = 0
    updated = 0
    errors = []
    
    # Raggruppa per dipendente/mese/anno e somma gli importi
    grouped_data = {}
    
    for idx, row in df.iterrows():
        try:
            dipendente = normalize_name(str(row[col_dipendente]))
            if not dipendente or dipendente == "NAN":
                continue
            
            mese_str = str(row[col_mese]).strip()
            mese = get_mese_numero(mese_str)
            if mese == 0:
                continue
            
            anno_val = row[col_anno]
            if isinstance(anno_val, datetime):
                anno = anno_val.year
            else:
                anno = int(anno_val)
            
            importo = float(row[col_importo]) if pd.notna(row[col_importo]) else 0
            
            # Chiave univoca
            key = (dipendente, anno, mese)
            if key not in grouped_data:
                grouped_data[key] = 0
            grouped_data[key] += importo
            
        except Exception as e:
            errors.append(f"Riga {idx + 2}: {str(e)}")
    
    # Inserisci/aggiorna nel database
    for (dipendente, anno, mese), importo_bonifico in grouped_data.items():
        # Cerca record esistente
        existing = await db["prima_nota_salari"].find_one({
            "dipendente": dipendente,
            "anno": anno,
            "mese": mese
        })
        
        if existing:
            # Aggiorna importo_bonifico
            importo_busta = existing.get("importo_busta") or 0
            await db["prima_nota_salari"].update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "importo_bonifico": round(importo_bonifico, 2),
                    "saldo": round(importo_bonifico - importo_busta, 2),  # Bonifico - Busta
                    "updated_at": datetime.utcnow().isoformat()
                }}
            )
            updated += 1
        else:
            # Crea nuovo record (solo bonifico senza busta)
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1],
                "importo_busta": 0,
                "importo_bonifico": round(importo_bonifico, 2),
                "saldo": round(importo_bonifico, 2),  # Bonifico - Busta(0) = positivo (anticipo)
                "progressivo": 0,
                "riconciliato": False,
                "created_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat()
            }
            await db["prima_nota_salari"].insert_one(new_record)
            created += 1
    
    # Ricalcola progressivi per tutti i dipendenti
    await ricalcola_progressivi_tutti(db)
    
    return {
        "success": True,
        "message": f"Import BONIFICI completato",
        "created": created,
        "updated": updated,
        "errors": errors[:10] if errors else []
    }


async def ricalcola_progressivi_tutti(db):
    """
    Ricalcola saldi e progressivi per tutti i dipendenti.
    
    Formula Saldo: Bonifico - Busta
    - Saldo positivo = dipendente ha ricevuto più di quanto spettava (ci deve soldi)
    - Saldo negativo = dipendente ha ricevuto meno di quanto spettava (gli dobbiamo soldi)
    
    Progressivo = Σ(saldi) da inizio a mese corrente
    """
    # Ottieni tutti i dipendenti unici
    dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    
    for dipendente in dipendenti:
        # Ordina per anno e mese (dal più vecchio al più recente)
        records = await db["prima_nota_salari"].find(
            {"dipendente": dipendente}
        ).sort([("anno", 1), ("mese", 1)]).to_list(500)
        
        progressivo = 0
        
        for record in records:
            # Ricalcola saldo: Bonifico - Busta
            importo_busta = record.get("importo_busta", 0) or 0
            importo_bonifico = record.get("importo_bonifico", 0) or 0
            saldo = importo_bonifico - importo_busta
            
            # Aggiunge al progressivo
            progressivo += saldo
            
            # Aggiorna il record con saldo e progressivo
            await db["prima_nota_salari"].update_one(
                {"_id": record["_id"]},
                {"$set": {
                    "saldo": round(saldo, 2),
                    "progressivo": round(progressivo, 2)
                }}
            )


@router.post("/ricalcola-progressivi")
async def ricalcola_progressivi() -> Dict[str, str]:
    """Ricalcola tutti i progressivi dipendenti."""
    db = Database.get_db()
    await ricalcola_progressivi_tutti(db)
    return {"message": "Progressivi ricalcolati"}


@router.get("/dipendenti-lista")
async def get_dipendenti_lista() -> List[str]:
    """Lista nomi dipendenti unici dalla prima nota salari."""
    db = Database.get_db()
    dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    return sorted(dipendenti)


@router.delete("/salari/reset")
async def reset_prima_nota_salari() -> Dict[str, Any]:
    """Elimina tutti i record della prima nota salari."""
    db = Database.get_db()
    result = await db["prima_nota_salari"].delete_many({})
    return {
        "message": f"Eliminati {result.deleted_count} record",
        "deleted_count": result.deleted_count
    }


@router.delete("/salari/{record_id}")
async def delete_salario(record_id: str) -> Dict[str, str]:
    """Elimina un singolo record."""
    db = Database.get_db()
    result = await db["prima_nota_salari"].delete_one({"id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    return {"message": "Record eliminato"}


@router.put("/salari/{record_id}")
async def update_salario(
    record_id: str,
    data: Dict[str, Any] = Body(...)
) -> Dict[str, str]:
    """Aggiorna un record della prima nota salari."""
    db = Database.get_db()
    
    # Calcola saldo: Bonifico - Busta
    importo_busta = float(data.get("importo_busta", 0))
    importo_bonifico = float(data.get("importo_bonifico", 0))
    saldo = importo_bonifico - importo_busta
    
    # Aggiorna il mese_nome se cambia il mese
    mese = int(data.get("mese", 1))
    mese_nome = MESI_NOMI[mese - 1] if 1 <= mese <= 12 else ""
    
    update_data = {
        "dipendente": data.get("dipendente", ""),
        "anno": int(data.get("anno", 2025)),
        "mese": mese,
        "mese_nome": mese_nome,
        "importo_busta": importo_busta,
        "importo_bonifico": importo_bonifico,
        "saldo": round(saldo, 2),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    result = await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Record aggiornato"}


@router.put("/salari/{record_id}/riconcilia")
async def riconcilia_salario(
    record_id: str,
    riconciliato: bool = Query(True)
) -> Dict[str, str]:
    """Marca un salario come riconciliato."""
    db = Database.get_db()
    result = await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": {
            "riconciliato": riconciliato,
            "data_riconciliazione": datetime.utcnow().isoformat() if riconciliato else None,
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    return {"message": "Stato riconciliazione aggiornato"}


@router.get("/export-excel")
async def export_prima_nota_salari_excel(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None)
):
    """Esporta prima nota salari in Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    
    salari = await db["prima_nota_salari"].find(
        query, {"_id": 0}
    ).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prima Nota Salari"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    
    # Headers
    headers = ["Dipendente", "Anno", "Mese", "Importo Busta", "Importo Bonifico", "Saldo", "Progressivo", "Riconciliato"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for row_num, s in enumerate(salari, 2):
        ws.cell(row=row_num, column=1, value=s.get("dipendente", ""))
        ws.cell(row=row_num, column=2, value=s.get("anno", ""))
        ws.cell(row=row_num, column=3, value=s.get("mese_nome", ""))
        ws.cell(row=row_num, column=4, value=s.get("importo_busta", 0))
        ws.cell(row=row_num, column=5, value=s.get("importo_bonifico", 0))
        ws.cell(row=row_num, column=6, value=s.get("saldo", 0))
        ws.cell(row=row_num, column=7, value=s.get("progressivo", 0))
        ws.cell(row=row_num, column=8, value="Sì" if s.get("riconciliato") else "No")
    
    # Larghezze
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"prima_nota_salari"
    if anno:
        filename += f"_{anno}"
    if mese:
        filename += f"_{mese:02d}"
    filename += ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
