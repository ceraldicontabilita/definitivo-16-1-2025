"""
Gestione Dipendenti - Router API completo.
Anagrafica, turni, libro unico, libretti sanitari.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body
from typing import Dict, Any, List, Optional
from datetime import datetime, date, timedelta
import uuid
import logging
import io
import re

from app.database import Database, Collections

logger = logging.getLogger(__name__)
router = APIRouter()

# Costanti
TURNI_TIPI = {
    "mattina": {"label": "Mattina", "orario": "06:00 - 14:00", "color": "#4caf50"},
    "pomeriggio": {"label": "Pomeriggio", "orario": "14:00 - 22:00", "color": "#2196f3"},
    "sera": {"label": "Sera", "orario": "18:00 - 02:00", "color": "#9c27b0"},
    "full": {"label": "Full Day", "orario": "10:00 - 22:00", "color": "#ff9800"},
    "riposo": {"label": "Riposo", "orario": "-", "color": "#9e9e9e"},
    "ferie": {"label": "Ferie", "orario": "-", "color": "#e91e63"},
    "malattia": {"label": "Malattia", "orario": "-", "color": "#f44336"}
}

MANSIONI = [
    "Cameriere", "Cuoco", "Aiuto Cuoco", "Barista", "Pizzaiolo", 
    "Lavapiatti", "Cassiera", "Responsabile Sala", "Chef", "Sommelier"
]

CONTRATTI_TIPI = [
    "Tempo Indeterminato", "Tempo Determinato", "Apprendistato", 
    "Stage/Tirocinio", "Collaborazione", "Part-time"
]


@router.get("")
async def list_dipendenti(
    skip: int = Query(0, ge=0),
    limit: int = Query(1000, ge=1, le=10000),
    attivo: Optional[bool] = Query(None),
    mansione: Optional[str] = Query(None),
    search: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """Lista dipendenti con filtri."""
    db = Database.get_db()
    
    query = {}
    if attivo is not None:
        query["attivo"] = attivo
    if mansione:
        query["mansione"] = mansione
    if search:
        query["$or"] = [
            {"nome_completo": {"$regex": search, "$options": "i"}},
            {"codice_fiscale": {"$regex": search, "$options": "i"}}
        ]
    
    dipendenti = await db[Collections.EMPLOYEES].find(query, {"_id": 0}).sort("nome_completo", 1).skip(skip).limit(limit).to_list(limit)
    return dipendenti


@router.get("/stats")
async def get_dipendenti_stats() -> Dict[str, Any]:
    """Statistiche dipendenti."""
    db = Database.get_db()
    
    total = await db[Collections.EMPLOYEES].count_documents({})
    attivi = await db[Collections.EMPLOYEES].count_documents({"attivo": {"$ne": False}})
    
    # Per mansione
    pipeline = [
        {"$group": {"_id": "$mansione", "count": {"$sum": 1}}}
    ]
    by_mansione = await db[Collections.EMPLOYEES].aggregate(pipeline).to_list(100)
    
    # Libretti in scadenza (prossimi 30 giorni)
    today = datetime.utcnow()
    deadline = today + timedelta(days=30)
    libretti_scadenza = await db[Collections.EMPLOYEES].count_documents({
        "libretto_scadenza": {"$lte": deadline.isoformat()[:10], "$gte": today.isoformat()[:10]}
    })
    
    return {
        "totale": total,
        "attivi": attivi,
        "inattivi": total - attivi,
        "per_mansione": {item["_id"] or "N/D": item["count"] for item in by_mansione},
        "libretti_in_scadenza": libretti_scadenza
    }


@router.get("/tipi-turno")
async def get_tipi_turno() -> Dict[str, Any]:
    """Ritorna i tipi di turno disponibili."""
    return TURNI_TIPI


@router.get("/mansioni")
async def get_mansioni() -> List[str]:
    """Ritorna le mansioni disponibili."""
    return MANSIONI


@router.get("/tipi-contratto")
async def get_tipi_contratto() -> List[str]:
    """Ritorna i tipi di contratto disponibili."""
    return CONTRATTI_TIPI


@router.post("")
async def create_dipendente(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo dipendente."""
    db = Database.get_db()
    
    # Campi obbligatori
    required = ["nome_completo"]
    for field in required:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Campo obbligatorio mancante: {field}")
    
    # Parse nome
    nome_parts = data["nome_completo"].split()
    cognome = nome_parts[0] if nome_parts else ""
    nome = " ".join(nome_parts[1:]) if len(nome_parts) > 1 else ""
    
    dipendente = {
        "id": str(uuid.uuid4()),
        "nome_completo": data["nome_completo"],
        "cognome": cognome,
        "nome": nome,
        "codice_fiscale": data.get("codice_fiscale", ""),
        "matricola": data.get("matricola", ""),
        "email": data.get("email", ""),
        "telefono": data.get("telefono", ""),
        "indirizzo": data.get("indirizzo", ""),
        "data_nascita": data.get("data_nascita"),
        "luogo_nascita": data.get("luogo_nascita", ""),
        "mansione": data.get("mansione", ""),
        "qualifica": data.get("qualifica", data.get("mansione", "")),
        "livello": data.get("livello", ""),
        "tipo_contratto": data.get("tipo_contratto", "Tempo Indeterminato"),
        "data_assunzione": data.get("data_assunzione"),
        "data_fine_contratto": data.get("data_fine_contratto"),
        "ore_settimanali": data.get("ore_settimanali", 40),
        "iban": data.get("iban", ""),
        # Libretto sanitario
        "libretto_numero": data.get("libretto_numero", ""),
        "libretto_scadenza": data.get("libretto_scadenza"),
        "libretto_file": data.get("libretto_file"),
        # Portale
        "portale_invitato": False,
        "portale_registrato": False,
        "portale_ultimo_accesso": None,
        # Status
        "attivo": True,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    # Verifica duplicato CF
    if dipendente["codice_fiscale"]:
        existing = await db[Collections.EMPLOYEES].find_one({"codice_fiscale": dipendente["codice_fiscale"]})
        if existing:
            raise HTTPException(status_code=409, detail="Dipendente con questo codice fiscale già esistente")
    
    await db[Collections.EMPLOYEES].insert_one(dipendente)
    dipendente.pop("_id", None)
    
    return dipendente


# ============== BUSTE PAGA (must be before /{dipendente_id} to avoid route conflict) ==============

@router.get("/buste-paga")
async def get_buste_paga(
    anno: int = Query(...),
    mese: str = Query(...)
) -> List[Dict[str, Any]]:
    """
    Ottiene le buste paga per un determinato mese.
    Le buste paga vengono create automaticamente dai movimenti salari.
    """
    db = Database.get_db()
    
    periodo = f"{anno}-{mese}"
    
    # Cerca buste paga esistenti
    buste = await db["buste_paga"].find(
        {"periodo": periodo},
        {"_id": 0}
    ).to_list(1000)
    
    return buste


@router.post("/buste-paga")
async def create_busta_paga(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea o aggiorna una busta paga."""
    db = Database.get_db()
    
    required = ["dipendente_id", "periodo"]
    for field in required:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Campo {field} obbligatorio")
    
    # Cerca busta esistente
    existing = await db["buste_paga"].find_one({
        "dipendente_id": data["dipendente_id"],
        "periodo": data["periodo"]
    })
    
    busta = {
        "dipendente_id": data["dipendente_id"],
        "periodo": data["periodo"],
        "lordo": float(data.get("lordo", 0) or 0),
        "netto": float(data.get("netto", 0) or 0),
        "contributi": float(data.get("contributi", 0) or 0),
        "trattenute": float(data.get("trattenute", 0) or 0),
        "pagata": bool(data.get("pagata", False)),
        "data_pagamento": data.get("data_pagamento"),
        "note": data.get("note", ""),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    if existing:
        await db["buste_paga"].update_one(
            {"id": existing["id"]},
            {"$set": busta}
        )
        busta["id"] = existing["id"]
    else:
        busta["id"] = str(uuid.uuid4())
        busta["created_at"] = datetime.utcnow().isoformat()
        await db["buste_paga"].insert_one(busta)
    
    busta.pop("_id", None)
    return busta


# ============== SALARI (must be before /{dipendente_id} to avoid route conflict) ==============

MESI_MAP = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12
}

@router.post("/import-salari")
async def import_salari_excel(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa salari da file Excel con aggregazione per dipendente/mese.
    
    Formato atteso:
    - Colonna 1: Dipendente (nome completo)
    - Colonna 2: Mese (italiano: Gennaio, Febbraio, etc.)
    - Colonna 3: Anno
    - Colonna 4: Stipendio Netto (importo busta) - può essere su riga separata
    - Colonna 5: Importo Erogato (bonifico) - può essere su riga separata o multipli
    
    Il sistema aggrega automaticamente più righe dello stesso dipendente/mese.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installato")
    
    db = Database.get_db()
    
    # Leggi file Excel
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = wb.active
    
    # Prima passata: aggrega dati per dipendente/mese/anno
    aggregati = {}  # key: (dipendente, mese, anno) -> {stipendio_netto, importo_erogato}
    
    for row_num in range(2, sheet.max_row + 1):
        dipendente_nome = sheet.cell(row=row_num, column=1).value
        mese_str = sheet.cell(row=row_num, column=2).value
        anno_val = sheet.cell(row=row_num, column=3).value
        stipendio_netto = sheet.cell(row=row_num, column=4).value
        importo_erogato = sheet.cell(row=row_num, column=5).value
        
        if not dipendente_nome or not mese_str or not anno_val:
            continue
        
        # Gestisci anno (può essere int, float, o datetime)
        try:
            if isinstance(anno_val, datetime):
                anno = anno_val.year
            elif isinstance(anno_val, date):
                anno = anno_val.year
            else:
                anno = int(float(str(anno_val)))
            
            # Valida anno ragionevole
            if anno < 2000 or anno > 2100:
                continue
        except (ValueError, TypeError):
            continue
        
        # Converti mese
        mese_lower = str(mese_str).lower().strip()
        mese = MESI_MAP.get(mese_lower)
        if not mese:
            continue
        
        # Chiave univoca
        key = (str(dipendente_nome).strip(), mese, anno)
        
        if key not in aggregati:
            aggregati[key] = {"stipendio_netto": 0, "importo_erogato": 0}
        
        # Aggrega i valori (somma se ci sono più righe)
        if stipendio_netto and float(stipendio_netto) > 0:
            # Per stipendio netto, prendi il MAX (non sommare)
            aggregati[key]["stipendio_netto"] = max(
                aggregati[key]["stipendio_netto"],
                float(stipendio_netto)
            )
        
        if importo_erogato and float(importo_erogato) > 0:
            # Per bonifici, SOMMA (potrebbero essere pagamenti multipli)
            aggregati[key]["importo_erogato"] += float(importo_erogato)
    
    # Seconda passata: inserisci/aggiorna nel DB
    imported = 0
    updated = 0
    skipped = 0
    errors = []
    
    for (dipendente_nome, mese, anno), valori in aggregati.items():
        try:
            stipendio = valori["stipendio_netto"]
            erogato = valori["importo_erogato"]
            
            # Se entrambi sono 0, salta
            if stipendio == 0 and erogato == 0:
                skipped += 1
                continue
            
            # Se stipendio è 0, usa erogato
            if stipendio == 0:
                stipendio = erogato
            
            # Crea data per il movimento (ultimo giorno del mese)
            from calendar import monthrange
            _, last_day = monthrange(int(anno), mese)
            data_movimento = f"{anno}-{mese:02d}-{last_day:02d}"
            
            # Crea ID univoco
            movimento_id = f"SAL-{anno}-{mese:02d}-{dipendente_nome.replace(' ', '-')}"
            
            # Cerca se esiste già
            existing = await db["prima_nota_salari"].find_one({"id": movimento_id})
            
            movimento = {
                "id": movimento_id,
                "dipendente": dipendente_nome,
                "mese": mese,
                "mese_nome": list(MESI_MAP.keys())[mese-1].capitalize(),
                "anno": int(anno),
                "data": data_movimento,
                "stipendio_netto": round(stipendio, 2),
                "importo_erogato": round(erogato, 2) if erogato > 0 else round(stipendio, 2),
                "importo": round(erogato, 2) if erogato > 0 else round(stipendio, 2),
                "tipo": "uscita",
                "categoria": "SALARIO",
                "descrizione": f"Stipendio {list(MESI_MAP.keys())[mese-1].capitalize()} {anno} - {dipendente_nome}",
                "imported": True
            }
            
            if existing:
                # Aggiorna
                await db["prima_nota_salari"].update_one(
                    {"id": movimento_id},
                    {"$set": movimento}
                )
                updated += 1
            else:
                movimento["created_at"] = datetime.utcnow().isoformat()
                await db["prima_nota_salari"].insert_one(movimento)
                imported += 1
            
        except Exception as e:
            errors.append(f"{dipendente_nome} {mese}/{anno}: {str(e)}")
            skipped += 1
    
    return {
        "message": "Importazione completata",
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20] if errors else [],
        "total_rows": sheet.max_row - 1,
        "aggregati": len(aggregati)
    }


@router.post("/import-estratto-conto")
async def import_estratto_conto(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa estratto conto bancario (CSV o Excel) e riconcilia con i salari.
    
    Formati supportati:
    - CSV (separatore ;): Data, Importo, Descrizione, Categoria
    - Excel (.xlsx, .xls): Colonne simili
    
    Nota: PDF non più supportato per attendibilità dati.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    contents = await file.read()
    
    movimenti_banca = []
    
    if filename.endswith('.csv'):
        # Parse CSV
        import csv
        text = contents.decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
        
        for row in reader:
            # Filtra solo bonifici stipendi
            categoria = row.get('Categoria/sottocategoria', '') or row.get('Categoria', '')
            descrizione = row.get('Descrizione', '')
            
            # Accetta sia "Risorse Umane - Salari" che descrizioni con "stip" o "FAVORE"
            is_stipendio = (
                'salari' in categoria.lower() or 
                'stipendi' in categoria.lower() or
                'stip' in descrizione.lower() or
                'favore' in descrizione.lower()
            )
            
            if not is_stipendio:
                continue
            
            # Parse data
            data_str = row.get('Data contabile') or row.get('Data valuta') or row.get('Data')
            if not data_str:
                continue
            
            # Parse importo (formato italiano: -1.234,56 o -1234,56)
            importo_str = row.get('Importo', '0')
            importo_str = importo_str.replace('.', '').replace(',', '.')
            try:
                importo = abs(float(importo_str))  # Prendi valore assoluto
            except:
                continue
            
            # Parse data (DD/MM/YYYY)
            try:
                if '/' in data_str:
                    parts = data_str.split('/')
                    data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    data_obj = date.fromisoformat(data_str)
            except:
                continue
            
            movimenti_banca.append({
                "data": data_obj,
                "importo": importo,
                "descrizione": descrizione,
                "nome_estratto": estrai_nome_da_descrizione(descrizione)
            })
    
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Parse Excel
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            
            # Trova header
            headers = [str(cell.value or '').lower() for cell in sheet[1]]
            
            for row_num in range(2, sheet.max_row + 1):
                row_data = [sheet.cell(row=row_num, column=i+1).value for i in range(len(headers))]
                
                # Cerca colonne rilevanti
                importo = None
                data_val = None
                descrizione = ""
                
                for i, h in enumerate(headers):
                    val = row_data[i]
                    if 'importo' in h and val:
                        if isinstance(val, (int, float)):
                            importo = abs(val)
                        else:
                            importo_str = str(val).replace('.', '').replace(',', '.')
                            try:
                                importo = abs(float(importo_str))
                            except:
                                pass
                    elif 'data' in h and val:
                        if isinstance(val, datetime):
                            data_val = val.date()
                        elif isinstance(val, date):
                            data_val = val
                    elif 'descri' in h and val:
                        descrizione = str(val)
                
                if importo and data_val:
                    movimenti_banca.append({
                        "data": data_val,
                        "importo": importo,
                        "descrizione": descrizione,
                        "nome_estratto": estrai_nome_da_descrizione(descrizione)
                    })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Formato file non supportato. Usa CSV o Excel.")
    
    # Carica tutti i salari non riconciliati
    salari = await db["prima_nota_salari"].find(
        {"riconciliato": {"$ne": True}},
        {"_id": 0}
    ).to_list(10000)
    
    # Crea indice per matching veloce: (nome_normalizzato, importo) -> lista salari
    salari_index = {}
    for s in salari:
        importo_s = s.get("importo_erogato") or s.get("importo") or 0
        if importo_s <= 0:
            continue  # Salta salari senza importo
        key = (normalizza_nome(s.get("dipendente", "")), round(importo_s, 2))
        if key not in salari_index:
            salari_index[key] = []
        salari_index[key].append(s)
    
    # Riconcilia
    riconciliati = 0
    non_trovati = []
    gia_riconciliati = 0
    
    for mov in movimenti_banca:
        nome_banca = mov.get("nome_estratto")
        if isinstance(nome_banca, list):
            nome_banca = " ".join(nome_banca)
        
        if not nome_banca:
            non_trovati.append({
                "data": mov["data"].isoformat() if hasattr(mov["data"], 'isoformat') else str(mov["data"]),
                "importo": mov["importo"],
                "descrizione": mov.get("descrizione", "")[:100],
                "nome": "N/D - nome non estratto"
            })
            continue
        
        nome_norm = normalizza_nome(nome_banca)
        importo = round(mov["importo"], 2)
        data_mov = mov["data"]
        
        # Cerca match: nome + importo (con tolleranza) + periodo compatibile
        found = False
        best_match = None
        best_match_score = -1
        
        for (key_nome, key_importo), salari_list in salari_index.items():
            # Tolleranza importo: 1% o 5 euro
            importo_diff = abs(key_importo - importo)
            if importo_diff > max(importo * 0.01, 5):
                continue
            
            # Verifica match nome usando fuzzy matching
            nome_match = match_nomi_fuzzy(key_nome, nome_norm)
            
            if not nome_match:
                continue
            
            # Cerca il salario con periodo più vicino alla data del movimento
            for salario in salari_list:
                if salario.get("riconciliato"):
                    continue
                
                # Calcola score basato sulla vicinanza temporale
                sal_anno = salario.get("anno", 0)
                sal_mese = salario.get("mese", 0)
                mov_anno = data_mov.year
                mov_mese = data_mov.month
                
                # Il bonifico di solito è nel mese successivo allo stipendio
                # o nello stesso mese (fine mese)
                mesi_diff = abs((mov_anno * 12 + mov_mese) - (sal_anno * 12 + sal_mese))
                
                # Score: 0 = stesso mese, 1 = mese successivo (ideale), 2+ = peggiore
                # Bonus per mese successivo (tipico per stipendi)
                if mesi_diff == 0:
                    score = 100  # Stesso mese
                elif mesi_diff == 1 and (mov_anno * 12 + mov_mese) > (sal_anno * 12 + sal_mese):
                    score = 110  # Mese successivo (caso tipico)
                elif mesi_diff <= 2:
                    score = 50
                elif mesi_diff <= 12:
                    score = 10
                else:
                    score = 0  # Troppo lontano
                
                # Bonus per match importo esatto
                if importo_diff < 0.01:
                    score += 20
                
                if score > best_match_score:
                    best_match_score = score
                    best_match = salario
        
        # Applica il match migliore (se trovato con score ragionevole)
        if best_match and best_match_score >= 10:
            await db["prima_nota_salari"].update_one(
                {"id": best_match["id"]},
                {"$set": {
                    "riconciliato": True,
                    "data_riconciliazione": datetime.utcnow().isoformat(),
                    "riferimento_banca": mov.get("descrizione", "")[:200],
                    "data_banca": data_mov.isoformat()
                }}
            )
            best_match["riconciliato"] = True
            riconciliati += 1
            found = True
        
        if not found:
            # Controlla se è già stato riconciliato in precedenza
            existing = await db["estratto_conto_salari"].find_one({
                "data": mov["data"].isoformat(),
                "importo": importo,
                "descrizione": mov["descrizione"][:100]
            })
            if existing:
                gia_riconciliati += 1
            else:
                non_trovati.append({
                    "data": mov["data"].isoformat(),
                    "importo": importo,
                    "descrizione": mov["descrizione"][:100],
                    "nome": nome_banca
                })
    
    # Salva movimenti banca importati
    for mov in movimenti_banca:
        mov_record = {
            "id": f"EC-{mov['data'].isoformat()}-{mov['importo']:.2f}",
            "data": mov["data"].isoformat(),
            "importo": mov["importo"],
            "descrizione": mov["descrizione"][:200] if mov.get("descrizione") else "",
            "nome_dipendente": mov.get("nome_estratto") if isinstance(mov.get("nome_estratto"), str) else " ".join(mov.get("nome_estratto", [])),
            "imported_at": datetime.utcnow().isoformat()
        }
        
        # Upsert per evitare duplicati
        await db["estratto_conto_salari"].update_one(
            {"id": mov_record["id"]},
            {"$set": mov_record},
            upsert=True
        )
    
    return {
        "message": "Importazione e riconciliazione completate",
        "movimenti_banca": len(movimenti_banca),
        "riconciliati": riconciliati,
        "gia_riconciliati": gia_riconciliati,
        "non_trovati": len(non_trovati),
        "dettaglio_non_trovati": non_trovati[:20]
    }


@router.get("/salari")
async def get_salari(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    dipendente: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """Lista movimenti salari con filtri."""
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    if dipendente:
        query["dipendente"] = {"$regex": dipendente, "$options": "i"}
    
    salari = await db["prima_nota_salari"].find(query, {"_id": 0}).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)
    return salari


@router.delete("/salari/reset-reconciliation")
async def reset_salari_reconciliation(
    anno: Optional[int] = Query(None),
    dipendente: Optional[str] = Query(None)
) -> Dict[str, Any]:
    """
    Reset dello stato di riconciliazione dei salari.
    Permette di testare nuovamente la riconciliazione.
    """
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if dipendente:
        query["dipendente"] = {"$regex": dipendente, "$options": "i"}
    
    # Reset solo i salari riconciliati
    query["riconciliato"] = True
    
    result = await db["prima_nota_salari"].update_many(
        query,
        {"$unset": {
            "riconciliato": "",
            "data_riconciliazione": "",
            "riferimento_banca": "",
            "data_banca": ""
        }}
    )
    
    # Opzionalmente, pulisci anche estratto_conto_salari
    ec_query = {}
    if anno:
        # Filtra per anno nella data (formato: YYYY-MM-DD)
        ec_query["data"] = {"$regex": f"^{anno}"}
    
    ec_deleted = await db["estratto_conto_salari"].delete_many(ec_query) if ec_query else {"deleted_count": 0}
    
    return {
        "message": "Reset riconciliazione completato",
        "salari_resettati": result.modified_count,
        "estratti_conto_eliminati": ec_deleted.deleted_count if hasattr(ec_deleted, 'deleted_count') else 0
    }


@router.delete("/salari/bulk/anno/{anno}")
async def delete_salari_anno(anno: int) -> Dict[str, Any]:
    """Elimina tutti i salari di un anno (per reimportazione)."""
    db = Database.get_db()
    
    result = await db["prima_nota_salari"].delete_many({"anno": anno})
    
    return {"message": f"Eliminati {result.deleted_count} salari per l'anno {anno}"}


@router.delete("/salari/{salario_id}")
async def delete_salario(salario_id: str) -> Dict[str, str]:
    """Elimina un movimento salario."""
    db = Database.get_db()
    
    result = await db["prima_nota_salari"].delete_one({"id": salario_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Salario non trovato")
    
    return {"message": "Salario eliminato"}


# ============== RICONCILIAZIONE BANCARIA SALARI ==============

# Dizionario alias dipendenti: nome_variante -> nome_standard
# Questo viene popolato dinamicamente + alias manuali
DIPENDENTI_ALIAS = {
    # Alias manuali per casi noti - TUTTE LE VARIANTI
    "dissanayaka": "dissanayaka",
    "sankapala arachchilage jananie ayachana dissanayaka": "dissanayaka",
    "sankapala jananie ayachana": "dissanayaka",
    "sankapala arachchilage jananie ayac": "dissanayaka",
    "sankapala": "dissanayaka",  # Primo nome che mappa al cognome
    "murolo": "murolo mario",
    "carotenuto": "carotenuto antonella",
    "capezzuto": "capezzuto alessandro",
}

def normalizza_nome(nome: str) -> str:
    """Normalizza un nome per il matching (rimuove accenti, lowercase, etc.)"""
    if not nome:
        return ""
    import unicodedata
    # Rimuovi accenti
    nfkd = unicodedata.normalize('NFKD', nome)
    nome_norm = ''.join([c for c in nfkd if not unicodedata.combining(c)])
    # Lowercase e rimuovi spazi extra
    nome_norm = ' '.join(nome_norm.lower().split())
    # Rimuovi "notprovide", "notprovided", etc.
    nome_norm = re.sub(r'\s*notprovid\w*', '', nome_norm)
    # Rimuovi forme societarie
    for forma in ['s.r.l.', 'srl', 's.p.a.', 'spa', 's.n.c.', 'snc', 's.a.s.', 'sas']:
        nome_norm = nome_norm.replace(forma, '')
    return ' '.join(nome_norm.split())


def estrai_cognome(nome: str) -> str:
    """Estrae il cognome (ultima parola significativa) da un nome."""
    nome_norm = normalizza_nome(nome)
    if not nome_norm:
        return ""
    words = nome_norm.split()
    # Per nomi italiani, il cognome è spesso la prima parola
    # Ma per stranieri può essere l'ultima
    # Restituiamo entrambe le possibilità
    return words[-1] if words else ""


def match_nomi_fuzzy(nome1: str, nome2: str) -> bool:
    """
    Match fuzzy tra due nomi. Gestisce:
    - Case diverse
    - Ordine cognome/nome invertito
    - Nomi parziali (solo cognome)
    - Nomi composti lunghi vs abbreviati
    - Alias noti
    """
    if not nome1 or not nome2:
        return False
    
    n1 = normalizza_nome(nome1)
    n2 = normalizza_nome(nome2)
    
    if not n1 or not n2:
        return False
    
    # Match esatto
    if n1 == n2:
        return True
    
    # Controlla alias - cerca tutte le varianti
    def get_canonical(nome):
        """Trova il nome canonico attraverso gli alias."""
        # Match esatto negli alias
        if nome in DIPENDENTI_ALIAS:
            return DIPENDENTI_ALIAS[nome]
        # Match parziale (contiene)
        for alias, canonical in DIPENDENTI_ALIAS.items():
            if alias in nome or nome in alias:
                return canonical
        return nome
    
    n1_canonical = get_canonical(n1)
    n2_canonical = get_canonical(n2)
    
    if n1_canonical == n2_canonical:
        return True
    
    # Se uno dei due è un alias conosciuto, verifica il canonico
    # contro le parole dell'altro
    if n1_canonical != n1:  # n1 aveva un alias
        if n1_canonical in n2 or n2 in n1_canonical:
            return True
        # Verifica se il canonico è una parola del n2
        for word in n2.split():
            if len(word) >= 4 and (word == n1_canonical or word in n1_canonical or n1_canonical in word):
                return True
    
    if n2_canonical != n2:  # n2 aveva un alias
        if n2_canonical in n1 or n1 in n2_canonical:
            return True
        for word in n1.split():
            if len(word) >= 4 and (word == n2_canonical or word in n2_canonical or n2_canonical in word):
                return True
    
    # Uno contenuto nell'altro (per nomi lunghi)
    if len(n1) > 5 and len(n2) > 5:
        if n1 in n2 or n2 in n1:
            return True
    
    # Split in parole
    words1 = set(n1.split())
    words2 = set(n2.split())
    
    # Se hanno almeno 2 parole in comune (nome e cognome)
    common = words1 & words2
    if len(common) >= 2:
        return True
    
    # Match per singola parola significativa (cognome)
    # Se uno dei due è una singola parola (solo cognome)
    if len(words1) == 1 or len(words2) == 1:
        single = list(words1)[0] if len(words1) == 1 else list(words2)[0]
        other = words2 if len(words1) == 1 else words1
        
        # Il cognome singolo deve matchare una parola dell'altro nome
        if len(single) >= 4:  # Almeno 4 caratteri per evitare falsi positivi
            for w in other:
                if len(w) >= 4 and (single == w or single in w or w in single):
                    return True
    
    # Match cognome: prima o ultima parola di uno con prima o ultima dell'altro
    list1 = n1.split()
    list2 = n2.split()
    
    if list1 and list2:
        # Cognomi possibili (prima e ultima parola)
        cognomi1 = {list1[0], list1[-1]}
        cognomi2 = {list2[0], list2[-1]}
        
        for c1 in cognomi1:
            for c2 in cognomi2:
                if len(c1) >= 4 and len(c2) >= 4:
                    # Match esatto o parziale
                    if c1 == c2:
                        return True
                    # Uno contenuto nell'altro (per cognomi composti)
                    if len(c1) >= 6 and len(c2) >= 6:
                        if c1 in c2 or c2 in c1:
                            return True
    
    return False


def estrai_nome_da_descrizione(descrizione: str) -> Optional[str]:
    """Estrae il nome del dipendente/fornitore dalla descrizione del bonifico."""
    if not descrizione:
        return None
    
    desc_upper = descrizione.upper()
    
    # Pattern: "FAVORE NomeCognome" o "FAVORE Nome Cognome"
    if "FAVORE" in desc_upper:
        idx = desc_upper.find("FAVORE")
        after = descrizione[idx + 7:].strip()
        
        # Prendi fino a " - " o fine stringa
        if " - " in after:
            nome = after.split(" - ")[0].strip()
        else:
            nome = after.strip()
        
        # Rimuovi "NOTPROVIDE", "NOTPROVIDED", etc.
        nome = re.sub(r'\s*NOTPROVID\w*', '', nome, flags=re.IGNORECASE)
        
        # Rimuovi codici e numeri alla fine (es. "FT 123")
        nome = re.sub(r'\s+FT\s*\d+.*$', '', nome, flags=re.IGNORECASE)
        nome = re.sub(r'\s+\d+\s*$', '', nome)
        
        return nome.strip() if nome.strip() else None
    
    # Pattern: "stip" indica stipendio, cerca nome prima
    if "STIP" in desc_upper:
        # Cerca pattern tipo "Nome stip MM YYYY"
        parts = descrizione.split()
        for i, p in enumerate(parts):
            if "stip" in p.lower() and i > 0:
                return " ".join(parts[:i]).strip()
    
    return None


@router.get("/salari/riepilogo")
async def get_salari_riepilogo(
    anno: Optional[int] = Query(None),
    dipendente: Optional[str] = Query(None)
) -> Dict[str, Any]:
    """
    Riepilogo salari con saldo per dipendente.
    Saldo = Importo Busta - Bonifico Riconciliato
    """
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if dipendente:
        query["dipendente"] = {"$regex": dipendente, "$options": "i"}
    
    salari = await db["prima_nota_salari"].find(query, {"_id": 0}).to_list(10000)
    
    # Raggruppa per dipendente
    riepilogo = {}
    for s in salari:
        nome = s.get("dipendente", "Sconosciuto")
        if nome not in riepilogo:
            riepilogo[nome] = {
                "dipendente": nome,
                "totale_busta": 0,
                "totale_bonifico": 0,
                "riconciliati": 0,
                "non_riconciliati": 0,
                "movimenti": []
            }
        
        busta = s.get("stipendio_netto") or s.get("importo") or 0
        bonifico = s.get("importo_erogato") or s.get("importo") or 0
        
        riepilogo[nome]["totale_busta"] += busta
        riepilogo[nome]["totale_bonifico"] += bonifico
        
        if s.get("riconciliato"):
            riepilogo[nome]["riconciliati"] += 1
        else:
            riepilogo[nome]["non_riconciliati"] += 1
    
    # Calcola saldo
    for nome, data in riepilogo.items():
        data["saldo"] = round(data["totale_busta"] - data["totale_bonifico"], 2)
    
    return {
        "anno": anno,
        "totale_dipendenti": len(riepilogo),
        "totale_buste": sum(r["totale_busta"] for r in riepilogo.values()),
        "totale_bonifici": sum(r["totale_bonifico"] for r in riepilogo.values()),
        "totale_riconciliati": sum(r["riconciliati"] for r in riepilogo.values()),
        "totale_non_riconciliati": sum(r["non_riconciliati"] for r in riepilogo.values()),
        "dipendenti": list(riepilogo.values())
    }


@router.get("/dipendenti-lista")
async def get_dipendenti_lista() -> List[str]:
    """Lista nomi unici dipendenti dai salari."""
    db = Database.get_db()
    
    dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    return sorted([d for d in dipendenti if d])


# ============== DIPENDENTE DETAIL (must be after specific routes) ==============

@router.get("/{dipendente_id}")
async def get_dipendente(dipendente_id: str) -> Dict[str, Any]:
    """Dettaglio singolo dipendente."""
    db = Database.get_db()
    
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return dipendente


@router.put("/{dipendente_id}")
async def update_dipendente(dipendente_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna dipendente."""
    db = Database.get_db()
    
    # Rimuovi campi non modificabili
    data.pop("id", None)
    data.pop("created_at", None)
    
    data["updated_at"] = datetime.utcnow().isoformat()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Dipendente aggiornato"}


@router.delete("/{dipendente_id}")
async def delete_dipendente(dipendente_id: str) -> Dict[str, str]:
    """Elimina dipendente."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].delete_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Dipendente eliminato"}


# ============== TURNI ==============

@router.get("/turni/settimana")
async def get_turni_settimana(
    data_inizio: str = Query(..., description="Data inizio settimana (YYYY-MM-DD)")
) -> Dict[str, Any]:
    """Ritorna i turni per una settimana."""
    db = Database.get_db()
    
    # Calcola date settimana
    start = datetime.strptime(data_inizio, "%Y-%m-%d")
    dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
    
    # Trova turni
    turni = await db["turni_dipendenti"].find(
        {"data": {"$in": dates}},
        {"_id": 0}
    ).to_list(1000)
    
    # Organizza per dipendente e data
    turni_by_employee = {}
    for t in turni:
        emp_id = t.get("dipendente_id")
        if emp_id not in turni_by_employee:
            turni_by_employee[emp_id] = {}
        turni_by_employee[emp_id][t.get("data")] = t.get("turno")
    
    # Carica dipendenti attivi
    dipendenti = await db[Collections.EMPLOYEES].find(
        {"attivo": {"$ne": False}},
        {"_id": 0, "id": 1, "nome_completo": 1, "mansione": 1}
    ).to_list(100)
    
    return {
        "settimana": dates,
        "dipendenti": dipendenti,
        "turni": turni_by_employee
    }


@router.post("/turni/salva")
async def salva_turni(data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Salva turni per una settimana."""
    db = Database.get_db()
    
    turni = data.get("turni", {})  # {dipendente_id: {data: turno}}
    
    for dip_id, turni_dip in turni.items():
        for data_turno, tipo_turno in turni_dip.items():
            await db["turni_dipendenti"].update_one(
                {"dipendente_id": dip_id, "data": data_turno},
                {"$set": {
                    "dipendente_id": dip_id,
                    "data": data_turno,
                    "turno": tipo_turno,
                    "updated_at": datetime.utcnow().isoformat()
                }},
                upsert=True
            )
    
    return {"message": "Turni salvati"}


# ============== LIBRETTI SANITARI ==============

@router.get("/libretti/scadenze")
async def get_libretti_scadenze(days: int = Query(30, ge=1, le=365)) -> List[Dict[str, Any]]:
    """Ritorna dipendenti con libretto in scadenza."""
    db = Database.get_db()
    
    today = datetime.utcnow()
    deadline = today + timedelta(days=days)
    
    dipendenti = await db[Collections.EMPLOYEES].find(
        {
            "libretto_scadenza": {"$ne": None},
            "$or": [
                {"libretto_scadenza": {"$lte": deadline.isoformat()[:10]}},
                {"libretto_scadenza": {"$lt": today.isoformat()[:10]}}  # Già scaduti
            ]
        },
        {"_id": 0}
    ).sort("libretto_scadenza", 1).to_list(100)
    
    return dipendenti


@router.put("/{dipendente_id}/libretto")
async def update_libretto(dipendente_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna dati libretto sanitario."""
    db = Database.get_db()
    
    update_data = {}
    if "libretto_numero" in data:
        update_data["libretto_numero"] = data["libretto_numero"]
    if "libretto_scadenza" in data:
        update_data["libretto_scadenza"] = data["libretto_scadenza"]
    if "libretto_file" in data:
        update_data["libretto_file"] = data["libretto_file"]
    
    update_data["updated_at"] = datetime.utcnow().isoformat()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Libretto aggiornato"}


# ============== PORTALE DIPENDENTI ==============

@router.post("/{dipendente_id}/invita-portale")
async def invita_portale(dipendente_id: str) -> Dict[str, str]:
    """Segna dipendente come invitato al portale."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": {
            "portale_invitato": True,
            "portale_data_invito": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Invito inviato"}


@router.post("/invita-multipli")
async def invita_multipli(dipendenti_ids: List[str] = Body(...)) -> Dict[str, Any]:
    """Invita multipli dipendenti al portale."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].update_many(
        {"id": {"$in": dipendenti_ids}},
        {"$set": {
            "portale_invitato": True,
            "portale_data_invito": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    
    return {"message": f"Invitati {result.modified_count} dipendenti"}


# ============== LIBRETTI SANITARI - COLLECTION SEPARATA ==============

@router.get("/libretti-sanitari/all")
async def get_all_libretti_sanitari() -> List[Dict[str, Any]]:
    """Lista tutti i libretti sanitari."""
    db = Database.get_db()
    
    libretti = await db["libretti_sanitari"].find({}, {"_id": 0}).sort("data_scadenza", 1).to_list(500)
    return libretti


@router.post("/libretti-sanitari")
async def create_libretto_sanitario(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo libretto sanitario."""
    db = Database.get_db()
    
    libretto = {
        "id": str(uuid.uuid4()),
        "dipendente_nome": data.get("dipendente_nome", ""),
        "dipendente_id": data.get("dipendente_id"),
        "numero_libretto": data.get("numero_libretto", ""),
        "data_rilascio": data.get("data_rilascio"),
        "data_scadenza": data.get("data_scadenza"),
        "stato": data.get("stato", "valido"),
        "note": data.get("note", ""),
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    await db["libretti_sanitari"].insert_one(libretto)
    libretto.pop("_id", None)
    
    return libretto


@router.put("/libretti-sanitari/{libretto_id}")
async def update_libretto_sanitario(libretto_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna libretto sanitario."""
    db = Database.get_db()
    
    data.pop("id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.utcnow().isoformat()
    
    result = await db["libretti_sanitari"].update_one(
        {"id": libretto_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Libretto non trovato")
    
    return {"message": "Libretto aggiornato"}


@router.delete("/libretti-sanitari/{libretto_id}")
async def delete_libretto_sanitario(libretto_id: str) -> Dict[str, str]:
    """Elimina libretto sanitario."""
    db = Database.get_db()
    
    result = await db["libretti_sanitari"].delete_one({"id": libretto_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Libretto non trovato")
    
    return {"message": "Libretto eliminato"}


# ============== LIBRO UNICO ==============

@router.get("/libro-unico/presenze")
async def get_libro_unico_presenze(month_year: str = Query(..., description="Formato: YYYY-MM")) -> List[Dict[str, Any]]:
    """Ottieni presenze dal libro unico per un mese."""
    db = Database.get_db()
    
    presenze = await db["libro_unico_presenze"].find(
        {"month_year": month_year},
        {"_id": 0}
    ).sort("dipendente_nome", 1).to_list(500)
    
    return presenze


@router.get("/libro-unico/salaries")
async def get_libro_unico_salaries(month_year: Optional[str] = Query(None, description="Formato: YYYY-MM")) -> List[Dict[str, Any]]:
    """Ottieni buste paga dal libro unico."""
    db = Database.get_db()
    
    query = {}
    if month_year:
        query["month_year"] = month_year
    
    salaries = await db["libro_unico_salaries"].find(query, {"_id": 0}).sort([("month_year", -1), ("dipendente_nome", 1)]).to_list(1000)
    
    return salaries


@router.post("/libro-unico/upload")
async def upload_libro_unico(
    file: UploadFile = File(...),
    month_year: str = Query(..., description="Formato: YYYY-MM")
) -> Dict[str, Any]:
    """
    Upload e parsing di PDF/Excel del libro unico.
    Estrae presenze, buste paga e aggiorna anagrafica.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    content = await file.read()
    
    presenze_count = 0
    salaries_count = 0
    payments_count = 0
    anagrafica_created = 0
    anagrafica_updated = 0
    
    if filename.endswith('.pdf'):
        # Parsing PDF - estrai dati buste paga
        try:
            import fitz  # PyMuPDF
            
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            
            for page in pdf_doc:
                text = page.get_text()
                
                # Cerca pattern busta paga
                # Pattern: NOME COGNOME seguito da valori numerici
                lines = text.split('\n')
                
                current_employee = None
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    # Cerca nome dipendente (tutte maiuscole, senza numeri)
                    if line and line.isupper() and not any(c.isdigit() for c in line) and len(line) > 3:
                        # Potrebbe essere un nome
                        if len(line.split()) >= 2:
                            current_employee = line
                    
                    # Cerca pattern "NETTO A PAGARE" o simili
                    if current_employee and 'netto' in line.lower():
                        # Cerca importo nelle righe successive
                        for j in range(i, min(i+5, len(lines))):
                            next_line = lines[j].strip()
                            # Cerca pattern €1.234,56 o 1234.56
                            import re
                            amounts = re.findall(r'[\d.,]+', next_line)
                            for amt in amounts:
                                try:
                                    # Converti formato italiano
                                    amt_clean = amt.replace('.', '').replace(',', '.')
                                    netto = float(amt_clean)
                                    if 100 < netto < 10000:  # Range plausibile per stipendio
                                        # Salva busta paga
                                        salary_doc = {
                                            "id": str(uuid.uuid4()),
                                            "dipendente_nome": current_employee,
                                            "month_year": month_year,
                                            "netto_a_pagare": netto,
                                            "acconto_pagato": 0,
                                            "differenza": netto,
                                            "note": f"Importato da PDF: {file.filename}",
                                            "created_at": datetime.utcnow().isoformat()
                                        }
                                        
                                        # Verifica se esiste già
                                        existing = await db["libro_unico_salaries"].find_one({
                                            "dipendente_nome": current_employee,
                                            "month_year": month_year
                                        })
                                        
                                        if not existing:
                                            await db["libro_unico_salaries"].insert_one(salary_doc)
                                            salaries_count += 1
                                        
                                        current_employee = None
                                        break
                                except:
                                    continue
                            if current_employee is None:
                                break
            
            pdf_doc.close()
            
        except ImportError:
            raise HTTPException(status_code=500, detail="PyMuPDF non installato per parsing PDF")
        except Exception as e:
            logger.error(f"Errore parsing PDF: {e}")
            raise HTTPException(status_code=400, detail=f"Errore parsing PDF: {str(e)}")
    
    elif filename.endswith(('.xlsx', '.xls')):
        # Parsing Excel
        try:
            import pandas as pd
            
            df = pd.read_excel(io.BytesIO(content))
            
            # Cerca colonne rilevanti
            columns_lower = {c.lower(): c for c in df.columns}
            
            for idx, row in df.iterrows():
                try:
                    # Cerca nome dipendente
                    nome = None
                    for col_key in ['nome', 'dipendente', 'cognome e nome', 'nominativo']:
                        if col_key in columns_lower:
                            nome = str(row[columns_lower[col_key]]) if pd.notna(row[columns_lower[col_key]]) else None
                            break
                    
                    if not nome or nome == 'nan':
                        continue
                    
                    # Cerca importo netto
                    netto = None
                    for col_key in ['netto', 'netto a pagare', 'importo', 'stipendio']:
                        if col_key in columns_lower:
                            val = row[columns_lower[col_key]]
                            if pd.notna(val):
                                try:
                                    netto = float(val)
                                except:
                                    pass
                            break
                    
                    if netto and netto > 0:
                        salary_doc = {
                            "id": str(uuid.uuid4()),
                            "dipendente_nome": nome.upper(),
                            "month_year": month_year,
                            "netto_a_pagare": netto,
                            "acconto_pagato": 0,
                            "differenza": netto,
                            "note": f"Importato da Excel: {file.filename}",
                            "created_at": datetime.utcnow().isoformat()
                        }
                        
                        existing = await db["libro_unico_salaries"].find_one({
                            "dipendente_nome": nome.upper(),
                            "month_year": month_year
                        })
                        
                        if not existing:
                            await db["libro_unico_salaries"].insert_one(salary_doc)
                            salaries_count += 1
                        
                except Exception as e:
                    logger.warning(f"Errore riga {idx}: {e}")
                    continue
                    
        except Exception as e:
            logger.error(f"Errore parsing Excel: {e}")
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa PDF o Excel.")
    
    return {
        "success": True,
        "message": f"Import completato per {month_year}",
        "presenze_count": presenze_count,
        "salaries_count": salaries_count,
        "payments_count": payments_count,
        "anagrafica_created": anagrafica_created,
        "anagrafica_updated": anagrafica_updated
    }


@router.get("/libro-unico/export-excel")
async def export_libro_unico_excel(month_year: str = Query(..., description="Formato: YYYY-MM")):
    """Esporta libro unico in Excel."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    db = Database.get_db()
    
    # Recupera dati
    salaries = await db["libro_unico_salaries"].find(
        {"month_year": month_year},
        {"_id": 0}
    ).sort("dipendente_nome", 1).to_list(500)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Libro Unico {month_year}"
    
    # Header
    headers = ["Dipendente", "Netto a Pagare", "Acconto", "Differenza", "Note"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for row_num, salary in enumerate(salaries, 2):
        ws.cell(row=row_num, column=1, value=salary.get("dipendente_nome", ""))
        ws.cell(row=row_num, column=2, value=salary.get("netto_a_pagare", 0))
        ws.cell(row=row_num, column=3, value=salary.get("acconto_pagato", 0))
        ws.cell(row=row_num, column=4, value=salary.get("differenza", 0))
        ws.cell(row=row_num, column=5, value=salary.get("note", ""))
    
    # Larghezze colonne
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=libro_unico_{month_year}.xlsx"}
    )


@router.put("/libro-unico/salaries/{salary_id}")
async def update_libro_unico_salary(
    salary_id: str,
    netto_a_pagare: float = Query(...),
    acconto_pagato: float = Query(0),
    differenza: float = Query(...),
    note: str = Query("")
) -> Dict[str, str]:
    """Aggiorna busta paga libro unico."""
    db = Database.get_db()
    
    result = await db["libro_unico_salaries"].update_one(
        {"id": salary_id},
        {"$set": {
            "netto_a_pagare": netto_a_pagare,
            "acconto_pagato": acconto_pagato,
            "differenza": differenza,
            "note": note,
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Aggiornato"}


@router.delete("/libro-unico/salaries/{salary_id}")
async def delete_libro_unico_salary(salary_id: str) -> Dict[str, str]:
    """Elimina busta paga libro unico."""
    db = Database.get_db()
    
    result = await db["libro_unico_salaries"].delete_one({"id": salary_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Eliminato"}


# Note: salari and buste-paga routes are defined earlier in the file to avoid route conflict with /{dipendente_id}



@router.get("/portale/stats")
async def get_portale_stats() -> Dict[str, Any]:
    """Statistiche portale dipendenti."""
    db = Database.get_db()
    
    total = await db[Collections.EMPLOYEES].count_documents({"attivo": {"$ne": False}})
    invitati = await db[Collections.EMPLOYEES].count_documents({"portale_invitato": True})
    registrati = await db[Collections.EMPLOYEES].count_documents({"portale_registrato": True})
    mai_invitati = await db[Collections.EMPLOYEES].count_documents({
        "attivo": {"$ne": False},
        "$or": [{"portale_invitato": False}, {"portale_invitato": {"$exists": False}}]
    })
    
    return {
        "totale": total,
        "mai_invitati": mai_invitati,
        "invitati": invitati,
        "registrati": registrati
    }


