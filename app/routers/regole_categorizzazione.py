"""
Router per gestione regole di categorizzazione contabile.
Permette download/upload Excel delle regole e gestione via UI.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Dict, Any, List, Optional
from datetime import datetime
import io
import uuid
import logging

from app.database import Database

logger = logging.getLogger(__name__)
router = APIRouter()


# ============== STRUTTURA DATI REGOLE ==============

DEFAULT_PIANO_CONTI = {
    "05.01.01": "Acquisto merci",
    "05.01.02": "Acquisto materie prime",
    "05.01.03": "Acquisto bevande alcoliche",
    "05.01.04": "Acquisto bevande analcoliche",
    "05.01.05": "Acquisto prodotti alimentari",
    "05.01.06": "Acquisto piccola utensileria",
    "05.01.07": "Materiali di consumo e imballaggio",
    "05.01.08": "Prodotti per pulizia e igiene",
    "05.01.09": "Acquisto caffe e affini",
    "05.01.10": "Acquisto surgelati",
    "05.01.11": "Acquisto prodotti da forno",
    "05.01.12": "Materiale edile e costruzioni",
    "05.01.13": "Additivi e ingredienti alimentari",
    "05.02.01": "Costi per servizi",
    "05.02.02": "Utenze (luce, gas, acqua)",
    "05.02.03": "Canoni di locazione",
    "05.02.04": "Utenze - Acqua",
    "05.02.05": "Utenze - Energia elettrica",
    "05.02.06": "Utenze - Gas",
    "05.02.07": "Telefonia e comunicazioni",
    "05.02.08": "Software e servizi cloud",
    "05.02.09": "Noleggi e locazioni operative",
    "05.02.10": "Manutenzioni e riparazioni",
    "05.02.11": "Carburanti e lubrificanti",
    "05.02.12": "Consulenze e prestazioni professionali",
    "05.02.13": "Assicurazioni",
    "05.02.14": "Pubblicita e marketing",
    "05.02.15": "Omaggi e spese promozionali",
    "05.02.16": "Trasporti su acquisti",
    "05.02.17": "Spese viaggio e trasferte",
    "05.02.18": "Spese di rappresentanza",
    "05.02.21": "Diritti SIAE e licenze",
    "05.02.22": "Noleggio automezzi",
    "05.02.23": "Canoni e abbonamenti",
    "05.02.24": "Arredi e tappezzeria",
    "05.03.05": "Buoni pasto dipendenti",
    "05.05.02": "Spese e commissioni bancarie",
}

DEFAULT_CATEGORIE = {
    "bevande_alcoliche": {"conto": "05.01.03", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "bevande_analcoliche": {"conto": "05.01.04", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "alimentari": {"conto": "05.01.05", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "caffe": {"conto": "05.01.09", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "surgelati": {"conto": "05.01.10", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "pasticceria": {"conto": "05.01.11", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "ferramenta": {"conto": "05.01.06", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "materiale_edile": {"conto": "05.01.12", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "imballaggi": {"conto": "05.01.07", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "pulizia": {"conto": "05.01.08", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "utenze_acqua": {"conto": "05.02.04", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "utenze_elettricita": {"conto": "05.02.05", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "utenze_gas": {"conto": "05.02.06", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "telefonia": {"conto": "05.02.07", "deducibilita_ires": 80, "deducibilita_irap": 80},
    "software_cloud": {"conto": "05.02.08", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "noleggio_attrezzature": {"conto": "05.02.09", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "noleggio_auto": {"conto": "05.02.22", "deducibilita_ires": 20, "deducibilita_irap": 20},
    "manutenzione": {"conto": "05.02.10", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "carburante": {"conto": "05.02.11", "deducibilita_ires": 20, "deducibilita_irap": 20},
    "consulenze": {"conto": "05.02.12", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "assicurazioni": {"conto": "05.02.13", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "pubblicita": {"conto": "05.02.14", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "trasporti": {"conto": "05.02.16", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "rappresentanza": {"conto": "05.02.18", "deducibilita_ires": 75, "deducibilita_irap": 100},
    "diritti_autore": {"conto": "05.02.21", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "canoni_abbonamenti": {"conto": "05.02.23", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "buoni_pasto": {"conto": "05.03.05", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "spese_bancarie": {"conto": "05.05.02", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "tappezzeria": {"conto": "05.02.24", "deducibilita_ires": 100, "deducibilita_irap": 100},
    "merci_generiche": {"conto": "05.01.01", "deducibilita_ires": 100, "deducibilita_irap": 100},
}


@router.get("/download-regole")
async def download_regole_excel():
    """
    Scarica un file Excel con tutte le regole di categorizzazione:
    - Foglio 1: Regole per Fornitore (fornitore → categoria)
    - Foglio 2: Regole per Descrizione (parola chiave → categoria) 
    - Foglio 3: Categorie (categoria → conto + deducibilità)
    - Foglio 4: Piano dei Conti (codice → nome)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    db = Database.get_db()
    
    # Carica regole dal database o usa default
    regole_fornitori = await db["regole_categorizzazione_fornitori"].find({}, {"_id": 0}).to_list(5000)
    regole_descrizioni = await db["regole_categorizzazione_descrizioni"].find({}, {"_id": 0}).to_list(5000)
    categorie = await db["regole_categorie"].find({}, {"_id": 0}).to_list(100)
    
    # Se non ci sono regole nel DB, usa quelle hardcoded
    if not regole_fornitori:
        regole_fornitori = await _get_default_regole_fornitori()
    if not categorie:
        categorie = [{"categoria": k, **v} for k, v in DEFAULT_CATEGORIE.items()]
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wb = Workbook()
    
    # === FOGLIO 1: REGOLE FORNITORI ===
    ws_forn = wb.active
    ws_forn.title = "Regole Fornitori"
    
    headers_forn = ['Fornitore (contiene)', 'Categoria', 'Note']
    for col, header in enumerate(headers_forn, 1):
        cell = ws_forn.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row, regola in enumerate(regole_fornitori, 2):
        ws_forn.cell(row=row, column=1, value=regola.get("pattern", "")).border = border
        ws_forn.cell(row=row, column=2, value=regola.get("categoria", "")).border = border
        ws_forn.cell(row=row, column=3, value=regola.get("note", "")).border = border
    
    # Aggiungi righe vuote per nuove voci
    for row in range(len(regole_fornitori) + 2, len(regole_fornitori) + 52):
        for col in range(1, 4):
            ws_forn.cell(row=row, column=col).border = border
    
    ws_forn.column_dimensions['A'].width = 40
    ws_forn.column_dimensions['B'].width = 25
    ws_forn.column_dimensions['C'].width = 40
    
    # === FOGLIO 2: REGOLE DESCRIZIONI ===
    ws_desc = wb.create_sheet("Regole Descrizioni")
    
    headers_desc = ['Parola Chiave (contiene)', 'Categoria', 'Note']
    for col, header in enumerate(headers_desc, 1):
        cell = ws_desc.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    if not regole_descrizioni:
        regole_descrizioni = await _get_default_regole_descrizioni()
    
    for row, regola in enumerate(regole_descrizioni, 2):
        ws_desc.cell(row=row, column=1, value=regola.get("pattern", "")).border = border
        ws_desc.cell(row=row, column=2, value=regola.get("categoria", "")).border = border
        ws_desc.cell(row=row, column=3, value=regola.get("note", "")).border = border
    
    # Righe vuote
    for row in range(len(regole_descrizioni) + 2, len(regole_descrizioni) + 52):
        for col in range(1, 4):
            ws_desc.cell(row=row, column=col).border = border
    
    ws_desc.column_dimensions['A'].width = 40
    ws_desc.column_dimensions['B'].width = 25
    ws_desc.column_dimensions['C'].width = 40
    
    # === FOGLIO 3: CATEGORIE ===
    ws_cat = wb.create_sheet("Categorie")
    
    headers_cat = ['Categoria', 'Codice Conto', 'Nome Conto', 'Deducibilita IRES %', 'Deducibilita IRAP %', 'Note']
    for col, header in enumerate(headers_cat, 1):
        cell = ws_cat.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row, cat in enumerate(categorie, 2):
        conto = cat.get("conto", "")
        ws_cat.cell(row=row, column=1, value=cat.get("categoria", "")).border = border
        ws_cat.cell(row=row, column=2, value=conto).border = border
        ws_cat.cell(row=row, column=3, value=DEFAULT_PIANO_CONTI.get(conto, "")).border = border
        ws_cat.cell(row=row, column=4, value=cat.get("deducibilita_ires", 100)).border = border
        ws_cat.cell(row=row, column=5, value=cat.get("deducibilita_irap", 100)).border = border
        ws_cat.cell(row=row, column=6, value=cat.get("note", "")).border = border
    
    # Righe vuote
    for row in range(len(categorie) + 2, len(categorie) + 22):
        for col in range(1, 7):
            ws_cat.cell(row=row, column=col).border = border
    
    for col, width in enumerate([20, 15, 35, 18, 18, 40], 1):
        ws_cat.column_dimensions[get_column_letter(col)].width = width
    
    # === FOGLIO 4: PIANO DEI CONTI ===
    ws_piano = wb.create_sheet("Piano dei Conti")
    
    headers_piano = ['Codice Conto', 'Nome Conto', 'Categoria Bilancio']
    for col, header in enumerate(headers_piano, 1):
        cell = ws_piano.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    piano_conti = await db["piano_conti"].find({}, {"_id": 0, "codice": 1, "nome": 1, "categoria": 1}).sort("codice", 1).to_list(200)
    
    for row, conto in enumerate(piano_conti, 2):
        ws_piano.cell(row=row, column=1, value=conto.get("codice", "")).border = border
        ws_piano.cell(row=row, column=2, value=conto.get("nome", "")).border = border
        ws_piano.cell(row=row, column=3, value=conto.get("categoria", "")).border = border
    
    ws_piano.column_dimensions['A'].width = 15
    ws_piano.column_dimensions['B'].width = 40
    ws_piano.column_dimensions['C'].width = 20
    
    # === FOGLIO 5: ISTRUZIONI ===
    ws_help = wb.create_sheet("Istruzioni")
    
    istruzioni = [
        ("ISTRUZIONI PER LA MODIFICA DELLE REGOLE DI CATEGORIZZAZIONE", ""),
        ("", ""),
        ("FOGLIO 'Regole Fornitori':", ""),
        ("- Colonna A: Inserire parte del nome fornitore (es. 'KIMBO' per KIMBO S.P.A.)", ""),
        ("- Colonna B: Categoria da assegnare (deve esistere nel foglio Categorie)", ""),
        ("- Colonna C: Note opzionali", ""),
        ("- Le regole sono case-insensitive (KIMBO = kimbo = Kimbo)", ""),
        ("", ""),
        ("FOGLIO 'Regole Descrizioni':", ""),
        ("- Colonna A: Parola chiave nella descrizione prodotto", ""),
        ("- Colonna B: Categoria da assegnare", ""),
        ("- Colonna C: Note opzionali", ""),
        ("", ""),
        ("FOGLIO 'Categorie':", ""),
        ("- Colonna A: Nome categoria (usato nei fogli precedenti)", ""),
        ("- Colonna B: Codice conto del Piano dei Conti", ""),
        ("- Colonna D: Percentuale deducibilita IRES (0-100)", ""),
        ("- Colonna E: Percentuale deducibilita IRAP (0-100)", ""),
        ("", ""),
        ("PRIORITA' CATEGORIZZAZIONE:", ""),
        ("1. Prima si cerca nel nome fornitore", ""),
        ("2. Poi si cerca nella descrizione prodotto", ""),
        ("3. Se non trova nulla, assegna 'merci_generiche'", ""),
        ("", ""),
        ("DOPO LE MODIFICHE:", ""),
        ("- Salvare il file Excel", ""),
        ("- Caricare dalla pagina Regole Categorizzazione", ""),
        ("- Cliccare 'Ricategorizza Fatture' per applicare", ""),
    ]
    
    for row, (text, _) in enumerate(istruzioni, 1):
        cell = ws_help.cell(row=row, column=1, value=text)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif "FOGLIO" in text or "PRIORITA'" in text or "DOPO" in text:
            cell.font = Font(bold=True)
    
    ws_help.column_dimensions['A'].width = 70
    
    # Genera file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=regole_categorizzazione.xlsx'}
    )


@router.post("/upload-regole")
async def upload_regole_excel(file: UploadFile = File(...)):
    """
    Carica un file Excel con le regole di categorizzazione modificate.
    Aggiorna le regole nel database.
    """
    from openpyxl import load_workbook
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato Excel (.xlsx)")
    
    db = Database.get_db()
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        
        stats = {
            "regole_fornitori_caricate": 0,
            "regole_descrizioni_caricate": 0,
            "categorie_caricate": 0,
            "errori": []
        }
        
        # === PROCESSA FOGLIO FORNITORI ===
        if "Regole Fornitori" in wb.sheetnames:
            ws = wb["Regole Fornitori"]
            regole_forn = []
            
            for row in range(2, ws.max_row + 1):
                pattern = ws.cell(row=row, column=1).value
                categoria = ws.cell(row=row, column=2).value
                note = ws.cell(row=row, column=3).value
                
                if pattern and categoria:
                    pattern = str(pattern).strip()
                    categoria = str(categoria).strip().lower().replace(" ", "_")
                    
                    regole_forn.append({
                        "id": str(uuid.uuid4()),
                        "pattern": pattern,
                        "categoria": categoria,
                        "note": str(note or ""),
                        "tipo": "fornitore",
                        "created_at": datetime.utcnow().isoformat(),
                        "attivo": True
                    })
            
            if regole_forn:
                # Elimina vecchie regole e inserisci nuove
                await db["regole_categorizzazione_fornitori"].delete_many({})
                await db["regole_categorizzazione_fornitori"].insert_many(regole_forn)
                stats["regole_fornitori_caricate"] = len(regole_forn)
        
        # === PROCESSA FOGLIO DESCRIZIONI ===
        if "Regole Descrizioni" in wb.sheetnames:
            ws = wb["Regole Descrizioni"]
            regole_desc = []
            
            for row in range(2, ws.max_row + 1):
                pattern = ws.cell(row=row, column=1).value
                categoria = ws.cell(row=row, column=2).value
                note = ws.cell(row=row, column=3).value
                
                if pattern and categoria:
                    pattern = str(pattern).strip()
                    categoria = str(categoria).strip().lower().replace(" ", "_")
                    
                    regole_desc.append({
                        "id": str(uuid.uuid4()),
                        "pattern": pattern,
                        "categoria": categoria,
                        "note": str(note or ""),
                        "tipo": "descrizione",
                        "created_at": datetime.utcnow().isoformat(),
                        "attivo": True
                    })
            
            if regole_desc:
                await db["regole_categorizzazione_descrizioni"].delete_many({})
                await db["regole_categorizzazione_descrizioni"].insert_many(regole_desc)
                stats["regole_descrizioni_caricate"] = len(regole_desc)
        
        # === PROCESSA FOGLIO CATEGORIE ===
        if "Categorie" in wb.sheetnames:
            ws = wb["Categorie"]
            categorie = []
            
            for row in range(2, ws.max_row + 1):
                categoria = ws.cell(row=row, column=1).value
                conto = ws.cell(row=row, column=2).value
                ded_ires = ws.cell(row=row, column=4).value
                ded_irap = ws.cell(row=row, column=5).value
                note = ws.cell(row=row, column=6).value
                
                if categoria and conto:
                    categoria = str(categoria).strip().lower().replace(" ", "_")
                    
                    try:
                        ded_ires = float(ded_ires) if ded_ires else 100
                        ded_irap = float(ded_irap) if ded_irap else 100
                    except (ValueError, TypeError):
                        ded_ires = 100
                        ded_irap = 100
                    
                    categorie.append({
                        "id": str(uuid.uuid4()),
                        "categoria": categoria,
                        "conto": str(conto).strip(),
                        "deducibilita_ires": min(100, max(0, ded_ires)),
                        "deducibilita_irap": min(100, max(0, ded_irap)),
                        "note": str(note or ""),
                        "created_at": datetime.utcnow().isoformat()
                    })
            
            if categorie:
                await db["regole_categorie"].delete_many({})
                await db["regole_categorie"].insert_many(categorie)
                stats["categorie_caricate"] = len(categorie)
        
        return {
            "success": True,
            "message": f"Regole caricate con successo",
            **stats
        }
        
    except Exception as e:
        logger.error(f"Errore upload regole: {e}")
        raise HTTPException(status_code=500, detail=f"Errore nel processamento del file: {str(e)}")


@router.get("/regole")
async def get_regole() -> Dict[str, Any]:
    """Ottiene tutte le regole di categorizzazione (dal database + default)."""
    db = Database.get_db()
    
    # Regole dal database
    regole_fornitori_db = await db["regole_categorizzazione_fornitori"].find(
        {"attivo": True}, {"_id": 0}
    ).to_list(5000)
    
    regole_descrizioni_db = await db["regole_categorizzazione_descrizioni"].find(
        {"attivo": True}, {"_id": 0}
    ).to_list(5000)
    
    categorie = await db["regole_categorie"].find({}, {"_id": 0}).to_list(100)
    
    # Carica sempre le regole di default e uniscile alle regole DB
    regole_fornitori_default = await _get_default_regole_fornitori()
    regole_descrizioni_default = await _get_default_regole_descrizioni()
    
    # Unisci: prima le regole DB (priorità), poi quelle di default
    patterns_db_forn = {r.get("pattern", "").lower() for r in regole_fornitori_db}
    patterns_db_desc = {r.get("pattern", "").lower() for r in regole_descrizioni_db}
    
    # Aggiungi regole default non presenti nel DB
    for r in regole_fornitori_default:
        if r["pattern"].lower() not in patterns_db_forn:
            r["source"] = "default"
            regole_fornitori_db.append(r)
    
    for r in regole_descrizioni_default:
        if r["pattern"].lower() not in patterns_db_desc:
            r["source"] = "default"
            regole_descrizioni_db.append(r)
    
    # Se non ci sono categorie nel DB, usa default
    if not categorie:
        categorie = [{"categoria": k, **v} for k, v in DEFAULT_CATEGORIE.items()]
    
    return {
        "regole_fornitori": regole_fornitori_db,
        "regole_descrizioni": regole_descrizioni_db,
        "categorie": categorie,
        "piano_conti": DEFAULT_PIANO_CONTI,
        "totale_regole": len(regole_fornitori_db) + len(regole_descrizioni_db)
    }


@router.post("/regole/fornitore")
async def aggiungi_regola_fornitore(data: Dict[str, Any]) -> Dict[str, Any]:
    """Aggiunge una nuova regola per fornitore."""
    db = Database.get_db()
    
    pattern = data.get("pattern", "").strip()
    categoria = data.get("categoria", "").strip().lower().replace(" ", "_")
    note = data.get("note", "")
    
    if not pattern or not categoria:
        raise HTTPException(status_code=400, detail="Pattern e categoria sono obbligatori")
    
    # Verifica se esiste già
    existing = await db["regole_categorizzazione_fornitori"].find_one({"pattern": pattern})
    if existing:
        # Aggiorna
        await db["regole_categorizzazione_fornitori"].update_one(
            {"pattern": pattern},
            {"$set": {"categoria": categoria, "note": note, "updated_at": datetime.utcnow().isoformat()}}
        )
        return {"success": True, "message": "Regola aggiornata", "action": "updated"}
    
    # Inserisci nuova
    regola = {
        "id": str(uuid.uuid4()),
        "pattern": pattern,
        "categoria": categoria,
        "note": note,
        "tipo": "fornitore",
        "created_at": datetime.utcnow().isoformat(),
        "attivo": True
    }
    await db["regole_categorizzazione_fornitori"].insert_one(regola)
    
    return {"success": True, "message": "Regola aggiunta", "action": "created", "id": regola["id"]}


@router.post("/regole/descrizione")
async def aggiungi_regola_descrizione(data: Dict[str, Any]) -> Dict[str, Any]:
    """Aggiunge una nuova regola per descrizione prodotto."""
    db = Database.get_db()
    
    pattern = data.get("pattern", "").strip()
    categoria = data.get("categoria", "").strip().lower().replace(" ", "_")
    note = data.get("note", "")
    
    if not pattern or not categoria:
        raise HTTPException(status_code=400, detail="Pattern e categoria sono obbligatori")
    
    existing = await db["regole_categorizzazione_descrizioni"].find_one({"pattern": pattern})
    if existing:
        await db["regole_categorizzazione_descrizioni"].update_one(
            {"pattern": pattern},
            {"$set": {"categoria": categoria, "note": note, "updated_at": datetime.utcnow().isoformat()}}
        )
        return {"success": True, "message": "Regola aggiornata", "action": "updated"}
    
    regola = {
        "id": str(uuid.uuid4()),
        "pattern": pattern,
        "categoria": categoria,
        "note": note,
        "tipo": "descrizione",
        "created_at": datetime.utcnow().isoformat(),
        "attivo": True
    }
    await db["regole_categorizzazione_descrizioni"].insert_one(regola)
    
    return {"success": True, "message": "Regola aggiunta", "action": "created", "id": regola["id"]}


@router.delete("/regole/{tipo}/{pattern}")
async def elimina_regola(tipo: str, pattern: str) -> Dict[str, Any]:
    """Elimina una regola."""
    db = Database.get_db()
    
    collection = "regole_categorizzazione_fornitori" if tipo == "fornitore" else "regole_categorizzazione_descrizioni"
    
    result = await db[collection].delete_one({"pattern": pattern})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Regola non trovata")
    
    return {"success": True, "message": "Regola eliminata"}


@router.post("/categorie")
async def aggiorna_categoria(data: Dict[str, Any]) -> Dict[str, Any]:
    """Aggiorna o crea una categoria con deducibilità."""
    db = Database.get_db()
    
    categoria = data.get("categoria", "").strip().lower().replace(" ", "_")
    conto = data.get("conto", "").strip()
    ded_ires = data.get("deducibilita_ires", 100)
    ded_irap = data.get("deducibilita_irap", 100)
    
    if not categoria or not conto:
        raise HTTPException(status_code=400, detail="Categoria e conto sono obbligatori")
    
    try:
        ded_ires = float(ded_ires)
        ded_irap = float(ded_irap)
    except (ValueError, TypeError):
        ded_ires = 100
        ded_irap = 100
    
    # Upsert categoria
    result = await db["regole_categorie"].update_one(
        {"categoria": categoria},
        {"$set": {
            "categoria": categoria,
            "conto": conto,
            "deducibilita_ires": min(100, max(0, ded_ires)),
            "deducibilita_irap": min(100, max(0, ded_irap)),
            "updated_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    
    action = "created" if result.upserted_id else "updated"
    return {"success": True, "message": f"Categoria {action}", "action": action}


# ============== FUNZIONI HELPER ==============

async def _get_default_regole_fornitori() -> List[Dict]:
    """Restituisce le regole fornitori di default."""
    from app.services.categorizzazione_contabile import PATTERNS_FORNITORE
    
    regole = []
    for pattern, categoria in PATTERNS_FORNITORE.items():
        # Converti regex in testo leggibile
        pattern_clean = pattern.replace("\\s*", " ").replace("\\s+", " ")
        pattern_clean = pattern_clean.replace("\\.", ".").replace("\\b", "")
        pattern_clean = pattern_clean.replace("|", " / ")
        
        regole.append({
            "pattern": pattern_clean,
            "categoria": categoria,
            "note": "Regola predefinita"
        })
    
    return regole


async def _get_default_regole_descrizioni() -> List[Dict]:
    """Restituisce le regole descrizioni di default (versione semplificata)."""
    # Estrai alcune regole principali
    regole_base = [
        ("limoncello", "bevande_alcoliche", "Liquori"),
        ("amaro", "bevande_alcoliche", "Liquori"),
        ("vino", "bevande_alcoliche", "Vini"),
        ("birra", "bevande_alcoliche", "Birra"),
        ("prosecco", "bevande_alcoliche", "Spumanti"),
        ("caffe", "caffe", "Caffè"),
        ("espresso", "caffe", "Caffè"),
        ("pasta", "alimentari", "Pasta"),
        ("formaggio", "alimentari", "Formaggi"),
        ("prosciutto", "alimentari", "Salumi"),
        ("cornetto", "pasticceria", "Prodotti forno"),
        ("brioche", "pasticceria", "Prodotti forno"),
        ("surgelato", "surgelati", "Surgelati"),
        ("congelato", "surgelati", "Surgelati"),
        ("manutenzione", "manutenzione", "Manutenzione"),
        ("riparazione", "manutenzione", "Riparazione"),
        ("benzina", "carburante", "Carburante"),
        ("gasolio", "carburante", "Carburante"),
        ("guanti", "pulizia", "DPI/Igiene"),
        ("detergente", "pulizia", "Pulizia"),
        ("noleggio auto", "noleggio_auto", "Noleggio veicoli"),
        ("canone finanziario", "noleggio_auto", "Leasing"),
        ("telefono", "telefonia", "Telefonia"),
        ("internet", "telefonia", "Internet"),
        ("assicurazione", "assicurazioni", "Assicurazioni"),
        ("polizza", "assicurazioni", "Assicurazioni"),
    ]
    
    return [{"pattern": p, "categoria": c, "note": n} for p, c, n in regole_base]
