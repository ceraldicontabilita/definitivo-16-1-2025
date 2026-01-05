"""
Riconciliazione Bonifici Fornitori
Abbina i movimenti bancari (bonifici) alle fatture di acquisto.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from typing import Dict, Any, List, Optional
from datetime import datetime, date
import uuid
import logging
import io
import re

from app.database import Database

logger = logging.getLogger(__name__)
router = APIRouter()

# Dizionario alias fornitori: nome_variante -> nome_standard
# Popolato dinamicamente + alias manuali per casi noti
FORNITORI_ALIAS = {
    # Aggiungi qui alias manuali per fornitori con nomi diversi
    "dolciaria acquaviva": "dolciaria acquaviva s.p.a.",
    "df baldassarre": "df baldassarre srl",
    "saima": "saima s.p.a.",
    "arval service lease italia": "arval service lease italia spa",
    "edenred italia": "edenred italia s.r.l.",
}


def normalizza_nome_fornitore(nome: str) -> str:
    """Normalizza il nome fornitore per il matching."""
    if not nome:
        return ""
    import unicodedata
    # Rimuovi accenti
    nfkd = unicodedata.normalize('NFKD', nome)
    nome_norm = ''.join([c for c in nfkd if not unicodedata.combining(c)])
    # Lowercase
    nome_norm = nome_norm.lower()
    # Rimuovi NOTPROVIDE e varianti
    nome_norm = re.sub(r'\s*notprovid\w*', '', nome_norm)
    # Rimuovi parti legali lunghe (es. "sog. all'att. di dir. e coord. di ...")
    nome_norm = re.sub(r'\s+sog\.?\s+all.*$', '', nome_norm)
    # Rimuovi forme societarie comuni
    for forma in ['s.r.l.', 'srl', 's.p.a.', 'spa', 's.n.c.', 'snc', 's.a.s.', 'sas', 
                  's.r.l', 's.p.a', 'n.p.', 'unipersonale']:
        nome_norm = nome_norm.replace(forma, '')
    # Rimuovi punteggiatura extra
    nome_norm = re.sub(r'[.,]+', ' ', nome_norm)
    return ' '.join(nome_norm.split())


def match_fornitori_fuzzy(nome1: str, nome2: str) -> bool:
    """Match fuzzy tra nomi fornitori."""
    if not nome1 or not nome2:
        return False
    
    n1 = normalizza_nome_fornitore(nome1)
    n2 = normalizza_nome_fornitore(nome2)
    
    if not n1 or not n2:
        return False
    
    # Match esatto
    if n1 == n2:
        return True
    
    # Controlla alias
    n1_canonical = FORNITORI_ALIAS.get(n1, n1)
    n2_canonical = FORNITORI_ALIAS.get(n2, n2)
    if n1_canonical == n2_canonical:
        return True
    # Controlla anche alias parziali
    for alias, canonical in FORNITORI_ALIAS.items():
        if alias in n1 or alias in n2:
            if normalizza_nome_fornitore(canonical) in [n1, n2]:
                return True
    
    # Uno contenuto nell'altro (per nomi con suffissi legali diversi)
    if len(n1) >= 6 and len(n2) >= 6:
        if n1 in n2 or n2 in n1:
            return True
    
    # Prima parola uguale (spesso è il nome distintivo)
    words1 = n1.split()
    words2 = n2.split()
    if words1 and words2:
        # Prima parola identica e significativa
        if words1[0] == words2[0] and len(words1[0]) >= 4:
            return True
        
        # Prime due parole identiche
        if len(words1) >= 2 and len(words2) >= 2:
            if words1[0] == words2[0] and words1[1] == words2[1]:
                return True
        
        # Match parole significative (>= 5 caratteri)
        sig_words1 = {w for w in words1 if len(w) >= 5}
        sig_words2 = {w for w in words2 if len(w) >= 5}
        common = sig_words1 & sig_words2
        if len(common) >= 1:
            return True
    
    return False


def estrai_fornitore_da_descrizione(descrizione: str) -> Optional[str]:
    """Estrae il nome fornitore dalla descrizione del bonifico."""
    if not descrizione:
        return None
    
    desc_upper = descrizione.upper()
    
    # Pattern: "FAVORE NomeFornitore"
    if "FAVORE" in desc_upper:
        idx = desc_upper.find("FAVORE")
        after = descrizione[idx + 7:].strip()
        
        # Prendi fino a " - " o fine stringa
        if " - " in after:
            nome = after.split(" - ")[0].strip()
        else:
            nome = after.strip()
        
        # Rimuovi "NOTPROVIDE", "NOTPROVIDED", etc.
        nome = re.sub(r'\s*NOTPROVID\w*', '', nome, flags=re.IGNORECASE)
        
        # Rimuovi codici alla fine (es. "FT 123", numeri)
        nome = re.sub(r'\s+FT\s*\d+.*$', '', nome, flags=re.IGNORECASE)
        nome = re.sub(r'\s+\d+\s*$', '', nome)
        
        # Limita a max 6 parole (nomi aziendali lunghi)
        words = nome.split()
        if len(words) > 6:
            nome = " ".join(words[:6])
        
        return nome.strip() if nome.strip() else None
    
    # Pattern: SDD con nome alla fine (addebito diretto)
    if "SDD" in desc_upper:
        # Cerca nome dopo il codice
        parts = descrizione.split()
        # Nome di solito è dopo i numeri
        nome_parts = []
        for p in parts:
            if not re.match(r'^[\d:]+$', p) and len(p) > 2 and p.upper() not in ['SDD', 'B2B']:
                nome_parts.append(p)
        if nome_parts:
            return " ".join(nome_parts[-4:])  # Ultime 4 parole significative
    
    return None


@router.post("/import-estratto-conto-fornitori")
async def import_estratto_conto_fornitori(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa estratto conto bancario e riconcilia con le fatture fornitori.
    
    Cerca le fatture di acquisto (passive) non pagate e le abbina
    ai bonifici con stesso importo e nome fornitore.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    contents = await file.read()
    
    movimenti_banca = []
    
    if filename.endswith('.csv'):
        # Parse CSV
        import csv
        text = contents.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
        
        for row in reader:
            categoria = row.get('Categoria/sottocategoria', '') or row.get('Categoria', '')
            descrizione = row.get('Descrizione', '')
            
            # Filtra solo bonifici fornitori (esclude salari)
            is_fornitore = (
                'fornitori' in categoria.lower() or 
                'materie prime' in categoria.lower() or
                'servizi' in categoria.lower() or
                'utenze' in categoria.lower()
            )
            
            # Escludi salari
            if 'salari' in categoria.lower() or 'stipend' in categoria.lower():
                continue
            
            if not is_fornitore:
                continue
            
            # Parse data
            data_str = row.get('Data contabile') or row.get('Data valuta') or row.get('Data')
            if not data_str:
                continue
            
            # Parse importo
            importo_str = row.get('Importo', '0')
            importo_str = importo_str.replace('.', '').replace(',', '.')
            try:
                importo = abs(float(importo_str))
            except:
                continue
            
            # Parse data (DD/MM/YYYY)
            try:
                if '/' in data_str:
                    parts = data_str.split('/')
                    data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    data_obj = date.fromisoformat(data_str)
            except:
                continue
            
            # Estrai nome fornitore
            nome_fornitore = estrai_fornitore_da_descrizione(descrizione)
            
            movimenti_banca.append({
                "data": data_obj,
                "importo": importo,
                "descrizione": descrizione,
                "nome_fornitore": nome_fornitore,
                "categoria": categoria
            })
    
    elif filename.endswith(('.xlsx', '.xls')):
        # Parse Excel
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            
            headers = [str(cell.value or '').lower() for cell in sheet[1]]
            
            for row_num in range(2, sheet.max_row + 1):
                row_data = [sheet.cell(row=row_num, column=i+1).value for i in range(len(headers))]
                
                importo = None
                data_val = None
                descrizione = ""
                categoria = ""
                
                for i, h in enumerate(headers):
                    val = row_data[i]
                    if 'importo' in h and val:
                        if isinstance(val, (int, float)):
                            importo = abs(val)
                        else:
                            try:
                                importo = abs(float(str(val).replace('.', '').replace(',', '.')))
                            except:
                                pass
                    elif 'data' in h and val:
                        if isinstance(val, datetime):
                            data_val = val.date()
                        elif isinstance(val, date):
                            data_val = val
                    elif 'descri' in h and val:
                        descrizione = str(val)
                    elif 'categoria' in h and val:
                        categoria = str(val)
                
                if importo and data_val:
                    # Filtra fornitori
                    if 'salari' in categoria.lower():
                        continue
                    
                    movimenti_banca.append({
                        "data": data_val,
                        "importo": importo,
                        "descrizione": descrizione,
                        "nome_fornitore": estrai_fornitore_da_descrizione(descrizione),
                        "categoria": categoria
                    })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa CSV o Excel.")
    
    # Carica fatture non pagate
    fatture = await db["invoices"].find(
        {"$or": [{"pagato": False}, {"pagato": {"$exists": False}}]},
        {"_id": 0}
    ).to_list(10000)
    
    # Crea indice per matching: (nome_normalizzato, importo) -> fatture
    fatture_index = {}
    for f in fatture:
        importo = f.get("total_amount", 0)
        nome = f.get("supplier_name", "")
        if importo <= 0 or not nome:
            continue
        key = (normalizza_nome_fornitore(nome), round(importo, 2))
        if key not in fatture_index:
            fatture_index[key] = []
        fatture_index[key].append(f)
    
    # Riconcilia
    riconciliati = 0
    non_trovati = []
    
    for mov in movimenti_banca:
        nome_banca = mov.get("nome_fornitore")
        if not nome_banca:
            non_trovati.append({
                "data": mov["data"].isoformat(),
                "importo": mov["importo"],
                "descrizione": mov["descrizione"][:100],
                "nome": "N/D - nome non estratto"
            })
            continue
        
        nome_norm = normalizza_nome_fornitore(nome_banca)
        importo = round(mov["importo"], 2)
        data_mov = mov["data"]
        
        found = False
        best_match = None
        best_match_score = -1
        
        for (key_nome, key_importo), fatture_list in fatture_index.items():
            # Tolleranza importo: 1% o 5 euro
            importo_diff = abs(key_importo - importo)
            if importo_diff > max(importo * 0.01, 5):
                continue
            
            # Match nome fuzzy
            if not match_fornitori_fuzzy(key_nome, nome_norm):
                continue
            
            # Cerca la fattura con data più vicina
            for fattura in fatture_list:
                if fattura.get("pagato"):
                    continue
                
                # Calcola score basato sulla vicinanza temporale
                data_fattura_str = fattura.get("invoice_date") or fattura.get("data_ricezione")
                if data_fattura_str:
                    try:
                        data_fattura = date.fromisoformat(data_fattura_str[:10])
                        giorni_diff = abs((data_mov - data_fattura).days)
                        
                        # Score basato su vicinanza
                        if giorni_diff <= 30:
                            score = 100 - giorni_diff
                        elif giorni_diff <= 90:
                            score = 50
                        else:
                            score = 10
                    except:
                        score = 50
                else:
                    score = 50
                
                # Bonus per match importo esatto
                if importo_diff < 0.01:
                    score += 20
                
                if score > best_match_score:
                    best_match_score = score
                    best_match = fattura
        
        if best_match and best_match_score >= 10:
            # Aggiorna fattura come pagata
            await db["invoices"].update_one(
                {"id": best_match["id"]},
                {"$set": {
                    "pagato": True,
                    "data_pagamento": data_mov.isoformat(),
                    "metodo_pagamento": "bonifico",
                    "riconciliato_da_estratto": True,
                    "riferimento_banca": mov["descrizione"][:200],
                    "updated_at": datetime.utcnow().isoformat()
                }}
            )
            best_match["pagato"] = True
            riconciliati += 1
            found = True
        
        if not found:
            non_trovati.append({
                "data": mov["data"].isoformat(),
                "importo": mov["importo"],
                "descrizione": mov["descrizione"][:100],
                "nome": nome_banca
            })
    
    # Salva movimenti banca per riferimento
    for mov in movimenti_banca:
        mov_record = {
            "id": f"ECF-{mov['data'].isoformat()}-{mov['importo']:.2f}-{uuid.uuid4().hex[:6]}",
            "data": mov["data"].isoformat(),
            "importo": mov["importo"],
            "descrizione": mov["descrizione"][:200],
            "nome_fornitore": mov.get("nome_fornitore", ""),
            "categoria": mov.get("categoria", ""),
            "tipo": "fornitore",
            "imported_at": datetime.utcnow().isoformat()
        }
        
        await db["estratto_conto_fornitori"].update_one(
            {"data": mov_record["data"], "importo": mov_record["importo"], "descrizione": mov_record["descrizione"][:50]},
            {"$set": mov_record},
            upsert=True
        )
    
    return {
        "message": "Importazione e riconciliazione fornitori completate",
        "movimenti_banca": len(movimenti_banca),
        "riconciliati": riconciliati,
        "non_trovati": len(non_trovati),
        "dettaglio_non_trovati": non_trovati[:20]
    }


@router.get("/fatture-non-pagate")
async def get_fatture_non_pagate(
    fornitore: Optional[str] = Query(None),
    limit: int = Query(100, le=1000)
) -> List[Dict[str, Any]]:
    """Lista fatture non pagate (per riconciliazione manuale)."""
    db = Database.get_db()
    
    query = {"$or": [{"pagato": False}, {"pagato": {"$exists": False}}]}
    if fornitore:
        query["supplier_name"] = {"$regex": fornitore, "$options": "i"}
    
    fatture = await db["invoices"].find(
        query,
        {"_id": 0, "id": 1, "invoice_number": 1, "invoice_date": 1, "supplier_name": 1, 
         "total_amount": 1, "pagato": 1, "data_pagamento": 1}
    ).sort("invoice_date", -1).limit(limit).to_list(limit)
    
    return fatture


@router.get("/riepilogo-fornitori")
async def get_riepilogo_fornitori() -> Dict[str, Any]:
    """Riepilogo pagamenti fornitori."""
    db = Database.get_db()
    
    total = await db["invoices"].count_documents({})
    pagate = await db["invoices"].count_documents({"pagato": True})
    non_pagate = await db["invoices"].count_documents(
        {"$or": [{"pagato": False}, {"pagato": {"$exists": False}}]}
    )
    
    # Totale importi
    pipeline = [
        {"$group": {
            "_id": "$pagato",
            "totale": {"$sum": "$total_amount"},
            "count": {"$sum": 1}
        }}
    ]
    by_stato = await db["invoices"].aggregate(pipeline).to_list(10)
    
    totale_pagate = sum(r["totale"] for r in by_stato if r["_id"] == True)
    totale_non_pagate = sum(r["totale"] for r in by_stato if r["_id"] != True)
    
    return {
        "totale_fatture": total,
        "fatture_pagate": pagate,
        "fatture_non_pagate": non_pagate,
        "importo_pagato": round(totale_pagate, 2),
        "importo_da_pagare": round(totale_non_pagate, 2)
    }


@router.delete("/reset-riconciliazione-fornitori")
async def reset_riconciliazione_fornitori(
    fornitore: Optional[str] = Query(None)
) -> Dict[str, Any]:
    """Reset riconciliazione fatture fornitori."""
    db = Database.get_db()
    
    query = {"riconciliato_da_estratto": True}
    if fornitore:
        query["supplier_name"] = {"$regex": fornitore, "$options": "i"}
    
    result = await db["invoices"].update_many(
        query,
        {"$unset": {
            "pagato": "",
            "data_pagamento": "",
            "riconciliato_da_estratto": "",
            "riferimento_banca": ""
        }}
    )
    
    # Pulisci anche estratto_conto_fornitori
    ec_deleted = await db["estratto_conto_fornitori"].delete_many({})
    
    return {
        "message": "Reset riconciliazione fornitori completato",
        "fatture_resettate": result.modified_count,
        "estratti_conto_eliminati": ec_deleted.deleted_count
    }
