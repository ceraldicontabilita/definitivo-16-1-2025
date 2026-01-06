"""
Gestione Estratto Conto
Salva e visualizza tutti i movimenti bancari importati con campi strutturati.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from typing import Dict, Any, List, Optional
from datetime import datetime, date
import uuid
import logging
import io
import re
import csv

from app.database import Database

logger = logging.getLogger(__name__)
router = APIRouter()


def estrai_numero_fattura(descrizione: str) -> Optional[str]:
    """Estrae il numero/i fattura dalla descrizione dopo NOTPROVIDE."""
    if not descrizione:
        return None
    
    # Pattern: dopo "NOTPROVIDE - " c'è il riferimento fatture
    match = re.search(r'NOTPROVIDE\s*-?\s*(.+)$', descrizione, re.IGNORECASE)
    if match:
        riferimento = match.group(1).strip()
        # Pulisci e restituisci
        # Rimuovi prefissi comuni
        riferimento = re.sub(r'^(saldo|pagamento)\s+(fattur[ae]|ft)\s*', '', riferimento, flags=re.IGNORECASE)
        return riferimento.strip()[:200] if riferimento.strip() else None
    
    # Pattern alternativo: "fattura/e" seguito da numeri
    match = re.search(r'fattur[ae]\s+(.+?)(?:\s*$|\s+-)', descrizione, re.IGNORECASE)
    if match:
        return match.group(1).strip()[:200]
    
    return None


def estrai_fornitore_pulito(descrizione: str) -> Optional[str]:
    """Estrae il nome fornitore dalla descrizione, pulendolo."""
    if not descrizione:
        return None
    
    desc_upper = descrizione.upper()
    
    if "FAVORE" in desc_upper:
        idx = desc_upper.find("FAVORE")
        after = descrizione[idx + 7:].strip()
        
        # Prendi fino a "NOTPROVIDE" o " - " o fine
        for sep in ["NOTPROVIDE", " - ADD.", " - "]:
            if sep.upper() in after.upper():
                idx_sep = after.upper().find(sep.upper())
                after = after[:idx_sep].strip()
                break
        
        # Rimuovi forme societarie alla fine per pulizia display
        # Ma mantienile se sono parte del nome
        nome = after.strip()
        
        return nome if nome else None
    
    return None


@router.post("/import")
async def import_estratto_conto(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa estratto conto bancario e salva tutti i movimenti con campi strutturati:
    - data: data contabile
    - fornitore: nome estratto dalla descrizione
    - importo: importo del movimento
    - numero_fattura: numero fatture pagate (se presente dopo NOTPROVIDE)
    - data_pagamento: data valuta
    - categoria: categoria del movimento
    - descrizione_originale: descrizione completa originale
    
    Evita duplicati controllando data + importo + descrizione.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    contents = await file.read()
    
    movimenti = []
    
    if filename.endswith('.csv'):
        text = contents.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
        
        for row in reader:
            # Estrai dati
            data_contabile = row.get('Data contabile', '')
            data_valuta = row.get('Data valuta', '')
            importo_str = row.get('Importo', '0')
            descrizione = row.get('Descrizione', '')
            categoria = row.get('Categoria/sottocategoria', '') or row.get('Categoria', '')
            
            # Parse importo
            importo_str = importo_str.replace('.', '').replace(',', '.')
            try:
                importo = float(importo_str)
            except (ValueError, TypeError):
                continue
            
            # Parse data contabile (DD/MM/YYYY)
            try:
                if '/' in data_contabile:
                    parts = data_contabile.split('/')
                    data_obj = date(int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    continue
            except (ValueError, TypeError, IndexError):
                continue
            
            # Parse data valuta
            data_pagamento = None
            try:
                if '/' in data_valuta:
                    parts = data_valuta.split('/')
                    data_pagamento = date(int(parts[2]), int(parts[1]), int(parts[0]))
            except (ValueError, TypeError, IndexError):
                pass
            
            # Estrai fornitore/beneficiario
            fornitore = estrai_fornitore_pulito(descrizione)
            
            # Estrai numero fattura
            numero_fattura = estrai_numero_fattura(descrizione)
            
            movimenti.append({
                "data": data_obj,
                "fornitore": fornitore,
                "importo": importo,
                "numero_fattura": numero_fattura,
                "data_pagamento": data_pagamento,
                "categoria": categoria,
                "descrizione_originale": descrizione,
                "tipo": "uscita" if importo < 0 else "entrata"
            })
    
    elif filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = wb.active
            
            headers = [str(cell.value or '').lower() for cell in sheet[1]]
            
            for row_num in range(2, sheet.max_row + 1):
                row_data = {headers[i]: sheet.cell(row=row_num, column=i+1).value 
                           for i in range(len(headers))}
                
                # Trova colonne
                data_contabile = None
                importo = None
                descrizione = ""
                categoria = ""
                data_valuta = None
                
                for h, v in row_data.items():
                    if not v:
                        continue
                    if 'data contabile' in h or (h == 'data' and not data_contabile):
                        if isinstance(v, (datetime, date)):
                            data_contabile = v if isinstance(v, date) else v.date()
                        elif '/' in str(v):
                            parts = str(v).split('/')
                            try:
                                data_contabile = date(int(parts[2]), int(parts[1]), int(parts[0]))
                            except:
                                pass
                    elif 'data valuta' in h:
                        if isinstance(v, (datetime, date)):
                            data_valuta = v if isinstance(v, date) else v.date()
                    elif 'importo' in h:
                        if isinstance(v, (int, float)):
                            importo = float(v)
                        else:
                            try:
                                importo = float(str(v).replace('.', '').replace(',', '.'))
                            except:
                                pass
                    elif 'descri' in h:
                        descrizione = str(v)
                    elif 'categoria' in h:
                        categoria = str(v)
                
                if data_contabile and importo is not None:
                    movimenti.append({
                        "data": data_contabile,
                        "fornitore": estrai_fornitore_pulito(descrizione),
                        "importo": importo,
                        "numero_fattura": estrai_numero_fattura(descrizione),
                        "data_pagamento": data_valuta,
                        "categoria": categoria,
                        "descrizione_originale": descrizione,
                        "tipo": "uscita" if importo < 0 else "entrata"
                    })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa CSV o Excel.")
    
    # Salva nel database, evitando duplicati
    inserted = 0
    duplicates = 0
    
    for mov in movimenti:
        # Crea ID univoco basato su data + importo + primi 50 char descrizione
        desc_hash = (mov["descrizione_originale"] or "")[:50]
        mov_id = f"EC-{mov['data'].isoformat()}-{mov['importo']:.2f}-{hash(desc_hash) % 100000:05d}"
        
        # Controlla duplicati
        existing = await db["estratto_conto_movimenti"].find_one({
            "data": mov["data"].isoformat(),
            "importo": mov["importo"],
            "descrizione_hash": desc_hash
        })
        
        if existing:
            duplicates += 1
            continue
        
        # Salva
        record = {
            "id": mov_id,
            "data": mov["data"].isoformat(),
            "fornitore": mov["fornitore"],
            "importo": mov["importo"],
            "numero_fattura": mov["numero_fattura"],
            "data_pagamento": mov["data_pagamento"].isoformat() if mov["data_pagamento"] else None,
            "categoria": mov["categoria"],
            "descrizione_originale": mov["descrizione_originale"],
            "tipo": mov["tipo"],
            "descrizione_hash": desc_hash,
            "created_at": datetime.utcnow().isoformat()
        }
        
        await db["estratto_conto_movimenti"].insert_one(record)
        inserted += 1
    
    return {
        "message": "Importazione estratto conto completata",
        "movimenti_trovati": len(movimenti),
        "inseriti": inserted,
        "duplicati_saltati": duplicates
    }


@router.get("/movimenti")
async def get_movimenti(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),  # "entrata" | "uscita"
    limit: int = Query(500, le=5000),
    offset: int = Query(0)
) -> Dict[str, Any]:
    """
    Recupera i movimenti dell'estratto conto con filtri.
    Ordinati per data decrescente.
    """
    db = Database.get_db()
    
    query = {}
    
    if anno:
        # Filtra per anno nella data (formato ISO: YYYY-MM-DD)
        query["data"] = {"$regex": f"^{anno}"}
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    # Count totale
    total = await db["estratto_conto_movimenti"].count_documents(query)
    
    # Recupera movimenti
    movimenti = await db["estratto_conto_movimenti"].find(
        query,
        {"_id": 0}
    ).sort("data", -1).skip(offset).limit(limit).to_list(limit)
    
    # Calcola totali
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None,
            "totale_entrate": {"$sum": {"$cond": [{"$gt": ["$importo", 0]}, "$importo", 0]}},
            "totale_uscite": {"$sum": {"$cond": [{"$lt": ["$importo", 0]}, {"$abs": "$importo"}, 0]}}
        }}
    ]
    totali_result = await db["estratto_conto_movimenti"].aggregate(pipeline).to_list(1)
    totali = totali_result[0] if totali_result else {"totale_entrate": 0, "totale_uscite": 0}
    
    return {
        "movimenti": movimenti,
        "totale": total,
        "offset": offset,
        "limit": limit,
        "totale_entrate": round(totali.get("totale_entrate", 0), 2),
        "totale_uscite": round(totali.get("totale_uscite", 0), 2)
    }


@router.get("/categorie")
async def get_categorie() -> List[str]:
    """Restituisce lista categorie uniche."""
    db = Database.get_db()
    categorie = await db["estratto_conto_movimenti"].distinct("categoria")
    return sorted([c for c in categorie if c])


@router.get("/fornitori")
async def get_fornitori_unici() -> List[str]:
    """Restituisce lista fornitori unici."""
    db = Database.get_db()
    fornitori = await db["estratto_conto_movimenti"].distinct("fornitore")
    return sorted([f for f in fornitori if f])


@router.get("/riepilogo")
async def get_riepilogo(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None)
) -> Dict[str, Any]:
    """Riepilogo estratto conto con filtri."""
    db = Database.get_db()
    
    query = {}
    if anno:
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
        else:
            query["data"] = {"$regex": f"^{anno}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    total = await db["estratto_conto_movimenti"].count_documents(query)
    
    # Totali per tipo
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$tipo",
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }}
    ]
    by_tipo = await db["estratto_conto_movimenti"].aggregate(pipeline).to_list(10)
    
    entrate = next((t for t in by_tipo if t["_id"] == "entrata"), {"totale": 0, "count": 0})
    uscite = next((t for t in by_tipo if t["_id"] == "uscita"), {"totale": 0, "count": 0})
    
    # Movimenti per categoria (top 10)
    pipeline_cat = [
        {"$match": query},
        {"$group": {
            "_id": "$categoria",
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"totale": -1}},
        {"$limit": 10}
    ]
    by_categoria = await db["estratto_conto_movimenti"].aggregate(pipeline_cat).to_list(10)
    
    return {
        "totale_movimenti": total,
        "entrate": {"count": entrate["count"], "totale": round(entrate["totale"], 2)},
        "uscite": {"count": uscite["count"], "totale": round(uscite["totale"], 2)},
        "saldo": round(entrate["totale"] - uscite["totale"], 2),
        "per_categoria": [{"categoria": c["_id"] or "N/D", "totale": round(c["totale"], 2), "count": c["count"]} for c in by_categoria]
    }


@router.delete("/clear")
async def clear_estratto_conto(anno: Optional[int] = Query(None)) -> Dict[str, Any]:
    """Elimina movimenti estratto conto."""
    db = Database.get_db()
    
    query = {}
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    result = await db["estratto_conto_movimenti"].delete_many(query)
    
    return {"message": f"Eliminati {result.deleted_count} movimenti"}


@router.get("/export-excel")
async def export_estratto_conto_excel(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    categoria: Optional[str] = Query(None),
    fornitore: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None)
):
    """
    Esporta i movimenti dell'estratto conto in formato Excel.
    Applica gli stessi filtri della visualizzazione.
    """
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    db = Database.get_db()
    
    # Costruisci query con filtri
    query = {}
    
    if anno:
        query["data"] = {"$regex": f"^{anno}"}
        if mese:
            query["data"] = {"$regex": f"^{anno}-{mese:02d}"}
    
    if categoria:
        query["categoria"] = {"$regex": categoria, "$options": "i"}
    
    if fornitore:
        query["fornitore"] = {"$regex": fornitore, "$options": "i"}
    
    if tipo:
        query["tipo"] = tipo
    
    # Recupera tutti i movimenti (senza paginazione per export)
    movimenti = await db["estratto_conto_movimenti"].find(
        query,
        {"_id": 0}
    ).sort("data", -1).to_list(10000)
    
    # Crea workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estratto Conto"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Data", "Fornitore", "Importo (€)", "Tipo", "N. Fattura", "Data Pag.", "Categoria"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Larghezze colonne
    col_widths = [12, 35, 15, 10, 30, 12, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Dati
    entrata_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    uscita_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    totale_entrate = 0
    totale_uscite = 0
    
    for row_num, mov in enumerate(movimenti, 2):
        # Formatta data
        data_str = mov.get("data", "")
        if data_str:
            try:
                parts = data_str.split("-")
                data_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except:
                data_formatted = data_str
        else:
            data_formatted = ""
        
        data_pag = mov.get("data_pagamento", "")
        if data_pag:
            try:
                parts = data_pag.split("-")
                data_pag_formatted = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except:
                data_pag_formatted = data_pag
        else:
            data_pag_formatted = ""
        
        importo = mov.get("importo", 0)
        tipo_mov = "Entrata" if importo >= 0 else "Uscita"
        
        if importo >= 0:
            totale_entrate += importo
        else:
            totale_uscite += abs(importo)
        
        row_data = [
            data_formatted,
            mov.get("fornitore") or "",
            abs(importo),
            tipo_mov,
            mov.get("numero_fattura") or "",
            data_pag_formatted,
            mov.get("categoria") or ""
        ]
        
        row_fill = entrata_fill if importo >= 0 else uscita_fill
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = row_fill
            
            # Formato numerico per importo
            if col == 3:
                cell.number_format = '#,##0.00'
    
    # Riga totali
    last_row = len(movimenti) + 2
    totals_row = last_row + 1
    
    ws.cell(row=totals_row, column=1, value="TOTALI")
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    
    ws.cell(row=totals_row, column=2, value=f"Entrate: € {totale_entrate:,.2f}")
    ws.cell(row=totals_row, column=2).font = Font(bold=True, color="16A34A")
    
    ws.cell(row=totals_row, column=3, value=f"Uscite: € {totale_uscite:,.2f}")
    ws.cell(row=totals_row, column=3).font = Font(bold=True, color="DC2626")
    
    saldo = totale_entrate - totale_uscite
    ws.cell(row=totals_row, column=4, value=f"Saldo: € {saldo:,.2f}")
    ws.cell(row=totals_row, column=4).font = Font(bold=True, color="16A34A" if saldo >= 0 else "DC2626")
    
    # Salva in memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome file
    filename_parts = ["estratto_conto"]
    if anno:
        filename_parts.append(str(anno))
    if mese:
        mesi_nomi = ['gen', 'feb', 'mar', 'apr', 'mag', 'giu', 'lug', 'ago', 'set', 'ott', 'nov', 'dic']
        filename_parts.append(mesi_nomi[mese - 1])
    filename = "_".join(filename_parts) + ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
