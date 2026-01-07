"""
Prima Nota Salari - Nuovo sistema di gestione stipendi.

Struttura dati:
- Dipendente
- Mese di competenza
- Anno
- Importo Busta (da file paghe.xlsx)
- Importo Bonifico (da file bonifici.xlsx)
- Saldo (differenza)
- Progressivo (riporto da mesi precedenti)
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from typing import Dict, Any, List, Optional
from datetime import datetime
import uuid
import logging
import io

from app.database import Database

logger = logging.getLogger(__name__)
router = APIRouter()

# Mapping mesi italiano -> numero (include anche numeri come stringhe)
MESI_MAP = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
    "gen": 1, "feb": 2, "mar": 3, "apr": 4, "mag": 5, "giu": 6,
    "lug": 7, "ago": 8, "set": 9, "ott": 10, "nov": 11, "dic": 12,
    "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6,
    "7": 7, "8": 8, "9": 9, "10": 10, "11": 11, "12": 12,
    "01": 1, "02": 2, "03": 3, "04": 4, "05": 5, "06": 6,
    "07": 7, "08": 8, "09": 9
}

MESI_NOMI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
             "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]


def normalize_name(name: str) -> str:
    """Normalizza nome dipendente per matching."""
    if not name:
        return ""
    return " ".join(name.strip().upper().split())


def get_mese_numero(mese_str: str) -> int:
    """Converte nome mese o numero in numero."""
    if not mese_str:
        return 0
    # Prima prova a convertire direttamente in intero
    try:
        mese = int(float(mese_str))
        if 1 <= mese <= 12:
            return mese
    except (ValueError, TypeError):
        pass
    # Altrimenti cerca nel mapping
    return MESI_MAP.get(str(mese_str).lower().strip(), 0)


# ============== ENDPOINTS ==============

@router.get("/salari")
async def get_prima_nota_salari(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None),
    dipendente: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """
    Recupera la prima nota salari con tutti i campi:
    - dipendente, mese, anno
    - importo_busta, importo_bonifico, saldo, progressivo
    - riconciliato
    """
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    if dipendente:
        query["dipendente"] = {"$regex": dipendente, "$options": "i"}
    
    salari = await db["prima_nota_salari"].find(
        query, {"_id": 0}
    ).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)
    
    return salari


@router.get("/salari/riepilogo")
async def get_riepilogo_salari(
    anno: int = Query(...),
    mese: Optional[int] = Query(None)
) -> Dict[str, Any]:
    """Riepilogo statistiche salari."""
    db = Database.get_db()
    
    query = {"anno": anno}
    if mese:
        query["mese"] = mese
    
    salari = await db["prima_nota_salari"].find(query, {"_id": 0}).to_list(5000)
    
    totale_buste = sum(s.get("importo_busta", 0) or 0 for s in salari)
    totale_bonifici = sum(s.get("importo_bonifico", 0) or 0 for s in salari)
    totale_saldo = sum(s.get("saldo", 0) or 0 for s in salari)
    
    # Conta dipendenti unici
    dipendenti_unici = len(set(s.get("dipendente", "") for s in salari))
    
    # Conta riconciliati
    riconciliati = sum(1 for s in salari if s.get("riconciliato"))
    
    return {
        "anno": anno,
        "mese": mese,
        "totale_records": len(salari),
        "dipendenti_unici": dipendenti_unici,
        "totale_buste": round(totale_buste, 2),
        "totale_bonifici": round(totale_bonifici, 2),
        "totale_saldo": round(totale_saldo, 2),
        "riconciliati": riconciliati,
        "da_riconciliare": len(salari) - riconciliati
    }


@router.post("/import-paghe")
async def import_paghe(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa file PAGHE (stipendi netti).
    Formato atteso: Dipendente | Mese | Anno | Stipendio Netto
    
    REGOLA: MAI AGGREGARE - Crea un record per ogni riga del file.
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {str(e)}")
    
    # Normalizza nomi colonne
    df.columns = [c.strip().lower() for c in df.columns]
    
    # Mapping colonne
    col_dipendente = None
    col_mese = None
    col_anno = None
    col_importo = None
    
    for c in df.columns:
        if 'dipendente' in c or 'nome' in c or 'cognome' in c:
            col_dipendente = c
        elif 'mese' in c:
            col_mese = c
        elif 'anno' in c:
            col_anno = c
        elif any(x in c for x in ['stipendio', 'netto', 'busta', 'paga', 'retribuzione']):
            col_importo = c
    
    # Fallback
    if not col_importo:
        for c in df.columns:
            if 'importo' in c and 'bonifico' not in c and 'erogato' not in c:
                col_importo = c
                break
    
    if not all([col_dipendente, col_mese, col_anno, col_importo]):
        raise HTTPException(
            status_code=400, 
            detail=f"Colonne richieste non trovate. Trovate: {list(df.columns)}"
        )
    
    logger.info(f"IMPORT PAGHE - Righe nel file: {len(df)}")
    
    created = 0
    errors = []
    
    # CREA UN RECORD PER OGNI RIGA - MAI AGGREGARE
    for idx, row in df.iterrows():
        try:
            dipendente = normalize_name(str(row[col_dipendente]))
            if not dipendente or dipendente == "NAN":
                continue
            
            # Gestisci mese (può essere numero o data)
            mese_val = row[col_mese]
            if hasattr(mese_val, 'month'):
                mese = mese_val.month
            else:
                mese_str = str(mese_val).strip()
                mese = get_mese_numero(mese_str)
            if mese == 0:
                continue
            
            # Gestisci anno (può essere numero o data)
            anno_val = row[col_anno]
            if hasattr(anno_val, 'year'):
                anno = anno_val.year
            elif isinstance(anno_val, datetime):
                anno = anno_val.year
            else:
                anno = int(anno_val)
            
            importo = float(row[col_importo]) if pd.notna(row[col_importo]) else 0
            
            # Crea nuovo record per OGNI riga
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
                "importo_busta": round(importo, 2),
                "importo_bonifico": 0,
                "saldo": round(-importo, 2),
                "progressivo": 0,
                "riconciliato": False,
                "tipo": "busta",
                "created_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat()
            }
            await db["prima_nota_salari"].insert_one(new_record)
            created += 1
            
        except Exception as e:
            errors.append(f"Riga {idx + 2}: {str(e)}")
    
    logger.info(f"IMPORT PAGHE - Record creati: {created}")
    
    # Ricalcola progressivi
    await ricalcola_progressivi_tutti(db)
    
    return {
        "success": True,
        "message": "Import PAGHE completato",
        "created": created,
        "updated": 0,
        "righe_file": len(df),
        "errors": errors[:10] if errors else []
    }


@router.post("/import-bonifici")
async def import_bonifici(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Importa file BONIFICI (importi erogati).
    Formati supportati:
    - Dipendente | Mese | Anno | Importo Erogato
    - Dipendente | Mese (data completa) | Importo Erogato
    
    REGOLA: MAI AGGREGARE - Crea un record per ogni riga del file.
    """
    import pandas as pd
    
    db = Database.get_db()
    content = await file.read()
    
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore lettura Excel: {str(e)}")
    
    # Normalizza nomi colonne (rimuovi spazi extra)
    df.columns = [c.strip().lower() for c in df.columns]
    
    logger.info(f"IMPORT BONIFICI - Colonne trovate: {list(df.columns)}")
    
    # Mapping colonne
    col_dipendente = None
    col_mese = None
    col_anno = None
    col_importo = None
    
    for c in df.columns:
        if 'dipendente' in c or 'nome' in c or 'cognome' in c:
            col_dipendente = c
        elif 'mese' in c or 'data' in c:
            col_mese = c
        elif 'anno' in c:
            col_anno = c
        elif any(x in c for x in ['erogato', 'bonifico', 'pagato', 'versato', 'accredito', 'importo']):
            col_importo = c
    
    if not all([col_dipendente, col_mese, col_importo]):
        raise HTTPException(
            status_code=400, 
            detail=f"Colonne richieste non trovate. Trovate: {list(df.columns)}. Serve: dipendente, mese/data, importo"
        )
    
    logger.info(f"IMPORT BONIFICI - Righe nel file: {len(df)}")
    logger.info(f"IMPORT BONIFICI - Mapping: dip={col_dipendente}, mese={col_mese}, anno={col_anno}, importo={col_importo}")
    
    created = 0
    errors = []
    
    # CREA UN RECORD PER OGNI RIGA - MAI AGGREGARE
    for idx, row in df.iterrows():
        try:
            dipendente = normalize_name(str(row[col_dipendente]))
            if not dipendente or dipendente == "NAN":
                continue
            
            # Gestisci mese e anno
            mese_val = row[col_mese]
            
            # Se è una data completa, estrai mese e anno
            if hasattr(mese_val, 'month') and hasattr(mese_val, 'year'):
                mese = mese_val.month
                anno = mese_val.year
            elif isinstance(mese_val, str) and '-' in mese_val:
                # Formato "2021-06-21" o simile
                try:
                    from dateutil import parser
                    dt = parser.parse(mese_val)
                    mese = dt.month
                    anno = dt.year
                except:
                    mese = get_mese_numero(str(mese_val))
                    anno = int(row[col_anno]) if col_anno else 2024
            else:
                mese_str = str(mese_val).strip()
                mese = get_mese_numero(mese_str)
                # Anno da colonna separata o default
                if col_anno:
                    anno_val = row[col_anno]
                    if hasattr(anno_val, 'year'):
                        anno = anno_val.year
                    else:
                        anno = int(anno_val)
                else:
                    anno = 2024  # Default
            
            if mese == 0 or mese > 12:
                errors.append(f"Riga {idx + 2}: mese non valido ({mese_val})")
                continue
            
            importo = float(row[col_importo]) if pd.notna(row[col_importo]) else 0
            
            # Crea nuovo record per OGNI riga
            new_record = {
                "id": str(uuid.uuid4()),
                "dipendente": dipendente,
                "anno": anno,
                "mese": mese,
                "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
                "importo_busta": 0,
                "importo_bonifico": round(importo, 2),
                "saldo": round(importo, 2),
                "progressivo": 0,
                "riconciliato": False,
                "tipo": "bonifico",
                "created_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat()
            }
            await db["prima_nota_salari"].insert_one(new_record)
            created += 1
            
        except Exception as e:
            errors.append(f"Riga {idx + 2}: {str(e)}")
    
    logger.info(f"IMPORT BONIFICI - Record creati: {created}")
    
    # Ricalcola progressivi
    await ricalcola_progressivi_tutti(db)
    
    return {
        "success": True,
        "message": "Import BONIFICI completato",
        "created": created,
        "updated": 0,
        "righe_file": len(df),
        "errors": errors[:10] if errors else []
    }


async def ricalcola_progressivi_tutti(db, anno_inizio: int = None, dipendente_filtro: str = None, anni_esclusi: List[int] = None):
    """
    Ricalcola saldi e progressivi per tutti i dipendenti o uno specifico.
    
    Per ogni record:
    - Saldo = importo_bonifico - importo_busta (della singola riga)
    - Progressivo = Somma cumulativa di tutti i saldi precedenti + saldo corrente
    
    I record sono ordinati per anno, mese e data di creazione.
    Gli anni in anni_esclusi vengono saltati nel calcolo del progressivo.
    I record con vincolo=True non vengono modificati.
    """
    if anni_esclusi is None:
        anni_esclusi = []
    
    # Ottieni dipendenti da processare
    if dipendente_filtro:
        dipendenti = [dipendente_filtro]
    else:
        dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    
    for dipendente in dipendenti:
        # Ordina per anno, mese e data creazione (dal più vecchio al più recente)
        records = await db["prima_nota_salari"].find(
            {"dipendente": dipendente}
        ).sort([("anno", 1), ("mese", 1), ("created_at", 1)]).to_list(5000)
        
        # Progressivo parte da 0
        progressivo = 0
        
        for record in records:
            # Skip record con vincolo
            if record.get("vincolo"):
                continue
            
            # Saldo della singola riga: Bonifico - Busta
            importo_busta = record.get("importo_busta", 0) or 0
            importo_bonifico = record.get("importo_bonifico", 0) or 0
            saldo = importo_bonifico - importo_busta
            
            anno_record = record.get("anno", 0)
            
            # Se l'anno è escluso, non aggiorna il progressivo
            if anno_record in anni_esclusi:
                prog_value = 0
            elif anno_inizio is None or anno_record >= anno_inizio:
                progressivo += saldo
                prog_value = round(progressivo, 2)
            else:
                prog_value = 0
            
            # Aggiorna il record con saldo e progressivo
            await db["prima_nota_salari"].update_one(
                {"_id": record["_id"]},
                {"$set": {
                    "saldo": round(saldo, 2),
                    "progressivo": prog_value
                }}
            )
    
    return True  # Conferma che il ricalcolo è avvenuto


@router.post("/ricalcola-progressivi")
async def ricalcola_progressivi(
    anno_inizio: Optional[int] = Query(None, description="Anno da cui iniziare il calcolo del progressivo (es. 2023)"),
    dipendente: Optional[str] = Query(None, description="Nome dipendente specifico (opzionale)"),
    force_reset: bool = Query(False, description="Forza il reset dei progressivi prima del ricalcolo"),
    anni_esclusi: Optional[str] = Query(None, description="Anni da escludere (separati da virgola, es. '2018,2019,2020')")
) -> Dict[str, Any]:
    """
    Ricalcola i progressivi per uno o tutti i dipendenti.
    Se anno_inizio è specificato, il progressivo parte da 0 a gennaio di quell'anno.
    Se dipendente è specificato, ricalcola solo per quel dipendente.
    Se force_reset è True, prima azzera tutti i progressivi e poi li ricalcola.
    Se anni_esclusi è specificato, quei record vengono ignorati nel calcolo.
    """
    db = Database.get_db()
    
    # Parse anni esclusi
    anni_esclusi_list = []
    if anni_esclusi:
        try:
            anni_esclusi_list = [int(a.strip()) for a in anni_esclusi.split(',') if a.strip()]
        except:
            pass
    
    # Se force_reset, prima azzera i progressivi
    if force_reset:
        reset_query = {}
        if dipendente:
            reset_query["dipendente"] = dipendente
        await db["prima_nota_salari"].update_many(
            reset_query,
            {"$set": {"progressivo": 0, "saldo": 0}}
        )
    
    # Ricalcola
    await ricalcola_progressivi_tutti(db, anno_inizio, dipendente, anni_esclusi_list)
    
    return {
        "message": f"Progressivi ricalcolati{f' dal {anno_inizio}' if anno_inizio else ''}{f' per {dipendente}' if dipendente else ''}{f' (esclusi: {anni_esclusi})' if anni_esclusi else ''}",
        "anno_inizio": anno_inizio,
        "dipendente": dipendente,
        "force_reset": force_reset,
        "anni_esclusi": anni_esclusi_list
    }


@router.post("/salari/aggiustamento")
async def aggiungi_aggiustamento(
    data: Dict[str, Any] = Body(...)
) -> Dict[str, Any]:
    """
    Aggiunge una riga di aggiustamento per allineare il saldo con il commercialista.
    L'importo positivo aumenta il saldo (va nel bonifico), negativo lo diminuisce (va nella busta).
    """
    db = Database.get_db()
    
    dipendente = data.get("dipendente", "").strip().upper()
    anno = int(data.get("anno", datetime.utcnow().year))
    mese = int(data.get("mese", datetime.utcnow().month))
    importo_busta = float(data.get("importo_busta", 0))
    importo_bonifico = float(data.get("importo_bonifico", 0))
    descrizione = data.get("descrizione", "Aggiustamento saldo")
    
    if not dipendente:
        raise HTTPException(status_code=400, detail="Dipendente obbligatorio")
    
    # Crea il record di aggiustamento
    new_record = {
        "id": str(uuid.uuid4()),
        "dipendente": dipendente,
        "anno": anno,
        "mese": mese,
        "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
        "importo_busta": round(importo_busta, 2),
        "importo_bonifico": round(importo_bonifico, 2),
        "saldo": round(importo_bonifico - importo_busta, 2),
        "progressivo": 0,  # Verrà ricalcolato
        "riconciliato": False,
        "tipo": "aggiustamento",
        "descrizione": descrizione,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    await db["prima_nota_salari"].insert_one(new_record)
    
    # Ricalcola i progressivi per questo dipendente
    await ricalcola_progressivi_tutti(db, None, dipendente)
    
    return {
        "success": True,
        "message": f"Aggiustamento inserito per {dipendente}",
        "record_id": new_record["id"]
    }


@router.get("/dipendenti-lista")
async def get_dipendenti_lista() -> List[str]:
    """Lista nomi dipendenti unici dalla prima nota salari."""
    db = Database.get_db()
    dipendenti = await db["prima_nota_salari"].distinct("dipendente")
    return sorted(dipendenti)


@router.delete("/salari/reset")
async def reset_prima_nota_salari(
    tipo: Optional[str] = Query(None, description="Tipo di record da eliminare: 'busta', 'bonifico', 'aggiustamento' o None per tutti")
) -> Dict[str, Any]:
    """
    Elimina i record della prima nota salari.
    - tipo=None: elimina TUTTI i record
    - tipo='busta': elimina solo i record delle paghe
    - tipo='bonifico': elimina solo i record dei bonifici
    - tipo='aggiustamento': elimina solo i record di aggiustamento
    """
    db = Database.get_db()
    
    query = {}
    if tipo:
        query["tipo"] = tipo
    
    result = await db["prima_nota_salari"].delete_many(query)
    
    tipo_desc = tipo if tipo else "tutti"
    return {
        "message": f"Eliminati {result.deleted_count} record ({tipo_desc})",
        "deleted_count": result.deleted_count,
        "tipo_eliminato": tipo_desc
    }


@router.post("/consolida-record")
async def consolida_record() -> Dict[str, Any]:
    """
    Consolida i record esistenti: unisce buste e bonifici dello stesso mese/dipendente in un'unica riga.
    Utile dopo import separati o per pulire dati duplicati.
    """
    db = Database.get_db()
    
    # Ottieni tutti i record
    all_records = await db["prima_nota_salari"].find({}, {"_id": 0}).to_list(10000)
    
    # Raggruppa per dipendente/anno/mese
    grouped = {}
    for r in all_records:
        key = (r.get("dipendente"), r.get("anno"), r.get("mese"))
        if key not in grouped:
            grouped[key] = {"busta": 0, "bonifico": 0, "records": []}
        grouped[key]["busta"] += r.get("importo_busta") or 0
        grouped[key]["bonifico"] += r.get("importo_bonifico") or 0
        grouped[key]["records"].append(r)
    
    # Conta quanti gruppi hanno più di un record
    duplicates = sum(1 for v in grouped.values() if len(v["records"]) > 1)
    
    if duplicates == 0:
        return {"message": "Nessun record da consolidare", "duplicates": 0}
    
    # Elimina tutti e ricrea consolidati
    await db["prima_nota_salari"].delete_many({})
    
    created = 0
    for (dipendente, anno, mese), data in grouped.items():
        if not dipendente or not anno or not mese:
            continue
        
        importo_busta = round(data["busta"], 2)
        importo_bonifico = round(data["bonifico"], 2)
        saldo = round(importo_bonifico - importo_busta, 2)
        
        new_record = {
            "id": str(uuid.uuid4()),
            "dipendente": dipendente,
            "anno": anno,
            "mese": mese,
            "mese_nome": MESI_NOMI[mese - 1] if 1 <= mese <= 12 else "",
            "importo_busta": importo_busta,
            "importo_bonifico": importo_bonifico,
            "saldo": saldo,
            "progressivo": 0,
            "riconciliato": False,
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }
        await db["prima_nota_salari"].insert_one(new_record)
        created += 1
    
    # Ricalcola progressivi
    await ricalcola_progressivi_tutti(db)
    
    return {
        "message": f"Consolidamento completato",
        "record_originali": len(all_records),
        "record_consolidati": created,
        "duplicati_risolti": duplicates
    }


@router.delete("/salari/{record_id}")
async def delete_salario(record_id: str) -> Dict[str, str]:
    """Elimina un singolo record."""
    db = Database.get_db()
    result = await db["prima_nota_salari"].delete_one({"id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    return {"message": "Record eliminato"}


@router.put("/salari/{record_id}")
async def update_salario(
    record_id: str,
    data: Dict[str, Any] = Body(...)
) -> Dict[str, str]:
    """Aggiorna un record della prima nota salari."""
    db = Database.get_db()
    
    # Calcola saldo: Bonifico - Busta
    importo_busta = float(data.get("importo_busta", 0))
    importo_bonifico = float(data.get("importo_bonifico", 0))
    saldo = importo_bonifico - importo_busta
    
    # Aggiorna il mese_nome se cambia il mese
    mese = int(data.get("mese", 1))
    mese_nome = MESI_NOMI[mese - 1] if 1 <= mese <= 12 else ""
    
    update_data = {
        "dipendente": data.get("dipendente", ""),
        "anno": int(data.get("anno", 2025)),
        "mese": mese,
        "mese_nome": mese_nome,
        "importo_busta": importo_busta,
        "importo_bonifico": importo_bonifico,
        "saldo": round(saldo, 2),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    # Aggiungi vincolo se presente
    if "vincolo" in data:
        update_data["vincolo"] = bool(data.get("vincolo"))
    
    result = await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Record aggiornato"}


@router.put("/salari/{record_id}/riconcilia")
async def riconcilia_salario(
    record_id: str,
    riconciliato: bool = Query(True)
) -> Dict[str, str]:
    """Marca un salario come riconciliato."""
    db = Database.get_db()
    result = await db["prima_nota_salari"].update_one(
        {"id": record_id},
        {"$set": {
            "riconciliato": riconciliato,
            "data_riconciliazione": datetime.utcnow().isoformat() if riconciliato else None,
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    return {"message": "Stato riconciliazione aggiornato"}


@router.get("/export-excel")
async def export_prima_nota_salari_excel(
    anno: Optional[int] = Query(None),
    mese: Optional[int] = Query(None)
):
    """Esporta prima nota salari in Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    db = Database.get_db()
    
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    
    salari = await db["prima_nota_salari"].find(
        query, {"_id": 0}
    ).sort([("anno", -1), ("mese", -1), ("dipendente", 1)]).to_list(5000)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prima Nota Salari"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    
    # Headers
    headers = ["Dipendente", "Anno", "Mese", "Importo Busta", "Importo Bonifico", "Saldo", "Progressivo", "Riconciliato"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for row_num, s in enumerate(salari, 2):
        ws.cell(row=row_num, column=1, value=s.get("dipendente", ""))
        ws.cell(row=row_num, column=2, value=s.get("anno", ""))
        ws.cell(row=row_num, column=3, value=s.get("mese_nome", ""))
        ws.cell(row=row_num, column=4, value=s.get("importo_busta", 0))
        ws.cell(row=row_num, column=5, value=s.get("importo_bonifico", 0))
        ws.cell(row=row_num, column=6, value=s.get("saldo", 0))
        ws.cell(row=row_num, column=7, value=s.get("progressivo", 0))
        ws.cell(row=row_num, column=8, value="Sì" if s.get("riconciliato") else "No")
    
    # Larghezze
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "prima_nota_salari"
    if anno:
        filename += f"_{anno}"
    if mese:
        filename += f"_{mese:02d}"
    filename += ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
