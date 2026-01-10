"""
Import Templates Router
Fornisce template Excel scaricabili per ogni tipo di importazione
"""
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()

def create_styled_workbook():
    """Crea un workbook con stili predefiniti."""
    wb = openpyxl.Workbook()
    return wb

def style_header_row(ws, headers, descriptions=None):
    """Applica stile alle intestazioni."""
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    desc_fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
    desc_font = Font(color="666666", italic=True, size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Intestazioni
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(15, len(header) + 5)
    
    # Descrizioni (riga 2)
    if descriptions:
        for col, desc in enumerate(descriptions, 1):
            cell = ws.cell(row=2, column=col, value=desc)
            cell.fill = desc_fill
            cell.font = desc_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 30

def add_example_row(ws, example_data, start_row=3):
    """Aggiunge una riga di esempio."""
    example_font = Font(color="999999", italic=True)
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=start_row, column=col, value=value)
        cell.font = example_font


@router.get("/versamenti")
async def template_versamenti():
    """Template per importazione versamenti in banca - formato banca."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Versamenti"
    
    headers = ["Ragione Sociale", "Data contabile", "Data valuta", "Banca", "Rapporto", "Importo", "Divisa", "Descrizione", "Causale", "Saldo", "Competenza"]
    descriptions = [
        "Nome azienda",
        "Data operazione",
        "Data valuta",
        "Nome banca",
        "N. rapporto",
        "Importo (€)",
        "EUR",
        "Descrizione movimento",
        "Codice causale",
        "Saldo progressivo",
        "Tipo competenza"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["AZIENDA SRL", "16/12/2025", "16/12/2025", "BANCA XYZ", "123456", "4500", "EUR", "VERS. CONTANTI", "48", "100000", "Ricavi - Deposito contanti"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_versamenti.xlsx"}
    )


@router.get("/pos")
async def template_pos():
    """Template per importazione incassi POS - formato banca."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Incassi POS"
    
    headers = ["DATA", "CONTO", "IMPORTO"]
    descriptions = [
        "Data operazione",
        "Tipo conto (pos)",
        "Importo giornaliero (€)"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["2025-01-15", "pos", 1500.00])
    add_example_row(ws, ["2025-01-16", "pos", 2000.00], start_row=4)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_pos.xlsx"}
    )


@router.get("/corrispettivi")
async def template_corrispettivi():
    """Template per importazione corrispettivi giornalieri - formato registratore cassa."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Corrispettivi"
    
    headers = [
        "Id invio", 
        "Matricola dispositivo", 
        "Data e ora rilevazione", 
        "Data e ora trasmissione", 
        "Ammontare delle vendite (totale in euro)", 
        "Imponibile vendite (totale in euro)", 
        "Imposta vendite (totale in euro)", 
        "Periodo di inattivita' da", 
        "Periodo di inattivita' a"
    ]
    descriptions = [
        "ID trasmissione",
        "Matricola RT",
        "Data chiusura",
        "Data invio",
        "Totale vendite (€)",
        "Imponibile (€)",
        "IVA (€)",
        "Inizio inattività",
        "Fine inattività"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["'2709633383'", "'99MEY026532'", "2025-01-02", "2025-01-02", 3264.55, 3264.55, 326.45, "", ""])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_corrispettivi.xlsx"}
    )


@router.get("/estratto-conto")
async def template_estratto_conto():
    """Template per importazione estratto conto bancario."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Estratto Conto"
    
    # Intestazioni corrette come da file banca
    headers = ["Ragione Sociale", "Data contabile", "Data valuta", "Banca", "Rapporto", "Importo", "Divisa", "Descrizione", "Categoria/sottocategoria", "Hashtag"]
    descriptions = [
        "Nome azienda",
        "Data operazione (DD/MM/YYYY)",
        "Data valuta (DD/MM/YYYY)",
        "Nome banca e codice",
        "Numero rapporto/conto",
        "Importo (+ entrata, - uscita)",
        "Valuta (EUR)",
        "Descrizione movimento",
        "Categoria contabile",
        "Tag opzionale"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["AZIENDA SRL", "08/01/2026", "08/01/2026", "05034 - BANCO BPM S.P.A.", "5462 - 03406 - 178800005462", "254,50", "EUR", "INCAS. TRAMITE P.O.S - NUMIA-PGBNT DEL 07/01/26", "Ricavi - Incasso tramite POS", ""])
    add_example_row(ws, ["AZIENDA SRL", "08/01/2026", "08/01/2026", "05034 - BANCO BPM S.P.A.", "5462 - 03406 - 178800005462", "-1500,00", "EUR", "BONIFICO A FAVORE FORNITORE XYZ", "Costi - Pagamento fornitore", ""], start_row=4)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_estratto_conto.xlsx"}
    )


@router.get("/fornitori")
async def template_fornitori():
    """Template per importazione anagrafica fornitori."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Fornitori"
    
    headers = ["ragione_sociale", "partita_iva", "codice_fiscale", "indirizzo", "citta", "cap", "email", "telefono", "metodo_pagamento"]
    descriptions = [
        "Nome azienda *",
        "P.IVA (11 cifre) *",
        "Codice Fiscale",
        "Indirizzo",
        "Città",
        "CAP",
        "Email",
        "Telefono",
        "bonifico/cassa/carta"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["ACME SRL", "12345678901", "12345678901", "Via Roma 1", "Milano", "20100", "info@acme.it", "02123456", "bonifico"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_fornitori.xlsx"}
    )


@router.get("/prodotti")
async def template_prodotti():
    """Template per importazione prodotti magazzino."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Prodotti"
    
    headers = ["nome", "codice", "categoria", "unita_misura", "prezzo_acquisto", "prezzo_vendita", "giacenza", "scorta_minima"]
    descriptions = [
        "Nome prodotto *",
        "Codice/SKU",
        "Categoria",
        "pz/kg/lt/cf",
        "Prezzo acquisto (€)",
        "Prezzo vendita (€)",
        "Quantità attuale",
        "Quantità minima alert"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["Farina 00 kg 25", "FAR001", "Materie Prime", "kg", 18.50, 0, 100, 20])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_prodotti.xlsx"}
    )


@router.get("/dipendenti")
async def template_dipendenti():
    """Template per importazione anagrafica dipendenti."""
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Dipendenti"
    
    headers = ["nome", "cognome", "codice_fiscale", "data_nascita", "data_assunzione", "qualifica", "retribuzione_lorda", "email", "telefono"]
    descriptions = [
        "Nome *",
        "Cognome *",
        "Codice Fiscale *",
        "Data nascita",
        "Data assunzione",
        "Qualifica/Mansione",
        "Retribuzione lorda (€)",
        "Email",
        "Telefono"
    ]
    
    style_header_row(ws, headers, descriptions)
    add_example_row(ws, ["Mario", "Rossi", "RSSMRA80A01H501Z", "1980-01-01", "2020-03-15", "Operaio", 1800.00, "mario.rossi@email.it", "3331234567"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_dipendenti.xlsx"}
    )
