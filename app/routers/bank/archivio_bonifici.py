"""
Archivio Bonifici - Import PDF bonifici bancari con parsing automatico.
Estrae dati da PDF bancari e li salva nel database.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from pathlib import Path
import os
import io
import uuid
import re
import hashlib
import zipfile
import logging
import asyncio

# PDF parsing
from pdfminer.high_level import extract_text
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

from app.database import Database

router = APIRouter(prefix="/archivio-bonifici", tags=["Archivio Bonifici"])
logger = logging.getLogger(__name__)

# Directory per upload temporanei (fuori dalla directory monitorata)
UPLOAD_DIR = Path("/tmp/bonifici_uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ---- UTILS ----
IBAN_RE = re.compile(r"\b[A-Z]{2}[0-9]{2}[A-Z0-9]{1,30}\b")
DATE_RE = [
    re.compile(r"(\d{2})[\/-](\d{2})[\/-](\d{4})"),
    re.compile(r"(\d{4})[\/-](\d{2})[\/-](\d{2})"),
]
AMOUNT_RE = re.compile(r"([+-]?)\s?(\d{1,3}(?:[\.,]\d{3})*|\d+)([\.,]\d{2})")


def parse_date(text: str) -> Optional[datetime]:
    """Estrae data da testo."""
    for rx in DATE_RE:
        m = rx.search(text)
        if m:
            try:
                if rx is DATE_RE[0]:
                    d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                else:
                    y, mth, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                return datetime(y, mth, d, tzinfo=timezone.utc)
            except Exception:
                continue
    return None


def parse_amount(text: str) -> Optional[float]:
    """Estrae importo da testo."""
    t = text.replace("€", " ").replace("EUR", " ").replace("EURO", " ")
    m = AMOUNT_RE.search(t.replace(" ", ""))
    if not m:
        return None
    sign = -1.0 if m.group(1) == '-' else 1.0
    integer = m.group(2).replace('.', '').replace(',', '')
    cents = m.group(3).replace(',', '.').replace(' ', '')
    try:
        base = float(integer)
        cent_val = float(cents)
        return sign * (base + cent_val)
    except Exception:
        return None


def normalize_str(s: Optional[str]) -> Optional[str]:
    """Normalizza stringa rimuovendo spazi multipli."""
    if not s:
        return None
    return re.sub(r"\s+", " ", s).strip()


def safe_filename(name: str) -> str:
    """Rende sicuro un nome file."""
    base = re.sub(r"[^A-Za-z0-9._-]", "_", name)
    if len(base) > 128:
        root, ext = os.path.splitext(base)
        base = root[:110] + '_' + hashlib.sha1(base.encode()).hexdigest()[:8] + ext
    return base


def build_dedup_key(t: Dict[str, Any]) -> str:
    """Genera chiave univoca per deduplicazione."""
    if t.get('cro_trn'):
        return f"CRO:{re.sub(r'[^A-Z0-9]', '', str(t['cro_trn']).upper())}"
    d = None
    if isinstance(t.get('data'), datetime):
        d = t['data'].strftime('%Y-%m-%d')
    elif isinstance(t.get('data'), str):
        d = t['data'][:10]
    amt = t.get('importo')
    b_iban = (t.get('beneficiario') or {}).get('iban')
    b_name = (t.get('beneficiario') or {}).get('nome')
    base = f"{d}|{amt}|{(b_iban or '')[-12:]}|{b_name or ''}"
    return 'CMP:' + hashlib.sha1(base.encode('utf-8')).hexdigest()


def parse_filename_data(filename: str) -> Optional[Dict[str, Any]]:
    """
    Estrae dati dal nome del file PDF.
    Pattern supportati:
    - Bonifico - ricevuta per ordinante_DD-MM-YYYY_IMPORTO.pdf
    - Contabile di filiale_DD-MM-YYYY_IMPORTO.pdf
    - NomeFile_DD-MM-YYYY_IMPORTO_N.pdf
    """
    # Pattern: tipo_data_importo[_n].pdf
    pattern = r'^(.+?)_(\d{2}-\d{2}-\d{4})(?:_(\d+[,.]?\d*))?(?:_\d+)?\.pdf$'
    match = re.match(pattern, filename, re.IGNORECASE)
    
    if not match:
        return None
    
    tipo, data_str, importo_str = match.groups()
    
    # Parsing data (DD-MM-YYYY -> YYYY-MM-DD)
    try:
        parts = data_str.split('-')
        data = f"{parts[2]}-{parts[1]}-{parts[0]}"
    except:
        data = None
    
    # Parsing importo
    importo = None
    if importo_str:
        try:
            importo = float(importo_str.replace('.', '').replace(',', '.'))
        except:
            pass
    
    # Determina tipo documento
    tipo_lower = tipo.lower()
    causale = None
    if 'bonifico' in tipo_lower:
        causale = "Bonifico"
    elif 'contabile' in tipo_lower:
        causale = "Contabile di filiale"
    else:
        causale = tipo[:50]
    
    return {
        'data': data,
        'importo': importo,
        'valuta': 'EUR',
        'ordinante': {'nome': None, 'iban': None},
        'beneficiario': {'nome': None, 'iban': None},
        'causale': causale,
        'cro_trn': None,
        'banca': None,
        'note': f"Importato da: {filename}",
        'tipo_documento': tipo[:100]
    }


def read_pdf_text(pdf_path: Path) -> str:
    """Estrae testo da PDF."""
    try:
        text = extract_text(str(pdf_path)) or ""
        if text.strip():
            return text
    except Exception as e:
        logger.warning(f"pdfminer failed for {pdf_path}: {e}")
    try:
        if fitz:
            doc = fitz.open(str(pdf_path))
            parts = []
            for page in doc:
                parts.append(page.get_text("text"))
            doc.close()
            return "\n".join(parts)
    except Exception as e:
        logger.exception(f"PyMuPDF parse failed for {pdf_path}: {e}")
    return ""


def extract_transfers_from_text(text: str) -> List[Dict[str, Any]]:
    """Estrae bonifici dal testo PDF."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    
    # Cerca dati strutturati
    results: List[Dict[str, Any]] = []
    
    # Parsing base
    dt = parse_date(text)
    amt = None
    
    # Cerca importo
    m_amt = re.search(r"\b(EUR|EURO)?\s*([+-]?\d{1,3}(?:[\.,]\d{3})*(?:[\.,]\d{2}))\b", text, re.IGNORECASE)
    if m_amt:
        try:
            amt = float(m_amt.group(2).replace('.', '').replace(',', '.'))
        except Exception:
            pass
    
    # Cerca CRO/TRN
    mcro = re.search(r"\b(?:CRO|TRN|NS\s*RIF\.?|RIF\.?\s*(?:OPERAZIONE)?)[:\s]*([A-Z0-9]*[0-9][A-Z0-9]{3,39})\b", text, re.IGNORECASE)
    cro = mcro.group(1).strip() if mcro else None
    
    # Cerca causale
    caus = None
    mca = re.search(r"causale[:\s]*([^\n]+)", text, re.IGNORECASE)
    if mca:
        caus = normalize_str(mca.group(1))
    
    # Cerca IBAN
    ibans = IBAN_RE.findall(text.replace(' ', ''))
    ben_iban = ibans[0] if ibans else None
    ord_iban = ibans[1] if len(ibans) > 1 else None
    
    # Cerca nomi
    ord_nome = None
    ben_nome = None
    for idx, line in enumerate(lines):
        if re.search(r"beneficiario", line, re.IGNORECASE):
            after = re.sub(r"(?i).*beneficiario[:\s]*", "", line).strip()
            if after and len(after) > 2:
                ben_nome = normalize_str(after)
        if re.search(r"ordinante", line, re.IGNORECASE):
            after = re.sub(r"(?i).*ordinante[:\s]*", "", line).strip()
            if after and len(after) > 2:
                ord_nome = normalize_str(after)
    
    results.append({
        'data': dt,
        'importo': amt,
        'valuta': 'EUR',
        'ordinante': {'nome': ord_nome, 'iban': ord_iban},
        'beneficiario': {'nome': ben_nome, 'iban': ben_iban},
        'causale': caus,
        'cro_trn': cro,
        'banca': None,
        'note': None,
    })
    
    return results


# ---- ENDPOINTS ----

@router.post("/jobs")
async def create_job() -> Dict[str, Any]:
    """Crea un nuovo job di import."""
    db = Database.get_db()
    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    job_data = {
        'id': job_id,
        'status': 'created',
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'total_files': 0,
        'processed_files': 0,
        'errors': 0,
        'imported_files': 0,
    }
    await db.bonifici_jobs.insert_one({**job_data}.copy())
    return job_data


@router.get("/jobs")
async def list_jobs():
    """Lista tutti i job."""
    db = Database.get_db()
    jobs = await db.bonifici_jobs.find({}, {'_id': 0}).sort('created_at', -1).to_list(100)
    return jobs


@router.get("/jobs/{job_id}")
async def get_job(job_id: str):
    """Ottiene stato di un job."""
    db = Database.get_db()
    job = await db.bonifici_jobs.find_one({'id': job_id}, {'_id': 0})
    if not job:
        raise HTTPException(status_code=404, detail='Job not found')
    return job


@router.post("/jobs/{job_id}/upload")
async def upload_files(job_id: str, background: BackgroundTasks, files: List[UploadFile] = File(...)):
    """
    Carica file PDF o ZIP per elaborazione.
    Supporta ZIP con fino a 1500 PDF.
    """
    db = Database.get_db()
    job = await db.bonifici_jobs.find_one({'id': job_id})
    if not job:
        raise HTTPException(status_code=404, detail='Job not found')
    
    job_dir = UPLOAD_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    
    pdf_paths: List[Path] = []
    zip_errors: List[str] = []
    
    for f in files:
        name = safe_filename(Path(f.filename).name)
        
        if name.lower().endswith('.zip'):
            # Salva ZIP temporaneamente su disco per gestire file grandi
            zip_path = job_dir / f"temp_{name}"
            try:
                # Leggi in chunk per file grandi
                with open(zip_path, 'wb') as fd:
                    while chunk := await f.read(1024 * 1024):  # 1MB chunks
                        fd.write(chunk)
                
                # Estrai PDF da ZIP
                with zipfile.ZipFile(zip_path, 'r') as z:
                    pdf_count = 0
                    for info in z.infolist():
                        if info.is_dir():
                            continue
                        if not info.filename.lower().endswith('.pdf'):
                            continue
                        
                        pdf_name = safe_filename(Path(info.filename).name)
                        # Evita collisioni di nomi
                        out_path = job_dir / f"{pdf_count:04d}_{pdf_name}"
                        
                        try:
                            with z.open(info) as fsrc, open(out_path, 'wb') as fdst:
                                fdst.write(fsrc.read())
                            pdf_paths.append(out_path)
                            pdf_count += 1
                        except Exception as e:
                            zip_errors.append(f"{info.filename}: {str(e)}")
                
                # Rimuovi ZIP temporaneo
                zip_path.unlink(missing_ok=True)
                
            except zipfile.BadZipFile:
                zip_errors.append(f"{name}: File ZIP corrotto")
            except Exception as e:
                zip_errors.append(f"{name}: {str(e)}")
                
        elif name.lower().endswith('.pdf'):
            out = job_dir / name
            with open(out, 'wb') as fd:
                while chunk := await f.read(1024 * 1024):
                    fd.write(chunk)
            pdf_paths.append(out)
    
    # Aggiorna job
    await db.bonifici_jobs.update_one(
        {'id': job_id},
        {'$set': {
            'status': 'queued',
            'total_files': len(pdf_paths),
            'zip_errors': zip_errors[:50],  # Limita errori salvati
            'updated_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Avvia elaborazione in background
    logger.info(f"Starting background processing for job {job_id} with {len(pdf_paths)} files")
    background.add_task(process_files_background, job_id, pdf_paths)
    
    return {
        'job_id': job_id, 
        'accepted_files': len(pdf_paths),
        'extraction_errors': len(zip_errors),
        'errors_sample': zip_errors[:5] if zip_errors else []
    }


async def process_files_background(job_id: str, file_paths: List[Path]):
    """
    Elabora i file PDF in background con deduplicazione avanzata.
    Supporta fino a 1500 file con gestione memoria ottimizzata.
    """
    logger.info(f"process_files_background started for job {job_id}")
    db = Database.get_db()
    
    await db.bonifici_jobs.update_one({'id': job_id}, {'$set': {'status': 'processing'}})
    
    processed = 0
    errors = 0
    imported = 0
    duplicates = 0
    
    # Carica chiavi esistenti per deduplicazione veloce
    existing_keys = set()
    existing_docs = await db.bonifici_transfers.find({}, {'_id': 0, 'dedup_key': 1}).to_list(50000)
    for doc in existing_docs:
        if doc.get('dedup_key'):
            existing_keys.add(doc['dedup_key'])
    
    for p in file_paths:
        try:
            text = read_pdf_text(p)
            
            transfers = extract_transfers_from_text(text) if text.strip() else []
            
            # Fallback: estrai dati dal nome file se il parsing PDF fallisce
            if not transfers or (transfers and not transfers[0].get('importo')):
                filename_data = parse_filename_data(p.name)
                if filename_data:
                    if transfers:
                        # Arricchisci con dati dal filename
                        transfers[0].update({k: v for k, v in filename_data.items() if v and not transfers[0].get(k)})
                    else:
                        transfers = [filename_data]
            
            if not transfers:
                errors += 1
                continue
            
            for t in transfers:
                t['source_file'] = p.name
                t['job_id'] = job_id
                t['id'] = str(uuid.uuid4())
                t['dedup_key'] = build_dedup_key(t)
                t['created_at'] = datetime.now(timezone.utc).isoformat()
                
                # Converti data in stringa
                if isinstance(t.get('data'), datetime):
                    t['data'] = t['data'].isoformat()
                
                # Deduplicazione con cache in memoria
                if t['dedup_key'] in existing_keys:
                    duplicates += 1
                    continue
                
                # Inserisci nuovo bonifico
                await db.bonifici_transfers.insert_one(t.copy())
                existing_keys.add(t['dedup_key'])
                imported += 1
                
        except Exception as e:
            errors += 1
            logger.exception(f"Processing failed for {p}: {e}")
        finally:
            processed += 1
            
            # Aggiorna stato ogni 10 file o all'ultimo
            if processed % 10 == 0 or processed == len(file_paths):
                await db.bonifici_jobs.update_one(
                    {'id': job_id},
                    {'$set': {
                        'processed_files': processed,
                        'errors': errors,
                        'imported_files': imported,
                        'duplicates_skipped': duplicates,
                        'updated_at': datetime.now(timezone.utc).isoformat()
                    }}
                )
            
            # Elimina file processato per liberare spazio
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass
    
    # Pulisci directory job
    job_dir = UPLOAD_DIR / job_id
    try:
        import shutil
        shutil.rmtree(job_dir, ignore_errors=True)
    except Exception:
        pass
    
    # AUTO-ASSOCIAZIONE: Cerca match automatici per i nuovi bonifici
    auto_associated_salari = 0
    auto_associated_fatture = 0
    
    if imported > 0:
        try:
            # Prendi i bonifici appena importati
            new_bonifici = await db.bonifici_transfers.find(
                {"job_id": job_id, "salario_associato": {"$ne": True}, "fattura_associata": {"$ne": True}},
                {"_id": 0}
            ).to_list(500)
            
            for bonifico in new_bonifici:
                importo = abs(bonifico.get("importo", 0))
                beneficiario_nome = ((bonifico.get("beneficiario") or {}).get("nome") or "").lower()
                causale = (bonifico.get("causale") or "").lower()
                
                if importo <= 0:
                    continue
                
                # 1. Cerca match SALARI (score >= 80)
                salari_match = await db.prima_nota_salari.find_one({
                    "salario_associato": {"$ne": True},
                    "$or": [
                        {"importo_busta": {"$gte": importo * 0.98, "$lte": importo * 1.02}},
                        {"importo_bonifico": {"$gte": importo * 0.98, "$lte": importo * 1.02}}
                    ]
                }, {"_id": 0})
                
                if salari_match:
                    dipendente = (salari_match.get("dipendente") or "").lower()
                    # Verifica match nome dipendente
                    if dipendente and (dipendente in beneficiario_nome or dipendente in causale or 
                                       beneficiario_nome in dipendente):
                        # Associa automaticamente
                        await db.bonifici_transfers.update_one(
                            {"id": bonifico["id"]},
                            {"$set": {
                                "salario_associato": True,
                                "operazione_salario_id": salari_match.get("id"),
                                "operazione_salario_desc": f"{salari_match.get('dipendente', '')} - {salari_match.get('anno', '')}/{salari_match.get('mese', '')}",
                                "auto_associated": True,
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        await db.prima_nota_salari.update_one(
                            {"id": salari_match.get("id")},
                            {"$set": {"salario_associato": True, "bonifico_id": bonifico["id"]}}
                        )
                        auto_associated_salari += 1
                        continue
                
                # 2. Cerca match FATTURE (score >= 80)
                fattura_match = await db.invoices.find_one({
                    "fattura_associata": {"$ne": True},
                    "$or": [
                        {"total_amount": {"$gte": importo * 0.98, "$lte": importo * 1.02}},
                        {"importo_totale": {"$gte": importo * 0.98, "$lte": importo * 1.02}},
                        {"totale_documento": {"$gte": importo * 0.98, "$lte": importo * 1.02}}
                    ]
                }, {"_id": 0})
                
                if fattura_match:
                    fornitore = (fattura_match.get("supplier_name") or 
                                (fattura_match.get("fornitore", {}) or {}).get("denominazione") or "").lower()
                    # Verifica match nome fornitore
                    if fornitore and (fornitore in beneficiario_nome or fornitore in causale or
                                     beneficiario_nome in fornitore):
                        # Associa automaticamente
                        await db.bonifici_transfers.update_one(
                            {"id": bonifico["id"]},
                            {"$set": {
                                "fattura_associata": True,
                                "fattura_id": fattura_match.get("id"),
                                "fattura_numero": fattura_match.get("invoice_number") or fattura_match.get("numero_documento"),
                                "fattura_fornitore": fattura_match.get("supplier_name") or (fattura_match.get("fornitore", {}) or {}).get("denominazione"),
                                "auto_associated": True,
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        await db.invoices.update_one(
                            {"id": fattura_match.get("id")},
                            {"$set": {"bonifico_associato": True, "bonifico_id": bonifico["id"]}}
                        )
                        auto_associated_fatture += 1
        except Exception as e:
            logger.error(f"Auto-association error: {e}")
    
    await db.bonifici_jobs.update_one(
        {'id': job_id}, 
        {'$set': {
            'status': 'completed',
            'completed_at': datetime.now(timezone.utc).isoformat(),
            'auto_associated_salari': auto_associated_salari,
            'auto_associated_fatture': auto_associated_fatture
        }}
    )


@router.get("/transfers")
async def list_transfers(
    job_id: Optional[str] = None,
    search: Optional[str] = None,
    ordinante: Optional[str] = None,
    beneficiario: Optional[str] = None,
    year: Optional[str] = None,
    limit: int = Query(1000, le=10000)
):
    """Lista bonifici con filtri."""
    db = Database.get_db()
    
    query: Dict[str, Any] = {}
    if job_id:
        query['job_id'] = job_id
    
    ands = []
    if search:
        ands.append({'$or': [
            {'ordinante.nome': {'$regex': search, '$options': 'i'}},
            {'beneficiario.nome': {'$regex': search, '$options': 'i'}},
            {'causale': {'$regex': search, '$options': 'i'}},
            {'cro_trn': {'$regex': search, '$options': 'i'}},
        ]})
    if ordinante:
        ands.append({'ordinante.nome': {'$regex': ordinante, '$options': 'i'}})
    if beneficiario:
        ands.append({'beneficiario.nome': {'$regex': beneficiario, '$options': 'i'}})
    if year:
        ands.append({'data': {'$regex': f'^{year}-'}})
    
    if ands:
        query['$and'] = ands
    
    transfers = await db.bonifici_transfers.find(query, {'_id': 0}).sort('data', -1).to_list(limit)
    return transfers


@router.get("/transfers/count")
async def count_transfers(job_id: Optional[str] = None):
    """Conta bonifici totali."""
    db = Database.get_db()
    query = {'job_id': job_id} if job_id else {}
    count = await db.bonifici_transfers.count_documents(query)
    return {'count': count}


@router.get("/transfers/summary")
async def transfers_summary():
    """Riepilogo per anno."""
    db = Database.get_db()
    
    pipeline = [
        {'$addFields': {
            'year': {'$substr': ['$data', 0, 4]}
        }},
        {'$group': {
            '_id': '$year',
            'count': {'$sum': 1},
            'total': {'$sum': '$importo'}
        }},
        {'$sort': {'_id': -1}}
    ]
    
    results = await db.bonifici_transfers.aggregate(pipeline).to_list(100)
    return {r['_id']: {'count': r['count'], 'total': round(r['total'] or 0, 2)} for r in results if r['_id']}


@router.delete("/transfers/{transfer_id}")
async def delete_transfer(transfer_id: str):
    """Elimina un bonifico."""
    db = Database.get_db()
    result = await db.bonifici_transfers.delete_one({'id': transfer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Transfer not found')
    return {'deleted': True}


@router.get("/transfers/{transfer_id}/pdf")
async def get_bonifico_pdf(transfer_id: str):
    """
    Restituisce il PDF originale del bonifico se disponibile.
    Cerca il file nella directory di upload usando il source_file.
    """
    db = Database.get_db()
    
    # Recupera bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": transfer_id}, {"_id": 0})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    source_file = bonifico.get("source_file", "")
    job_id = bonifico.get("job_id", "")
    
    if not source_file:
        raise HTTPException(status_code=404, detail="File PDF non disponibile per questo bonifico")
    
    # Cerca il file in varie posizioni
    possible_paths = [
        UPLOAD_DIR / job_id / source_file,
        UPLOAD_DIR / source_file,
        Path(f"/tmp/bonifici_uploads/{job_id}/{source_file}"),
        Path(f"/tmp/bonifici_uploads/{source_file}"),
    ]
    
    pdf_path = None
    for p in possible_paths:
        if p.exists():
            pdf_path = p
            break
    
    if not pdf_path:
        # Il file è stato cancellato dopo l'elaborazione (normale comportamento)
        raise HTTPException(
            status_code=404, 
            detail="Il file PDF originale non è più disponibile. I file vengono rimossi dopo l'elaborazione per risparmiare spazio."
        )
    
    return StreamingResponse(
        open(pdf_path, "rb"),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{source_file}"'}
    )


@router.delete("/transfers/bulk")
async def bulk_delete(job_id: Optional[str] = None):
    """Elimina tutti i bonifici di un job."""
    db = Database.get_db()
    query = {'job_id': job_id} if job_id else {}
    result = await db.bonifici_transfers.delete_many(query)
    return {'deleted': result.deleted_count}


@router.get("/export")
async def export_transfers(
    format: str = Query('xlsx', pattern='^(csv|xlsx)$'),
    job_id: Optional[str] = None
):
    """Esporta bonifici in CSV o XLSX."""
    db = Database.get_db()
    query = {'job_id': job_id} if job_id else {}
    transfers = await db.bonifici_transfers.find(query, {'_id': 0}).to_list(10000)
    
    if format == 'csv':
        import csv
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=';')
        headers = ['data', 'importo', 'valuta', 'ordinante', 'ordinante_iban', 'beneficiario', 'beneficiario_iban', 'causale', 'cro_trn']
        w.writerow(headers)
        for t in transfers:
            w.writerow([
                t.get('data', ''),
                t.get('importo', ''),
                t.get('valuta', 'EUR'),
                (t.get('ordinante') or {}).get('nome', ''),
                (t.get('ordinante') or {}).get('iban', ''),
                (t.get('beneficiario') or {}).get('nome', ''),
                (t.get('beneficiario') or {}).get('iban', ''),
                t.get('causale', ''),
                t.get('cro_trn', '')
            ])
        data = buf.getvalue().encode('utf-8')
        headers = {'Content-Disposition': 'attachment; filename="bonifici_export.csv"'}
        return StreamingResponse(io.BytesIO(data), media_type='text/csv', headers=headers)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bonifici'
        ws.append(['Data', 'Importo', 'Valuta', 'Ordinante', 'IBAN Ord.', 'Beneficiario', 'IBAN Ben.', 'Causale', 'CRO/TRN'])
        for t in transfers:
            ws.append([
                t.get('data', ''),
                t.get('importo', ''),
                t.get('valuta', 'EUR'),
                (t.get('ordinante') or {}).get('nome', ''),
                (t.get('ordinante') or {}).get('iban', ''),
                (t.get('beneficiario') or {}).get('nome', ''),
                (t.get('beneficiario') or {}).get('iban', ''),
                t.get('causale', ''),
                t.get('cro_trn', '')
            ])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        headers = {'Content-Disposition': 'attachment; filename="bonifici_export.xlsx"'}
        return StreamingResponse(out, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


# =============================================================================
# RICONCILIAZIONE BONIFICI CON ESTRATTO CONTO
# =============================================================================

# Store per task riconciliazione in background
_riconciliazione_task: Dict[str, Any] = {}


async def _execute_riconciliazione_batch(task_id: str):
    """Esegue la riconciliazione in background con chunking."""
    db = Database.get_db()
    
    try:
        _riconciliazione_task[task_id]["status"] = "in_progress"
        _riconciliazione_task[task_id]["message"] = "Caricamento dati..."
        
        # Recupera tutti i bonifici non ancora riconciliati
        bonifici = await db.bonifici_transfers.find(
            {"riconciliato": {"$ne": True}},
            {"_id": 0}
        ).to_list(10000)
        
        # Recupera movimenti estratto conto
        movimenti = await db.estratto_conto_movimenti.find(
            {},
            {"_id": 0}
        ).to_list(50000)
        
        if not bonifici:
            _riconciliazione_task[task_id]["status"] = "completed"
            _riconciliazione_task[task_id]["result"] = {"riconciliati": 0, "message": "Nessun bonifico da riconciliare"}
            return
        
        if not movimenti:
            _riconciliazione_task[task_id]["status"] = "completed"
            _riconciliazione_task[task_id]["result"] = {"riconciliati": 0, "message": "Nessun movimento estratto conto"}
            return
        
        _riconciliazione_task[task_id]["total"] = len(bonifici)
        _riconciliazione_task[task_id]["message"] = f"Riconciliazione {len(bonifici)} bonifici..."
        
        riconciliati = 0
        movimenti_usati = set()
        processed = 0
        
        # Processa in chunk di 50 per evitare timeout
        chunk_size = 50
        
        for bonifico in bonifici:
            processed += 1
            
            # Fix: gestisce None esplicito per importo
            raw_importo = bonifico.get("importo")
            if raw_importo is None:
                continue
            bonifico_importo = abs(float(raw_importo))
            bonifico_data_str = bonifico.get("data", "")
            
            try:
                if "T" in bonifico_data_str:
                    bonifico_data = datetime.fromisoformat(bonifico_data_str.replace("+00:00", "").replace("Z", ""))
                else:
                    bonifico_data = datetime.strptime(bonifico_data_str[:10], "%Y-%m-%d")
            except:
                continue
            
            match_found = None
            
            for idx, mov in enumerate(movimenti):
                if idx in movimenti_usati:
                    continue
                
                # Fix: gestisce None esplicito per importo movimento
                raw_mov_importo = mov.get("importo")
                if raw_mov_importo is None:
                    continue
                mov_importo = abs(float(raw_mov_importo))
                mov_data_str = mov.get("data", "")
                
                try:
                    mov_data = datetime.strptime(mov_data_str[:10], "%Y-%m-%d")
                except:
                    continue
                
                if abs(bonifico_importo - mov_importo) > 0.01:
                    continue
                
                diff_giorni = abs((bonifico_data - mov_data).days)
                if diff_giorni > 1:
                    continue
                
                match_found = mov
                movimenti_usati.add(idx)
                break
            
            if match_found:
                await db.bonifici_transfers.update_one(
                    {"id": bonifico.get("id")},
                    {"$set": {
                        "riconciliato": True,
                        "data_riconciliazione": datetime.now(timezone.utc),
                        "movimento_estratto_conto_id": match_found.get("id"),
                        "movimento_data": match_found.get("data"),
                        "movimento_descrizione": match_found.get("descrizione_originale", "")[:100]
                    }}
                )
                riconciliati += 1
            
            # Aggiorna progress ogni chunk
            if processed % chunk_size == 0:
                _riconciliazione_task[task_id]["processed"] = processed
                _riconciliazione_task[task_id]["riconciliati"] = riconciliati
                # Yield per non bloccare l'event loop
                await asyncio.sleep(0)
        
        _riconciliazione_task[task_id]["status"] = "completed"
        _riconciliazione_task[task_id]["result"] = {
            "riconciliati": riconciliati,
            "totale_bonifici": len(bonifici),
            "non_riconciliati": len(bonifici) - riconciliati
        }
        _riconciliazione_task[task_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
        
    except Exception as e:
        logger.error(f"Errore riconciliazione task {task_id}: {e}")
        _riconciliazione_task[task_id]["status"] = "error"
        _riconciliazione_task[task_id]["error"] = str(e)


@router.post("/riconcilia")
async def riconcilia_bonifici_con_estratto(
    background: bool = Query(False, description="Esegui in background (consigliato)")
):
    """
    Riconcilia i bonifici con i movimenti dell'estratto conto.
    Match basato su: importo esatto e data (±1 giorno).
    
    Se background=true, avvia la riconciliazione in background e restituisce task_id.
    """
    import asyncio
    
    if background:
        # Modalità background
        task_id = str(uuid.uuid4())
        _riconciliazione_task[task_id] = {
            "task_id": task_id,
            "status": "pending",
            "message": "Avvio riconciliazione...",
            "started_at": datetime.now(timezone.utc).isoformat(),
            "processed": 0,
            "total": 0,
            "riconciliati": 0
        }
        
        # Avvia in background
        asyncio.create_task(_execute_riconciliazione_batch(task_id))
        
        return {
            "background": True,
            "task_id": task_id,
            "message": "Riconciliazione avviata in background. Usa /riconcilia/task/{task_id} per lo stato."
        }
    
    # Modalità sincrona (legacy)
    db = Database.get_db()
    
    bonifici = await db.bonifici_transfers.find(
        {"riconciliato": {"$ne": True}},
        {"_id": 0}
    ).to_list(10000)
    
    movimenti = await db.estratto_conto_movimenti.find(
        {},
        {"_id": 0}
    ).to_list(50000)
    
    if not bonifici:
        return {"success": True, "message": "Nessun bonifico da riconciliare", "riconciliati": 0}
    
    if not movimenti:
        return {"success": False, "message": "Nessun movimento estratto conto caricato", "riconciliati": 0}
    
    riconciliati = 0
    movimenti_usati = set()
    
    for bonifico in bonifici:
        # Fix: gestisce None esplicito per importo
        raw_importo = bonifico.get("importo")
        if raw_importo is None:
            continue
        bonifico_importo = abs(float(raw_importo))
        bonifico_data_str = bonifico.get("data", "")
        
        try:
            if "T" in bonifico_data_str:
                bonifico_data = datetime.fromisoformat(bonifico_data_str.replace("+00:00", "").replace("Z", ""))
            else:
                bonifico_data = datetime.strptime(bonifico_data_str[:10], "%Y-%m-%d")
        except:
            continue
        
        match_found = None
        
        for idx, mov in enumerate(movimenti):
            if idx in movimenti_usati:
                continue
            
            # Fix: gestisce None esplicito per importo movimento
            raw_mov_importo = mov.get("importo")
            if raw_mov_importo is None:
                continue
            mov_importo = abs(float(raw_mov_importo))
            mov_data_str = mov.get("data", "")
            
            try:
                mov_data = datetime.strptime(mov_data_str[:10], "%Y-%m-%d")
            except:
                continue
            
            if abs(bonifico_importo - mov_importo) > 0.01:
                continue
            
            diff_giorni = abs((bonifico_data - mov_data).days)
            if diff_giorni > 1:
                continue
            
            match_found = mov
            movimenti_usati.add(idx)
            break
        
        if match_found:
            await db.bonifici_transfers.update_one(
                {"id": bonifico.get("id")},
                {"$set": {
                    "riconciliato": True,
                    "data_riconciliazione": datetime.now(timezone.utc),
                    "movimento_estratto_conto_id": match_found.get("id"),
                    "movimento_data": match_found.get("data"),
                    "movimento_descrizione": match_found.get("descrizione_originale", "")[:100]
                }}
            )
            riconciliati += 1
    
    return {
        "success": True,
        "message": f"Riconciliazione completata: {riconciliati} bonifici riconciliati",
        "riconciliati": riconciliati,
        "totale_bonifici": len(bonifici),
        "non_riconciliati": len(bonifici) - riconciliati
    }


@router.get("/riconcilia/task/{task_id}")
async def get_riconciliazione_task(task_id: str):
    """Stato di un task di riconciliazione in background."""
    if task_id not in _riconciliazione_task:
        raise HTTPException(status_code=404, detail="Task non trovato")
    return _riconciliazione_task[task_id]


@router.get("/stato-riconciliazione")
async def stato_riconciliazione_bonifici():
    """Stato della riconciliazione bonifici."""
    db = Database.get_db()
    
    totale = await db.bonifici_transfers.count_documents({})
    riconciliati = await db.bonifici_transfers.count_documents({"riconciliato": True})
    non_riconciliati = totale - riconciliati
    
    # Totale importi
    pipeline = [
        {"$group": {
            "_id": "$riconciliato",
            "totale": {"$sum": "$importo"},
            "count": {"$sum": 1}
        }}
    ]
    stats = {doc["_id"]: {"totale": doc["totale"], "count": doc["count"]} 
             async for doc in db.bonifici_transfers.aggregate(pipeline)}
    
    return {
        "totale": totale,
        "riconciliati": riconciliati,
        "non_riconciliati": non_riconciliati,
        "percentuale": round(riconciliati / max(totale, 1) * 100, 1),
        "importo_riconciliato": round(stats.get(True, {}).get("totale", 0), 2),
        "importo_non_riconciliato": round(stats.get(False, {}).get("totale", 0) + stats.get(None, {}).get("totale", 0), 2)
    }


@router.get("/dashboard-riconciliazione")
async def dashboard_riconciliazione():
    """
    Dashboard completa riconciliazione con statistiche dettagliate.
    """
    db = Database.get_db()
    
    # === BONIFICI ===
    totale_bonifici = await db.bonifici_transfers.count_documents({})
    bonifici_riconciliati = await db.bonifici_transfers.count_documents({"riconciliato": True})
    bonifici_con_salario = await db.bonifici_transfers.count_documents({"salario_associato": True})
    bonifici_con_fattura = await db.bonifici_transfers.count_documents({"fattura_associata": True})
    
    # Importi bonifici per stato
    pipeline_bonifici = [
        {"$group": {
            "_id": {
                "riconciliato": {"$ifNull": ["$riconciliato", False]},
                "salario": {"$ifNull": ["$salario_associato", False]},
                "fattura": {"$ifNull": ["$fattura_associata", False]}
            },
            "totale": {"$sum": {"$abs": "$importo"}},
            "count": {"$sum": 1}
        }}
    ]
    bonifici_stats = await db.bonifici_transfers.aggregate(pipeline_bonifici).to_list(10)
    
    importo_totale = sum(s.get("totale", 0) for s in bonifici_stats)
    importo_riconciliato = sum(s.get("totale", 0) for s in bonifici_stats if s["_id"].get("riconciliato"))
    
    # === SCADENZE ===
    totale_scadenze = await db.scadenziario_fornitori.count_documents({})
    scadenze_pagate = await db.scadenziario_fornitori.count_documents({"status": "pagata"})
    scadenze_aperte = await db.scadenziario_fornitori.count_documents({"status": {"$ne": "pagata"}})
    
    # === PRIMA NOTA SALARI ===
    totale_salari = await db.prima_nota_salari.count_documents({})
    salari_associati = await db.prima_nota_salari.count_documents({"salario_associato": True})
    
    # === DIPENDENTI CON IBAN ===
    dipendenti_totali = await db.employees.count_documents({})
    dipendenti_con_iban = await db.employees.count_documents({"iban": {"$exists": True, "$nin": [None, ""]}})
    
    # === STATISTICHE PER MESE (ultimi 6 mesi) ===
    from datetime import datetime, timedelta
    sei_mesi_fa = datetime.utcnow() - timedelta(days=180)
    
    pipeline_mensile = [
        {"$match": {"data": {"$gte": sei_mesi_fa.isoformat()}}},
        {"$addFields": {
            "mese": {"$substr": ["$data", 0, 7]}  # "2025-11"
        }},
        {"$group": {
            "_id": "$mese",
            "totale": {"$sum": 1},
            "riconciliati": {"$sum": {"$cond": [{"$eq": ["$riconciliato", True]}, 1, 0]}},
            "importo": {"$sum": {"$abs": "$importo"}}
        }},
        {"$sort": {"_id": -1}},
        {"$limit": 6}
    ]
    stats_mensili = await db.bonifici_transfers.aggregate(pipeline_mensile).to_list(6)
    
    return {
        "bonifici": {
            "totale": totale_bonifici,
            "riconciliati": bonifici_riconciliati,
            "con_salario": bonifici_con_salario,
            "con_fattura": bonifici_con_fattura,
            "non_associati": totale_bonifici - bonifici_con_salario - bonifici_con_fattura,
            "percentuale_riconciliazione": round(bonifici_riconciliati / max(totale_bonifici, 1) * 100, 1),
            "importo_totale": round(importo_totale, 2),
            "importo_riconciliato": round(importo_riconciliato, 2)
        },
        "scadenze": {
            "totale": totale_scadenze,
            "pagate": scadenze_pagate,
            "aperte": scadenze_aperte
        },
        "salari": {
            "totale": totale_salari,
            "associati": salari_associati,
            "percentuale": round(salari_associati / max(totale_salari, 1) * 100, 1)
        },
        "dipendenti": {
            "totale": dipendenti_totali,
            "con_iban": dipendenti_con_iban,
            "percentuale_iban": round(dipendenti_con_iban / max(dipendenti_totali, 1) * 100, 1)
        },
        "trend_mensile": stats_mensili,
        "timestamp": datetime.utcnow().isoformat()
    }


@router.post("/reset-riconciliazione")
async def reset_riconciliazione():
    """Reset dello stato di riconciliazione di tutti i bonifici."""
    db = Database.get_db()
    
    result = await db.bonifici_transfers.update_many(
        {},
        {"$unset": {
            "riconciliato": "",
            "data_riconciliazione": "",
            "movimento_estratto_conto_id": "",
            "movimento_data": "",
            "movimento_descrizione": ""
        }}
    )
    
    return {"success": True, "reset": result.modified_count}


@router.patch("/transfers/{transfer_id}")
async def update_transfer(transfer_id: str, data: Dict[str, Any]):
    """Aggiorna campi di un bonifico (note, associazione cedolino, etc.)."""
    db = Database.get_db()
    
    # Campi permessi per l'aggiornamento
    allowed_fields = ["note", "cedolino_id", "dipendente_id", "categoria"]
    update_data = {k: v for k, v in data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun campo valido da aggiornare")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.bonifici_transfers.update_one(
        {"id": transfer_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    return {"success": True, "updated": True}


@router.get("/download-zip/{year}")
async def download_bonifici_zip(year: str):
    """
    Esporta tutti i bonifici di un anno in formato ZIP.
    I bonifici vengono organizzati in cartelle per dipendente.
    Ogni cartella contiene: XLSX riepilogo + file info per ogni bonifico.
    """
    db = Database.get_db()
    
    # Recupera bonifici dell'anno
    transfers = await db.bonifici_transfers.find(
        {"data": {"$regex": f"^{year}-"}},
        {"_id": 0}
    ).sort("data", 1).to_list(50000)
    
    if not transfers:
        raise HTTPException(status_code=404, detail=f"Nessun bonifico trovato per l'anno {year}")
    
    # Crea ZIP in memoria
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # ============================================
        # ORGANIZZA PER DIPENDENTE
        # ============================================
        
        # Raggruppa bonifici per dipendente
        per_dipendente = {}
        senza_dipendente = []
        
        for t in transfers:
            dip_id = t.get('dipendente_id')
            dip_nome = t.get('dipendente_nome', 'Non Assegnato')
            
            if dip_id:
                if dip_id not in per_dipendente:
                    per_dipendente[dip_id] = {
                        'nome': dip_nome,
                        'bonifici': []
                    }
                per_dipendente[dip_id]['bonifici'].append(t)
            else:
                senza_dipendente.append(t)
        
        # Crea cartella per ogni dipendente
        for dip_id, dip_data in per_dipendente.items():
            # Nome cartella sicuro
            folder_name = re.sub(r'[^A-Za-z0-9\s\-_àèéìòù]', '', dip_data['nome']).strip()
            if not folder_name:
                folder_name = f"Dipendente_{dip_id[:8]}"
            folder_name = folder_name.replace(' ', '_')
            
            bonifici_dip = dip_data['bonifici']
            
            # Crea XLSX per questo dipendente
            wb = Workbook()
            ws = wb.active
            ws.title = f"Bonifici {year}"
            
            headers = ['Data', 'Importo €', 'Causale', 'CRO/TRN', 'Riconciliato', 'Note']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            totale = 0
            for row, t in enumerate(bonifici_dip, 2):
                importo = t.get('importo', 0)
                totale += importo
                ws.cell(row=row, column=1, value=t.get('data', '')[:10])
                ws.cell(row=row, column=2, value=round(importo, 2))
                ws.cell(row=row, column=3, value=t.get('causale', ''))
                ws.cell(row=row, column=4, value=t.get('cro_trn', ''))
                ws.cell(row=row, column=5, value='SI' if t.get('riconciliato') else 'NO')
                ws.cell(row=row, column=6, value=t.get('note', ''))
            
            # Riga totale
            ws.cell(row=len(bonifici_dip)+2, column=1, value='TOTALE')
            ws.cell(row=len(bonifici_dip)+2, column=2, value=round(totale, 2))
            ws.cell(row=len(bonifici_dip)+2, column=1).font = Font(bold=True)
            ws.cell(row=len(bonifici_dip)+2, column=2).font = Font(bold=True)
            
            # Adatta colonne
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 30
            
            xlsx_buffer = io.BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)
            zf.writestr(f"{folder_name}/bonifici_{year}.xlsx", xlsx_buffer.read())
            
            # Crea riepilogo TXT per dipendente
            riepilogo_dip = f"""RIEPILOGO BONIFICI {year}
Dipendente: {dip_data['nome']}
================================

Totale bonifici: {len(bonifici_dip)}
Totale importo: € {totale:,.2f}

Dettaglio:
"""
            for t in bonifici_dip:
                data = t.get('data', '')[:10]
                importo = t.get('importo', 0)
                causale = t.get('causale', '-')[:60]
                riepilogo_dip += f"  {data}  € {importo:,.2f}  {causale}\n"
            
            zf.writestr(f"{folder_name}/riepilogo.txt", riepilogo_dip.encode('utf-8'))
        
        # ============================================
        # BONIFICI NON ASSEGNATI
        # ============================================
        if senza_dipendente:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Non Assegnati {year}"
            
            headers = ['Data', 'Importo €', 'Beneficiario', 'IBAN', 'Causale', 'CRO/TRN', 'Riconciliato']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            totale_na = 0
            for row, t in enumerate(senza_dipendente, 2):
                importo = t.get('importo', 0)
                totale_na += importo
                ws.cell(row=row, column=1, value=t.get('data', '')[:10])
                ws.cell(row=row, column=2, value=round(importo, 2))
                ws.cell(row=row, column=3, value=(t.get('beneficiario') or {}).get('nome', ''))
                ws.cell(row=row, column=4, value=(t.get('beneficiario') or {}).get('iban', ''))
                ws.cell(row=row, column=5, value=t.get('causale', ''))
                ws.cell(row=row, column=6, value=t.get('cro_trn', ''))
                ws.cell(row=row, column=7, value='SI' if t.get('riconciliato') else 'NO')
            
            ws.cell(row=len(senza_dipendente)+2, column=1, value='TOTALE')
            ws.cell(row=len(senza_dipendente)+2, column=2, value=round(totale_na, 2))
            
            xlsx_buffer = io.BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)
            zf.writestr("_NON_ASSEGNATI/bonifici_non_assegnati.xlsx", xlsx_buffer.read())
        
        # ============================================
        # FILE RIEPILOGO GENERALE
        # ============================================
        totale_generale = sum(t.get('importo', 0) for t in transfers)
        riconciliati = sum(1 for t in transfers if t.get('riconciliato'))
        
        riepilogo = f"""
RIEPILOGO BONIFICI {year}
========================

Totale bonifici: {len(transfers)}
Totale importo: € {totale_generale:,.2f}

Riconciliati: {riconciliati}
Non riconciliati: {len(transfers) - riconciliati}

PER DIPENDENTE:
"""
        for dip_id, dip_data in sorted(per_dipendente.items(), key=lambda x: x[1]['nome']):
            tot_dip = sum(b.get('importo', 0) for b in dip_data['bonifici'])
            riepilogo += f"  - {dip_data['nome']}: {len(dip_data['bonifici'])} bonifici, € {tot_dip:,.2f}\n"
        
        if senza_dipendente:
            tot_na = sum(b.get('importo', 0) for b in senza_dipendente)
            riepilogo += f"\n  - NON ASSEGNATI: {len(senza_dipendente)} bonifici, € {tot_na:,.2f}\n"
        
        riepilogo += f"""
Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}
"""
        zf.writestr(f"_RIEPILOGO_GENERALE_{year}.txt", riepilogo.encode('utf-8'))
        
        # ============================================
        # XLSX COMPLETO (per compatibilità)
        # ============================================
        wb = Workbook()
        ws = wb.active
        ws.title = f"Tutti i Bonifici {year}"
        
        headers = ['Data', 'Importo €', 'Dipendente', 'Beneficiario', 'IBAN', 'Causale', 'CRO/TRN', 'Riconciliato', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for row, t in enumerate(transfers, 2):
            ws.cell(row=row, column=1, value=t.get('data', '')[:10])
            ws.cell(row=row, column=2, value=round(t.get('importo', 0), 2))
            ws.cell(row=row, column=3, value=t.get('dipendente_nome', ''))
            ws.cell(row=row, column=4, value=(t.get('beneficiario') or {}).get('nome', ''))
            ws.cell(row=row, column=5, value=(t.get('beneficiario') or {}).get('iban', ''))
            ws.cell(row=row, column=6, value=t.get('causale', ''))
            ws.cell(row=row, column=7, value=t.get('cro_trn', ''))
            ws.cell(row=row, column=8, value='SI' if t.get('riconciliato') else 'NO')
            ws.cell(row=row, column=9, value=t.get('note', ''))
        
        ws.cell(row=len(transfers)+2, column=1, value='TOTALE')
        ws.cell(row=len(transfers)+2, column=2, value=round(totale_generale, 2))
        
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        zf.writestr(f"_TUTTI_BONIFICI_{year}.xlsx", xlsx_buffer.read())
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=bonifici_{year}_per_dipendente.zip"}
    )



# =============================================================================
# ASSOCIAZIONE BONIFICI - DIPENDENTI
# =============================================================================

@router.post("/associa-dipendenti")
async def associa_bonifici_dipendenti(dry_run: bool = Query(True)):
    """
    Associa automaticamente i bonifici ai dipendenti in base alla causale.
    Usa il nome/cognome del dipendente per il matching.
    
    - dry_run=True: mostra solo le associazioni trovate senza salvare
    - dry_run=False: salva le associazioni nel database
    """
    db = Database.get_db()
    
    # Carica dipendenti
    dipendenti = []
    async for dip in db.employees.find({}):
        nome = dip.get('nome_completo') or f"{dip.get('nome', '')} {dip.get('cognome', '')}".strip()
        nome_lower = nome.lower()
        
        parts = nome_lower.split()
        cognome = dip.get('cognome', '').lower() or (parts[-1] if parts else '')
        nome_proprio = dip.get('nome', '').lower() or (parts[0] if len(parts) > 1 else '')
        
        dipendenti.append({
            'id': dip.get('id'),
            'nome_completo': nome,
            'nome': nome_proprio,
            'cognome': cognome
        })
    
    # Carica bonifici senza associazione
    bonifici = await db.bonifici_transfers.find(
        {"$or": [{"dipendente_id": None}, {"dipendente_id": {"$exists": False}}]},
        {"_id": 0}
    ).to_list(10000)
    
    associazioni = []
    non_trovati = []
    
    for bon in bonifici:
        causale = (bon.get('causale') or '').lower()
        matched_dip = None
        best_score = 0
        
        for dip in dipendenti:
            score = 0
            
            # Score per cognome trovato
            if dip['cognome'] and len(dip['cognome']) > 2 and dip['cognome'] in causale:
                score += 2
            
            # Score per nome trovato
            if dip['nome'] and len(dip['nome']) > 2 and dip['nome'] in causale:
                score += 2
            
            # Bonus se entrambi trovati
            if score == 4:
                score += 2
            
            if score > best_score:
                best_score = score
                matched_dip = dip
        
        if matched_dip and best_score >= 2:
            associazioni.append({
                'bonifico_id': bon.get('id'),
                'dipendente_id': matched_dip['id'],
                'dipendente_nome': matched_dip['nome_completo'],
                'importo': bon.get('importo'),
                'causale': bon.get('causale'),
                'data': str(bon.get('data'))[:10],
                'score': best_score
            })
            
            if not dry_run:
                await db.bonifici_transfers.update_one(
                    {"id": bon.get('id')},
                    {"$set": {
                        "dipendente_id": matched_dip['id'],
                        "dipendente_nome": matched_dip['nome_completo'],
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
        else:
            non_trovati.append({
                'bonifico_id': bon.get('id'),
                'causale': bon.get('causale'),
                'importo': bon.get('importo')
            })
    
    return {
        "success": True,
        "dry_run": dry_run,
        "totale_bonifici": len(bonifici),
        "associati": len(associazioni),
        "non_trovati": len(non_trovati),
        "associazioni": associazioni,
        "non_trovati_list": non_trovati[:20]
    }


@router.get("/dipendente/{dipendente_id}")
async def get_bonifici_dipendente(dipendente_id: str):
    """
    Restituisce tutti i bonifici associati a un dipendente.
    """
    db = Database.get_db()
    
    # Verifica che il dipendente esista
    dipendente = await db.employees.find_one({"id": dipendente_id}, {"_id": 0})
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    nome = dipendente.get('nome_completo') or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}".strip()
    
    # Trova bonifici associati
    bonifici = await db.bonifici_transfers.find(
        {"dipendente_id": dipendente_id},
        {"_id": 0}
    ).sort("data", -1).to_list(1000)
    
    # Calcola totale
    totale = sum(b.get('importo', 0) for b in bonifici)
    
    return {
        "dipendente_id": dipendente_id,
        "dipendente_nome": nome,
        "totale_bonifici": len(bonifici),
        "totale_importo": round(totale, 2),
        "bonifici": bonifici
    }


@router.post("/associa-manuale/{bonifico_id}")
async def associa_manuale_bonifico(bonifico_id: str, dipendente_id: str):
    """
    Associa manualmente un bonifico a un dipendente.
    Se il bonifico ha un IBAN beneficiario, lo aggiunge automaticamente 
    all'anagrafica del dipendente.
    """
    db = Database.get_db()
    
    # Verifica bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    # Verifica dipendente
    dipendente = await db.employees.find_one({"id": dipendente_id}, {"_id": 0})
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    nome = dipendente.get('nome_completo') or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}".strip()
    
    # Aggiorna bonifico
    await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$set": {
            "dipendente_id": dipendente_id,
            "dipendente_nome": nome,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # SINCRONIZZAZIONE IBAN: Se il bonifico ha un IBAN beneficiario, aggiungilo all'anagrafica
    iban_sync_info = None
    iban_beneficiario = bonifico.get("beneficiario", {}).get("iban")
    if iban_beneficiario and len(iban_beneficiario) >= 15:
        iban_clean = iban_beneficiario.upper().replace(" ", "")
        
        # Leggi IBAN esistenti
        existing_ibans = set(dipendente.get("ibans", []))
        if dipendente.get("iban"):
            existing_ibans.add(dipendente["iban"].upper().replace(" ", ""))
        
        # Aggiungi nuovo IBAN se non già presente
        if iban_clean not in existing_ibans and len(existing_ibans) < 3:
            existing_ibans.add(iban_clean)
            
            await db.employees.update_one(
                {"id": dipendente_id},
                {"$set": {
                    "ibans": list(existing_ibans)[:3],
                    "iban": list(existing_ibans)[0] if existing_ibans else "",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            iban_sync_info = f"IBAN {iban_clean} aggiunto all'anagrafica"
    
    return {
        "success": True,
        "bonifico_id": bonifico_id,
        "dipendente_id": dipendente_id,
        "dipendente_nome": nome,
        "iban_sync": iban_sync_info
    }


@router.delete("/disassocia/{bonifico_id}")
async def disassocia_bonifico(bonifico_id: str):
    """
    Rimuove l'associazione di un bonifico da un dipendente.
    """
    db = Database.get_db()
    
    result = await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$unset": {"dipendente_id": "", "dipendente_nome": ""}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    return {"success": True, "bonifico_id": bonifico_id}


@router.post("/sync-iban-anagrafica")
async def sync_iban_bonifici_to_anagrafica():
    """
    Sincronizza gli IBAN dai bonifici all'anagrafica dipendenti.
    Se un bonifico è associato a un dipendente e ha un IBAN beneficiario,
    lo aggiunge all'array ibans del dipendente (se non già presente).
    """
    db = Database.get_db()
    
    # Trova tutti i bonifici con dipendente_id e IBAN beneficiario
    bonifici = await db.bonifici_transfers.find(
        {
            "dipendente_id": {"$exists": True, "$ne": None},
            "beneficiario.iban": {"$exists": True, "$nin": [None, ""]}
        },
        {"dipendente_id": 1, "beneficiario.iban": 1, "_id": 0}
    ).to_list(10000)
    
    # Raggruppa IBAN per dipendente
    iban_per_dipendente = {}
    for bon in bonifici:
        dip_id = bon.get("dipendente_id")
        iban = bon.get("beneficiario", {}).get("iban")
        if dip_id and iban and len(iban) >= 15:
            if dip_id not in iban_per_dipendente:
                iban_per_dipendente[dip_id] = set()
            iban_per_dipendente[dip_id].add(iban.upper().replace(" ", ""))
    
    aggiornati = 0
    for dip_id, new_ibans in iban_per_dipendente.items():
        # Leggi IBAN esistenti del dipendente
        dipendente = await db.employees.find_one({"id": dip_id}, {"ibans": 1, "iban": 1, "_id": 0})
        if not dipendente:
            continue
        
        existing_ibans = set(dipendente.get("ibans", []))
        if dipendente.get("iban"):
            existing_ibans.add(dipendente["iban"].upper().replace(" ", ""))
        
        # Combina vecchi e nuovi IBAN
        all_ibans = existing_ibans.union(new_ibans)
        
        # Aggiorna solo se ci sono nuovi IBAN
        if all_ibans != existing_ibans:
            await db.employees.update_one(
                {"id": dip_id},
                {"$set": {
                    "ibans": list(all_ibans)[:3],  # Max 3 IBAN
                    "iban": list(all_ibans)[0] if all_ibans else "",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            aggiornati += 1
    
    return {
        "success": True,
        "dipendenti_aggiornati": aggiornati,
        "totale_bonifici_analizzati": len(bonifici),
        "dipendenti_con_iban": len(iban_per_dipendente)
    }


# =============================================================================
# ASSOCIAZIONE BONIFICI ↔ PRIMA NOTA SALARI
# =============================================================================

@router.get("/operazioni-salari/{bonifico_id}")
async def get_operazioni_salari_compatibili(bonifico_id: str):
    """
    Cerca operazioni in Prima Nota Salari compatibili con il bonifico.
    Priorità di matching:
    1. IBAN del beneficiario = IBAN del dipendente (score +100)
    2. Importo simile (±10%)
    3. Nome dipendente in causale/beneficiario
    
    Esclude operazioni già associate ad altri bonifici.
    """
    db = Database.get_db()
    
    # Recupera bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id}, {"_id": 0})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    # Fix: gestisce None esplicito
    raw_importo = bonifico.get("importo")
    importo = abs(float(raw_importo)) if raw_importo is not None else 0
    data_bonifico = bonifico.get("data", "")
    causale = (bonifico.get("causale") or "").lower()
    beneficiario_nome = ((bonifico.get("beneficiario") or {}).get("nome") or "").lower()
    beneficiario_iban = ((bonifico.get("beneficiario") or {}).get("iban") or "").upper().replace(" ", "")
    
    # === STEP 1: Cerca dipendente per IBAN ===
    dipendente_iban_match = None
    if beneficiario_iban and len(beneficiario_iban) > 10:
        # Cerca dipendente con IBAN corrispondente (case-insensitive)
        dipendente_iban_match = await db.employees.find_one(
            {"iban": {"$regex": f"^{beneficiario_iban}$", "$options": "i"}},
            {"_id": 0, "id": 1, "nome_completo": 1, "nome": 1, "cognome": 1, "iban": 1}
        )
        if dipendente_iban_match:
            dipendente_iban_match["nome_display"] = (
                dipendente_iban_match.get("nome_completo") or 
                f"{dipendente_iban_match.get('nome', '')} {dipendente_iban_match.get('cognome', '')}".strip()
            )
    
    # === STEP 2: Pipeline per operazioni compatibili ===
    match_stage = {
        "salario_associato": {"$ne": True},
        "$or": [
            {"bonifico_id": {"$exists": False}},
            {"bonifico_id": None},
            {"bonifico_id": ""}
        ]
    }
    
    if importo > 0:
        match_stage["$and"] = [
            {"$or": [
                {"bonifico_id": {"$exists": False}},
                {"bonifico_id": None},
                {"bonifico_id": ""}
            ]},
            {"$or": [
                {"importo_busta": {"$gte": importo * 0.9, "$lte": importo * 1.1}},
                {"importo_bonifico": {"$gte": importo * 0.9, "$lte": importo * 1.1}},
                {"importo": {"$gte": importo * 0.9, "$lte": importo * 1.1}},
                {"netto": {"$gte": importo * 0.9, "$lte": importo * 1.1}}
            ]}
        ]
        del match_stage["$or"]
    
    pipeline = [
        {"$match": match_stage},
        {"$group": {
            "_id": {
                "dipendente": "$dipendente",
                "importo_busta": "$importo_busta",
                "anno": "$anno",
                "mese": "$mese"
            },
            "doc": {"$first": "$$ROOT"}
        }},
        {"$replaceRoot": {"newRoot": "$doc"}},
        {"$project": {"_id": 0}},
        {"$sort": {"anno": -1, "mese": -1}},
        {"$limit": 50}
    ]
    
    operazioni = await db.prima_nota_salari.aggregate(pipeline).to_list(50)
    
    # === STEP 3: Calcola score di compatibilità ===
    risultati = []
    seen_keys = set()
    
    for op in operazioni:
        # Chiave univoca per deduplicazione
        key = f"{op.get('dipendente', '')}_{op.get('importo_busta', 0)}_{op.get('anno', '')}_{op.get('mese', '')}"
        if key in seen_keys:
            continue
        seen_keys.add(key)
        
        score = 0
        iban_match = False
        op_importo = op.get("importo_busta") or op.get("importo_bonifico") or op.get("importo") or op.get("netto", 0)
        op_data = op.get("data", "")
        dipendente = op.get("dipendente", "").lower()
        
        # === BONUS IBAN (max 100) ===
        # Se il dipendente dell'operazione corrisponde al dipendente trovato per IBAN
        if dipendente_iban_match and dipendente:
            nome_iban = dipendente_iban_match.get("nome_display", "").lower()
            # Match nome dipendente (fuzzy)
            if nome_iban:
                # Match esatto o parziale
                parti_nome = nome_iban.split()
                parti_dip = dipendente.split()
                match_count = sum(1 for p in parti_nome if any(p in d or d in p for d in parti_dip if len(d) > 2))
                if match_count >= 2 or nome_iban in dipendente or dipendente in nome_iban:
                    score += 100
                    iban_match = True
        
        # Score importo (max 50)
        if importo > 0 and op_importo > 0:
            diff_pct = abs(importo - op_importo) / importo * 100
            if diff_pct < 1:
                score += 50
            elif diff_pct < 5:
                score += 40
            elif diff_pct < 10:
                score += 20
        
        # Score data (max 30)
        if data_bonifico and op_data:
            try:
                from datetime import timedelta
                d1 = datetime.fromisoformat(str(data_bonifico).replace("Z", "+00:00")) if isinstance(data_bonifico, str) else data_bonifico
                d2 = datetime.fromisoformat(str(op_data).replace("Z", "+00:00")) if isinstance(op_data, str) else op_data
                diff_days = abs((d1 - d2).days)
                if diff_days <= 3:
                    score += 30
                elif diff_days <= 7:
                    score += 20
                elif diff_days <= 30:
                    score += 10
            except:
                pass
        
        # Score causale/beneficiario match (max 30)
        if dipendente and not iban_match:
            if dipendente in causale:
                score += 30
            elif beneficiario_nome and dipendente in beneficiario_nome:
                score += 25
            else:
                cognome = dipendente.split()[-1] if dipendente else ""
                if cognome and len(cognome) > 3:
                    if cognome in causale or (beneficiario_nome and cognome in beneficiario_nome):
                        score += 15
        
        # Bonus per parole chiave stipendio
        if any(word in causale for word in ["stipendio", "salario", "netto", "busta", "stip"]):
            score += 10
        
        # Include risultato
        if score >= 0:
            risultati.append({
                **op,
                "compatibilita_score": max(score, 10) if op_importo > 0 else score,
                "importo_display": op_importo,
                "iban_match": iban_match
            })
    
    # Ordina per score (IBAN match prima)
    risultati.sort(key=lambda x: (x.get("iban_match", False), x.get("compatibilita_score", 0)), reverse=True)
    
    return {
        "bonifico": {
            "id": bonifico_id,
            "importo": importo,
            "data": data_bonifico,
            "causale": bonifico.get("causale"),
            "beneficiario": beneficiario_nome,
            "beneficiario_iban": beneficiario_iban
        },
        "dipendente_iban_match": dipendente_iban_match,
        "operazioni_compatibili": risultati[:15]
    }


@router.post("/associa-salario")
async def associa_bonifico_salario(
    bonifico_id: str = Query(...),
    operazione_id: str = Query(...)
):
    """
    Associa un bonifico a un'operazione di Prima Nota Salari.
    Aggiorna anche dipendente_id per mostrare il bonifico nella pagina dipendenti.
    """
    db = Database.get_db()
    
    # Verifica bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    # Verifica operazione
    operazione = await db.prima_nota_salari.find_one({"id": operazione_id}, {"_id": 0})
    if not operazione:
        raise HTTPException(status_code=404, detail="Operazione salari non trovata")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Trova dipendente_id dall'operazione o cercando per nome
    dipendente_id = operazione.get("dipendente_id")
    dipendente_nome = operazione.get("dipendente") or operazione.get("descrizione")
    
    # Se non c'è dipendente_id nell'operazione, cerca per nome
    if not dipendente_id and dipendente_nome:
        # Cerca dipendente per nome
        dipendente = await db.employees.find_one(
            {"$or": [
                {"nome_completo": {"$regex": dipendente_nome, "$options": "i"}},
                {"cognome": {"$regex": dipendente_nome.split()[-1] if dipendente_nome else "", "$options": "i"}}
            ]},
            {"_id": 0, "id": 1, "nome_completo": 1, "nome": 1, "cognome": 1}
        )
        if dipendente:
            dipendente_id = dipendente.get("id")
            dipendente_nome = dipendente.get("nome_completo") or f"{dipendente.get('nome', '')} {dipendente.get('cognome', '')}".strip()
    
    # Aggiorna bonifico con dipendente_id
    update_data = {
        "salario_associato": True,
        "operazione_salario_id": operazione_id,
        "operazione_salario_desc": dipendente_nome,
        "operazione_salario_data": operazione.get("data"),
        "operazione_salario_importo": operazione.get("importo") or operazione.get("netto"),
        "data_associazione": now,
        "updated_at": now
    }
    
    # Aggiungi dipendente_id se trovato
    if dipendente_id:
        update_data["dipendente_id"] = dipendente_id
        update_data["dipendente_nome"] = dipendente_nome
    
    await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$set": update_data}
    )
    
    # Aggiorna operazione salari con bonifico_id
    await db.prima_nota_salari.update_one(
        {"id": operazione_id},
        {"$set": {
            "bonifico_associato": True,
            "bonifico_id": bonifico_id,
            "updated_at": now
        }}
    )
    
    return {
        "success": True,
        "bonifico_id": bonifico_id,
        "operazione_id": operazione_id,
        "dipendente_id": dipendente_id,
        "descrizione": dipendente_nome
    }


@router.delete("/disassocia-salario/{bonifico_id}")
async def disassocia_bonifico_salario(bonifico_id: str):
    """
    Rimuove l'associazione bonifico ↔ operazione salari.
    """
    db = Database.get_db()
    
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    operazione_id = bonifico.get("operazione_salario_id")
    
    # Rimuovi associazione dal bonifico
    await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$unset": {
            "salario_associato": "",
            "operazione_salario_id": "",
            "operazione_salario_desc": "",
            "operazione_salario_data": "",
            "operazione_salario_importo": "",
            "data_associazione": ""
        }}
    )
    
    # Rimuovi riferimento dall'operazione
    if operazione_id:
        await db.prima_nota_salari.update_one(
            {"id": operazione_id},
            {"$unset": {"bonifico_associato": "", "bonifico_id": ""}}
        )
    
    return {"success": True, "bonifico_id": bonifico_id}



@router.post("/transfers/{bonifico_id}/link-estratto-conto")
async def link_bonifico_to_estratto_conto(
    bonifico_id: str, 
    estratto_conto_id: Optional[str] = None
) -> Dict[str, Any]:
    """
    Collega un bonifico PDF a un movimento dell'estratto conto.
    Se estratto_conto_id non è fornito, cerca automaticamente per importo e data.
    """
    db = Database.get_db()
    
    # Trova il bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id}, {"_id": 0})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    movimento_ec = None
    
    if estratto_conto_id:
        # Link manuale
        movimento_ec = await db["estratto_conto"].find_one({"id": estratto_conto_id}, {"_id": 0})
        if not movimento_ec:
            raise HTTPException(status_code=404, detail="Movimento estratto conto non trovato")
    else:
        # Cerca automaticamente
        importo_bonifico = bonifico.get("importo", 0)
        data_bonifico = bonifico.get("data")
        
        if isinstance(data_bonifico, datetime):
            data_str = data_bonifico.strftime("%Y-%m-%d")
        elif isinstance(data_bonifico, str):
            data_str = data_bonifico[:10]
        else:
            data_str = None
        
        # Cerca movimento per importo simile e data
        query = {"tipo": "uscita"}
        
        if importo_bonifico:
            query["importo"] = {
                "$gte": abs(importo_bonifico) - 1,
                "$lte": abs(importo_bonifico) + 1
            }
        
        if data_str:
            # Cerca nella stessa data o giorno prima/dopo
            query["data"] = {"$regex": f"^{data_str[:8]}"}  # Stesso mese
        
        # Preferisci movimenti non ancora associati
        candidates = await db["estratto_conto"].find(
            {**query, "bonifico_pdf_id": {"$exists": False}},
            {"_id": 0}
        ).limit(10).to_list(10)
        
        # Match per importo esatto prima
        for c in candidates:
            if abs(c.get("importo", 0) - abs(importo_bonifico)) < 0.10:
                movimento_ec = c
                break
        
        if not movimento_ec and candidates:
            movimento_ec = candidates[0]
    
    if not movimento_ec:
        return {
            "success": False,
            "message": "Nessun movimento estratto conto trovato per questo bonifico"
        }
    
    # Aggiorna il bonifico con il link all'estratto conto
    await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$set": {
            "estratto_conto_id": movimento_ec.get("id"),
            "estratto_conto_data": movimento_ec.get("data"),
            "estratto_conto_importo": movimento_ec.get("importo"),
            "estratto_conto_desc": movimento_ec.get("descrizione", "")[:100],
            "linked_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Aggiorna l'estratto conto con il link al bonifico PDF
    await db["estratto_conto"].update_one(
        {"id": movimento_ec.get("id")},
        {"$set": {
            "bonifico_pdf_id": bonifico_id,
            "bonifico_pdf_file": bonifico.get("file_name"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "success": True,
        "bonifico_id": bonifico_id,
        "estratto_conto_id": movimento_ec.get("id"),
        "message": f"Bonifico collegato al movimento del {movimento_ec.get('data')}"
    }


@router.post("/riconcilia-automatico")
async def riconcilia_bonifici_automatico(anno: Optional[int] = Query(None)) -> Dict[str, Any]:
    """
    Riconcilia automaticamente i bonifici PDF con i movimenti dell'estratto conto
    basandosi su importo, data e beneficiario.
    """
    db = Database.get_db()
    
    # Trova bonifici non ancora collegati
    query_bonifici = {"estratto_conto_id": {"$exists": False}}
    if anno:
        query_bonifici["data"] = {"$regex": f"^{anno}-"}
    
    bonifici = await db.bonifici_transfers.find(query_bonifici, {"_id": 0}).to_list(5000)
    
    # Trova movimenti estratto conto non collegati
    query_ec = {
        "tipo": "uscita",
        "bonifico_pdf_id": {"$exists": False}
    }
    if anno:
        query_ec["data"] = {"$regex": f"^{anno}"}
    
    movimenti_ec = await db["estratto_conto"].find(query_ec, {"_id": 0}).to_list(10000)
    
    # Indicizza movimenti per importo
    ec_by_importo = {}
    for m in movimenti_ec:
        imp = round(abs(m.get("importo", 0)), 2)
        if imp not in ec_by_importo:
            ec_by_importo[imp] = []
        ec_by_importo[imp].append(m)
    
    riconciliati = 0
    non_trovati = []
    
    for bonifico in bonifici:
        importo = round(abs(bonifico.get("importo", 0) or 0), 2)
        data_bonifico = bonifico.get("data")
        
        if isinstance(data_bonifico, datetime):
            data_str = data_bonifico.strftime("%Y-%m-%d")
        elif isinstance(data_bonifico, str):
            data_str = str(data_bonifico)[:10]
        else:
            data_str = None
        
        # Cerca per importo
        match = None
        candidates = ec_by_importo.get(importo, [])
        
        for c in candidates:
            c_data = c.get("data", "")[:10]
            # Preferisci match con stessa data o giorno vicino
            if data_str and c_data:
                if c_data == data_str or abs((datetime.fromisoformat(c_data) - datetime.fromisoformat(data_str[:10])).days) <= 3:
                    match = c
                    break
        
        if not match and candidates:
            # Prendi il primo disponibile
            match = candidates[0]
        
        if match:
            # Collega
            await db.bonifici_transfers.update_one(
                {"id": bonifico.get("id")},
                {"$set": {
                    "estratto_conto_id": match.get("id"),
                    "estratto_conto_data": match.get("data"),
                    "estratto_conto_importo": match.get("importo"),
                    "linked_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            await db["estratto_conto"].update_one(
                {"id": match.get("id")},
                {"$set": {
                    "bonifico_pdf_id": bonifico.get("id"),
                    "bonifico_pdf_file": bonifico.get("source_file") or bonifico.get("file_name"),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Rimuovi dalla lista per evitare duplicati


# =============================================================================
# ASSOCIAZIONE BONIFICI A FATTURE
# =============================================================================

@router.get("/fatture-compatibili/{bonifico_id}")
async def get_fatture_compatibili(bonifico_id: str):
    """
    Cerca fatture ricevute compatibili con il bonifico.
    Filtra per:
    - Importo simile (±10%)
    - Data vicina (±60 giorni)
    - Fornitore nel beneficiario/causale
    - Non già associate ad altri bonifici
    """
    db = Database.get_db()
    
    # Recupera bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id}, {"_id": 0})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    # Fix: gestisce None esplicito
    raw_importo = bonifico.get("importo")
    importo = abs(float(raw_importo)) if raw_importo is not None else 0
    data_bonifico = bonifico.get("data", "")
    causale = (bonifico.get("causale") or "").lower()
    beneficiario = ((bonifico.get("beneficiario") or {}).get("nome") or "").lower()
    
    # Pipeline con deduplicazione per numero fattura + fornitore
    # IMPORTANTE: Escludi fatture già associate a QUALSIASI bonifico
    match_stage = {
        "fattura_associata": {"$ne": True},
        "$or": [
            {"bonifico_id": {"$exists": False}},
            {"bonifico_id": None},
            {"bonifico_id": ""},
            {"bonifico_associato": {"$ne": True}}
        ]
    }
    
    if importo > 0:
        match_stage["$and"] = [
            {"$or": [
                {"bonifico_id": {"$exists": False}},
                {"bonifico_id": None},
                {"bonifico_id": ""},
                {"bonifico_associato": {"$ne": True}}
            ]},
            {"$or": [
                {"totale_documento": {"$gte": importo * 0.85, "$lte": importo * 1.15}},
                {"importo_totale": {"$gte": importo * 0.85, "$lte": importo * 1.15}},
                {"total_amount": {"$gte": importo * 0.85, "$lte": importo * 1.15}}
            ]}
        ]
        # Rimuovi il $or originale perché ora è in $and
        del match_stage["$or"]
    
    pipeline = [
        {"$match": match_stage},
        # Deduplica per numero fattura + P.IVA fornitore
        {"$group": {
            "_id": {
                "numero": {"$ifNull": ["$invoice_number", "$numero_documento"]},
                "piva": {"$ifNull": ["$supplier_vat", "$fornitore.partita_iva"]}
            },
            "doc": {"$first": "$$ROOT"}
        }},
        {"$replaceRoot": {"newRoot": "$doc"}},
        {"$project": {"_id": 0}},  # Esclude _id DOPO replaceRoot
        {"$sort": {"invoice_date": -1}},
        {"$limit": 50}
    ]
    
    fatture = await db.invoices.aggregate(pipeline).to_list(50)
    
    # Calcola score di compatibilità
    risultati = []
    seen_keys = set()  # Extra deduplicazione
    
    for f in fatture:
        # Chiave univoca
        num = f.get("invoice_number") or f.get("numero_documento", "")
        piva = f.get("supplier_vat") or (f.get("fornitore", {}) or {}).get("partita_iva", "")
        key = f"{num}_{piva}"
        if key in seen_keys:
            continue
        seen_keys.add(key)
        
        score = 0
        f_importo = f.get("totale_documento") or f.get("importo_totale") or f.get("total_amount", 0)
        f_data = f.get("data_documento") or f.get("invoice_date", "")
        fornitore = (f.get("fornitore", {}) or {})
        f_fornitore = (fornitore.get("denominazione") or fornitore.get("ragione_sociale") or f.get("supplier_name") or "").lower()
        
        # Score importo (max 50)
        if importo > 0 and f_importo > 0:
            diff_pct = abs(importo - f_importo) / importo * 100
            if diff_pct < 1:
                score += 50
            elif diff_pct < 5:
                score += 40
            elif diff_pct < 10:
                score += 25
            elif diff_pct < 15:
                score += 15
        
        # Score fornitore match (max 40)
        if f_fornitore:
            # Match esatto nel beneficiario
            if f_fornitore in beneficiario or beneficiario in f_fornitore:
                score += 40
            # Match nella causale
            elif f_fornitore in causale:
                score += 30
            # Match parziale (prime parole)
            else:
                parole_fornitore = [p for p in f_fornitore.split() if len(p) > 3]
                for parola in parole_fornitore[:2]:
                    if parola in causale or parola in beneficiario:
                        score += 15
                        break
        
        # Score data (max 10)
        if data_bonifico and f_data:
            try:
                d1 = datetime.fromisoformat(str(data_bonifico).replace("Z", "+00:00")) if isinstance(data_bonifico, str) else data_bonifico
                d2 = datetime.fromisoformat(str(f_data).replace("Z", "+00:00")) if isinstance(f_data, str) else f_data
                diff_days = abs((d1 - d2).days)
                if diff_days <= 7:
                    score += 10
                elif diff_days <= 30:
                    score += 5
            except:
                pass
        
        if score > 0:
            risultati.append({
                "id": f.get("id"),
                "numero_fattura": f.get("numero_documento") or f.get("invoice_number"),
                "data_fattura": f_data,
                "fornitore": f_fornitore.title() if f_fornitore else "N/A",
                "importo": f_importo,
                "compatibilita_score": score,
                "collection": "invoices"  # Collezione consolidata
            })
    
    # Ordina per score
    risultati.sort(key=lambda x: x.get("compatibilita_score", 0), reverse=True)
    
    return {
        "bonifico": {
            "id": bonifico_id,
            "importo": importo,
            "data": data_bonifico,
            "causale": bonifico.get("causale"),
            "beneficiario": beneficiario
        },
        "fatture_compatibili": risultati[:15]
    }


@router.post("/associa-fattura")
async def associa_bonifico_fattura(
    bonifico_id: str = Query(...),
    fattura_id: str = Query(...),
    collection: str = Query(default="invoices")
):
    """
    Associa un bonifico a una fattura ricevuta.
    """
    db = Database.get_db()
    now = datetime.now(timezone.utc).isoformat()
    
    # Verifica bonifico
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    # Verifica fattura
    coll = db[collection]
    fattura = await coll.find_one({"id": fattura_id})
    if not fattura:
        raise HTTPException(status_code=404, detail="Fattura non trovata")
    
    # Aggiorna bonifico
    await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$set": {
            "fattura_associata": True,
            "fattura_id": fattura_id,
            "fattura_numero": fattura.get("numero_documento") or fattura.get("invoice_number"),
            "fattura_fornitore": (fattura.get("fornitore", {}) or {}).get("denominazione") or fattura.get("supplier_name"),
            "fattura_importo": fattura.get("totale_documento") or fattura.get("importo_totale") or fattura.get("total_amount"),
            "data_associazione": now,
            "updated_at": now
        }}
    )
    
    # Aggiorna fattura
    await coll.update_one(
        {"id": fattura_id},
        {"$set": {
            "bonifico_associato": True,
            "bonifico_id": bonifico_id,
            "updated_at": now
        }}
    )
    
    return {
        "success": True,
        "bonifico_id": bonifico_id,
        "fattura_id": fattura_id,
        "fattura_numero": fattura.get("numero_documento") or fattura.get("invoice_number")
    }


@router.delete("/disassocia-fattura/{bonifico_id}")
async def disassocia_bonifico_fattura(bonifico_id: str):
    """
    Rimuove l'associazione bonifico ↔ fattura.
    """
    db = Database.get_db()
    
    bonifico = await db.bonifici_transfers.find_one({"id": bonifico_id})
    if not bonifico:
        raise HTTPException(status_code=404, detail="Bonifico non trovato")
    
    fattura_id = bonifico.get("fattura_id")
    
    # Rimuovi associazione dal bonifico
    await db.bonifici_transfers.update_one(
        {"id": bonifico_id},
        {"$unset": {
            "fattura_associata": "",
            "fattura_id": "",
            "fattura_numero": "",
            "fattura_fornitore": "",
            "fattura_importo": "",
            "data_associazione": ""
        }}
    )
    
    # Rimuovi riferimento dalla fattura
    if fattura_id:
        for coll_name in ["invoices"]:  # Consolidato: solo invoices
            await db[coll_name].update_one(
                {"id": fattura_id},
                {"$unset": {"bonifico_associato": "", "bonifico_id": ""}}
            )
    
    return {"success": True, "bonifico_id": bonifico_id}

