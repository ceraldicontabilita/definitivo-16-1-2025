#!/usr/bin/env python3
"""
Script per importare i dati Excel nella Prima Nota
Logica contabile:
- CASSA:
  - Corrispettivi -> DARE (entrata) - incassi giornalieri
  - POS -> AVERE (uscita) - i soldi escono dalla cassa per andare in banca
  - Versamenti -> AVERE (uscita) - versamento in banca
  - Finanziamento soci -> DARE (entrata)
  
- BANCA:
  - Versamenti -> DARE (entrata) - arrivano dalla cassa
  - POS accreditati -> DARE (entrata) - accredito POS
"""

import openpyxl
import requests
import os
from datetime import datetime
from io import BytesIO

API_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://invoice-flow-64.preview.emergentagent.com')

def download_excel(url):
    """Download Excel file from URL"""
    response = requests.get(url)
    response.raise_for_status()
    return BytesIO(response.content)

def parse_corrispettivi(url):
    """
    Parse corrispettivi.xlsx - vanno in CASSA come DARE (entrata)
    Colonne: Id invio, Matricola dispositivo, Data e ora rilevazione, Data e ora trasmissione, 
             Ammontare delle vendite (totale in euro), Imponibile vendite, Imposta vendite, ...
    """
    print(f"Parsing corrispettivi from {url}")
    file_content = download_excel(url)
    wb = openpyxl.load_workbook(file_content)
    ws = wb.active
    
    movements = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # Skip header
        if row and len(row) >= 5:
            data_rilevazione = row[2]  # Data e ora rilevazione
            ammontare_vendite = row[4]  # Ammontare delle vendite (totale in euro)
            
            if data_rilevazione and ammontare_vendite and ammontare_vendite > 0:
                if isinstance(data_rilevazione, datetime):
                    data_str = data_rilevazione.strftime('%Y-%m-%d')
                else:
                    continue
                
                movements.append({
                    "data": data_str,
                    "tipo": "entrata",  # DARE
                    "importo": round(float(ammontare_vendite), 2),
                    "descrizione": f"Corrispettivo giornaliero",
                    "categoria": "Corrispettivi",
                    "source": "excel_corrispettivi"
                })
    
    print(f"  Found {len(movements)} corrispettivi")
    return movements

def parse_pos(url):
    """
    Parse pos.xlsx - vanno in CASSA come AVERE (uscita) perché escono dalla cassa
    Colonne: DATA, CONTO, IMPORTO
    """
    print(f"Parsing POS from {url}")
    file_content = download_excel(url)
    wb = openpyxl.load_workbook(file_content)
    ws = wb.active
    
    cassa_movements = []  # AVERE (uscita da cassa)
    banca_movements = []  # DARE (entrata in banca quando accreditato)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3:
            data_val = row[0]  # DATA
            importo = row[2]    # IMPORTO
            
            if data_val and importo and float(importo) > 0:
                if isinstance(data_val, datetime):
                    data_str = data_val.strftime('%Y-%m-%d')
                else:
                    continue
                
                # POS in CASSA come AVERE (i soldi escono per andare in banca)
                cassa_movements.append({
                    "data": data_str,
                    "tipo": "uscita",  # AVERE - escono dalla cassa
                    "importo": round(float(importo), 2),
                    "descrizione": f"POS giornaliero - accredito banca",
                    "categoria": "POS",
                    "source": "excel_pos"
                })
                
                # POS in BANCA come DARE (i soldi entrano quando accreditati)
                banca_movements.append({
                    "data": data_str,
                    "tipo": "entrata",  # DARE - entrano in banca
                    "importo": round(float(importo), 2),
                    "descrizione": f"Accredito POS giornaliero",
                    "categoria": "POS",
                    "source": "excel_pos"
                })
    
    print(f"  Found {len(cassa_movements)} POS movements")
    return cassa_movements, banca_movements

def parse_versamenti(url):
    """
    Parse versamento.xlsx - CASSA AVERE (uscita), BANCA DARE (entrata)
    Colonne: DATA, CONTO, IMPORTO
    """
    print(f"Parsing versamenti from {url}")
    file_content = download_excel(url)
    wb = openpyxl.load_workbook(file_content)
    ws = wb.active
    
    cassa_movements = []  # AVERE (uscita da cassa)
    banca_movements = []  # DARE (entrata in banca)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 3:
            data_val = row[0]  # DATA
            importo = row[2]    # IMPORTO
            
            # Solo righe con importo valorizzato
            if data_val and importo and float(importo) > 0:
                if isinstance(data_val, datetime):
                    data_str = data_val.strftime('%Y-%m-%d')
                else:
                    continue
                
                # Versamento USCITA da CASSA
                cassa_movements.append({
                    "data": data_str,
                    "tipo": "uscita",  # AVERE
                    "importo": round(float(importo), 2),
                    "descrizione": f"Versamento in banca",
                    "categoria": "Versamento",
                    "source": "excel_versamento"
                })
                
                # Versamento ENTRATA in BANCA
                banca_movements.append({
                    "data": data_str,
                    "tipo": "entrata",  # DARE
                    "importo": round(float(importo), 2),
                    "descrizione": f"Versamento da cassa",
                    "categoria": "Versamento",
                    "source": "excel_versamento"
                })
    
    print(f"  Found {len(cassa_movements)} versamenti")
    return cassa_movements, banca_movements

def parse_finanziamento_soci(url):
    """
    Parse finanziamento soci.xlsx - vanno in CASSA come DARE (entrata)
    Colonne: Data, Numero Fattura, Fornitore, Descrizione, Entrate, Uscite
    """
    print(f"Parsing finanziamento soci from {url}")
    file_content = download_excel(url)
    wb = openpyxl.load_workbook(file_content)
    ws = wb.active
    
    movements = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5:
            data_val = row[0]    # Data
            fornitore = row[2]   # Fornitore (nome socio)
            descrizione = row[3] # Descrizione
            entrate = row[4]     # Entrate
            
            if data_val and entrate and float(entrate) > 0:
                if isinstance(data_val, datetime):
                    data_str = data_val.strftime('%Y-%m-%d')
                else:
                    continue
                
                desc = f"Finanziamento socio - {fornitore}" if fornitore else "Finanziamento soci"
                if descrizione:
                    desc = f"{desc} ({descrizione})"
                
                movements.append({
                    "data": data_str,
                    "tipo": "entrata",  # DARE - entrano soldi in cassa
                    "importo": round(float(entrate), 2),
                    "descrizione": desc,
                    "categoria": "Finanziamento soci",
                    "source": "excel_finanziamento"
                })
    
    print(f"  Found {len(movements)} finanziamenti soci")
    return movements

def import_to_api(cassa_movements, banca_movements):
    """Import movements to Prima Nota API"""
    payload = {
        "cassa": cassa_movements,
        "banca": banca_movements
    }
    
    response = requests.post(
        f"{API_URL}/api/prima-nota/import-batch",
        json=payload
    )
    
    return response.json()

def main():
    # URLs dei file Excel
    corrispettivi_url = "https://customer-assets.emergentagent.com/job_d33da9de-0c8a-4ce3-90d6-dbb93f270aa0/artifacts/x9v56oeo_corrispettivi.xlsx"
    pos_url = "https://customer-assets.emergentagent.com/job_d33da9de-0c8a-4ce3-90d6-dbb93f270aa0/artifacts/1hu7vd4q_pos.xlsx"
    versamento_url = "https://customer-assets.emergentagent.com/job_d33da9de-0c8a-4ce3-90d6-dbb93f270aa0/artifacts/cegmb0pu_versamento.xlsx"
    finanziamento_url = "https://customer-assets.emergentagent.com/job_d33da9de-0c8a-4ce3-90d6-dbb93f270aa0/artifacts/jk25dlqh_finanziamento%20soci.xlsx"
    
    # Raccolta movimenti
    all_cassa = []
    all_banca = []
    
    # 1. Corrispettivi -> CASSA DARE (entrata)
    corrisp = parse_corrispettivi(corrispettivi_url)
    all_cassa.extend(corrisp)
    
    # 2. POS -> CASSA AVERE (uscita), BANCA DARE (entrata)
    pos_cassa, pos_banca = parse_pos(pos_url)
    all_cassa.extend(pos_cassa)
    all_banca.extend(pos_banca)
    
    # 3. Versamenti -> CASSA AVERE (uscita), BANCA DARE (entrata)
    vers_cassa, vers_banca = parse_versamenti(versamento_url)
    all_cassa.extend(vers_cassa)
    all_banca.extend(vers_banca)
    
    # 4. Finanziamento soci -> CASSA DARE (entrata)
    fin = parse_finanziamento_soci(finanziamento_url)
    all_cassa.extend(fin)
    
    # Calcola totali per tipo
    entrate_cassa = sum(m['importo'] for m in all_cassa if m['tipo'] == 'entrata')
    uscite_cassa = sum(m['importo'] for m in all_cassa if m['tipo'] == 'uscita')
    entrate_banca = sum(m['importo'] for m in all_banca if m['tipo'] == 'entrata')
    uscite_banca = sum(m['importo'] for m in all_banca if m['tipo'] == 'uscita')
    
    print(f"\n=== RIEPILOGO ===")
    print(f"Totale movimenti CASSA: {len(all_cassa)}")
    print(f"  - Entrate (DARE): €{entrate_cassa:,.2f}")
    print(f"  - Uscite (AVERE): €{uscite_cassa:,.2f}")
    print(f"  - Saldo: €{entrate_cassa - uscite_cassa:,.2f}")
    print(f"\nTotale movimenti BANCA: {len(all_banca)}")
    print(f"  - Entrate (DARE): €{entrate_banca:,.2f}")
    print(f"  - Uscite (AVERE): €{uscite_banca:,.2f}")
    print(f"  - Saldo: €{entrate_banca - uscite_banca:,.2f}")
    
    # Import via API
    print("\nImportazione in corso...")
    result = import_to_api(all_cassa, all_banca)
    print(f"Risultato: {result}")

if __name__ == "__main__":
    main()
